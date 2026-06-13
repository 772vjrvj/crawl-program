import os
import pandas as pd
from openpyxl import load_workbook


def read_csv_with_encoding(csv_path: str) -> pd.DataFrame:
    encodings = ["utf-8-sig", "utf-8", "cp949", "euc-kr"]

    last_error = None

    for encoding in encodings:
        try:
            print(f"[INFO] CSV 읽기 시도: encoding={encoding}")

            return pd.read_csv(
                csv_path,
                encoding=encoding,
                dtype=str,              # 모든 컬럼 문자열로 읽기
                keep_default_na=False    # 빈 값이 NaN으로 바뀌지 않게
            )

        except UnicodeDecodeError as e:
            last_error = e
            print(f"[WARN] 인코딩 실패: {encoding}")

        except Exception as e:
            last_error = e
            print(f"[WARN] CSV 읽기 실패: {encoding} / {e}")

    raise RuntimeError(f"CSV 파일을 읽지 못했습니다. 마지막 오류: {last_error}")


def set_text_format_for_phone_columns(xlsx_path: str):
    wb = load_workbook(xlsx_path)
    ws = wb.active

    phone_keywords = [
        "연락처",
        "전화",
        "전화번호",
        "휴대폰",
        "휴대폰번호",
        "핸드폰",
        "핸드폰번호",
        "phone",
        "tel",
        "mobile"
    ]

    for col_idx, cell in enumerate(ws[1], start=1):
        header = str(cell.value).strip() if cell.value else ""

        if any(keyword.lower() in header.lower() for keyword in phone_keywords):
            col_letter = cell.column_letter

            for row in range(2, ws.max_row + 1):
                ws[f"{col_letter}{row}"].number_format = "@"

                if ws[f"{col_letter}{row}"].value is not None:
                    ws[f"{col_letter}{row}"].value = str(ws[f"{col_letter}{row}"].value)

            print(f"[INFO] 텍스트 형식 적용 컬럼: {header}")

    wb.save(xlsx_path)


def convert_csv_to_xlsx():
    base_dir = os.getcwd()

    input_csv = os.path.join(base_dir, "yanolja_local_accommodation_with_seller.csv")
    output_xlsx = os.path.join(base_dir, "yanolja_local_accommodation_with_seller.xlsx")

    if not os.path.exists(input_csv):
        raise FileNotFoundError(f"CSV 파일이 없습니다: {input_csv}")

    df = read_csv_with_encoding(input_csv)

    df.to_excel(output_xlsx, index=False, engine="openpyxl")

    set_text_format_for_phone_columns(output_xlsx)

    print("[DONE] 변환 완료")
    print(f"[INPUT]  {input_csv}")
    print(f"[OUTPUT] {output_xlsx}")
    print(f"[ROWS]   {len(df)}")


if __name__ == "__main__":
    convert_csv_to_xlsx()