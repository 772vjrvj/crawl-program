from __future__ import annotations

import argparse
import csv
import json
import os
import random
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

import requests
from bs4 import BeautifulSoup, Tag
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry


# ============================================================
# 파일 설정
# ============================================================
NORMAL_NO_FILE = Path("resume_no_progress.json")

# 이전에 만든 영문 파일명과 사용자가 말한 한글 파일명을 모두 지원합니다.
PREFERRED_NO_FILE_CANDIDATES = [
    Path("preferred_resume_no_progress.json"),
    Path("우대.json"),
]

# 상세정보 수집 중간 저장 파일
DETAIL_CHECKPOINT_CSV = Path("resume_detail_checkpoint.csv")
DETAIL_FAILED_CSV = Path("resume_detail_failed.csv")
DEBUG_DIR = Path("debug_resume_detail")

# 최종 엑셀 2개
XLSX_CERTIFICATES_ONE_CELL = Path("resume_detail_certificates_one_cell.xlsx")
XLSX_CERTIFICATES_ROWS = Path("resume_detail_certificates_rows.xlsx")


# ============================================================
# 사이트 / 요청 설정
# ============================================================
BASE_URL = "https://www.dorus.co.kr"
DETAIL_URL = f"{BASE_URL}/work/resume_detail.html"
REFERER_URL = f"{BASE_URL}/work/resume_list.html?page=1"

# 환경변수가 있으면 환경변수를 우선 사용합니다.
# 직접 입력하려면 아래 문자열을 현재 PHPSESSID로 교체하세요.
PHPSESSID = os.getenv(
    "DORUS_PHPSESSID",
    "9cf6262413847cb4882aa20a672035b6; _fwb=30zJXu4WKYnf47JJJHKi4r.1785756039940; _gid=GA1.3.982365059.1785756040; uwrToday=46678,38,37,46768,46903,46892,46846,46437,46910; wcs_bt=undefined:1785759939|undefined:1785756040; _gat_gtag_UA_86116564_1=1; _ga_F71F17H2DY=GS2.1.s1785759939$o2$g1$t1785759939$j60$l0$h0; _ga=GA1.1.1992543861.1785756040",
)

REQUEST_TIMEOUT_SEC = 25
MIN_DELAY_SEC = 0.8
MAX_DELAY_SEC = 1.4


CSV_HEADERS = [
    "이력서번호",
    "우대여부",
    "지역",
    "직종",
    "성명(성별,나이)",
    "핸드폰",
    "전화번호",
    "이메일",
    "주소",
    "보유자격증JSON",
    "최종수정일",
    "상세URL",
]

FAILED_HEADERS = [
    "처리일시",
    "이력서번호",
    "오류유형",
    "오류내용",
    "상세URL",
]


class DetailParseError(RuntimeError):
    """상세 페이지 HTML 구조에서 필수 정보를 찾지 못한 경우."""


class SessionAccessError(RuntimeError):
    """세션 만료 또는 접근 권한 문제로 판단되는 경우."""


# ============================================================
# 공통 유틸
# ============================================================
def normalize_text(value: str | None) -> str:
    """HTML에서 추출한 공백과 줄바꿈을 한 칸으로 정리합니다."""
    if not value:
        return ""

    value = value.replace("\xa0", " ")
    return re.sub(r"\s+", " ", value).strip()


def clean_tag_text(tag: Tag | None) -> str:
    if tag is None:
        return ""
    return normalize_text(tag.get_text(" ", strip=True))


def resolve_preferred_no_file() -> Path:
    for candidate in PREFERRED_NO_FILE_CANDIDATES:
        if candidate.exists():
            return candidate

    expected = ", ".join(str(path) for path in PREFERRED_NO_FILE_CANDIDATES)
    raise FileNotFoundError(
        f"우대 이력서 번호 파일이 없습니다. 다음 중 하나가 필요합니다: {expected}"
    )


def load_resume_nos(json_path: Path) -> list[int]:
    """JSON의 resume_nos 배열을 숫자 목록으로 읽고 내부 중복을 제거합니다."""
    if not json_path.exists():
        raise FileNotFoundError(f"번호 파일을 찾을 수 없습니다: {json_path}")

    with json_path.open("r", encoding="utf-8-sig") as file:
        data = json.load(file)

    raw_nos = data.get("resume_nos")
    if not isinstance(raw_nos, list):
        raise ValueError(f"resume_nos 배열이 없습니다: {json_path}")

    result: list[int] = []
    seen: set[int] = set()

    for raw_no in raw_nos:
        try:
            resume_no = int(raw_no)
        except (TypeError, ValueError):
            continue

        if resume_no <= 0 or resume_no in seen:
            continue

        seen.add(resume_no)
        result.append(resume_no)

    return result


def merge_resume_nos(
        normal_nos: Iterable[int],
        preferred_nos: Iterable[int],
) -> list[int]:
    """일반 번호 뒤에 우대 번호를 합치면서 순서를 유지하고 중복 제거합니다."""
    return list(dict.fromkeys([*normal_nos, *preferred_nos]))


def decode_response_html(response: requests.Response) -> str:
    """사이트의 CP949/EUC-KR 계열 응답을 한글로 디코딩합니다."""
    return response.content.decode("cp949", errors="replace")


# ============================================================
# HTTP 세션
# ============================================================
def create_session(php_session_id: str) -> requests.Session:
    if not php_session_id or php_session_id == "여기에_현재_PHPSESSID":
        raise ValueError(
            "PHPSESSID가 설정되지 않았습니다. 코드의 PHPSESSID를 교체하거나 "
            "환경변수 DORUS_PHPSESSID를 설정하세요."
        )

    session = requests.Session()

    retry = Retry(
        total=5,
        connect=5,
        read=5,
        status=5,
        backoff_factor=1.5,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=frozenset(["GET"]),
        respect_retry_after_header=True,
        raise_on_status=False,
    )

    adapter = HTTPAdapter(
        max_retries=retry,
        pool_connections=5,
        pool_maxsize=5,
    )

    session.mount("https://", adapter)
    session.mount("http://", adapter)

    session.headers.update(
        {
            "Accept": (
                "text/html,application/xhtml+xml,application/xml;q=0.9,"
                "image/avif,image/webp,image/apng,*/*;q=0.8"
            ),
            "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
            "Cache-Control": "no-cache",
            "Pragma": "no-cache",
            "Referer": REFERER_URL,
            "Upgrade-Insecure-Requests": "1",
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/151.0.0.0 Safari/537.36"
            ),
        }
    )

    # 분석/광고 쿠키는 필요 없고 로그인 세션 쿠키만 사용합니다.
    session.cookies.set(
        "PHPSESSID",
        php_session_id,
        domain="www.dorus.co.kr",
        path="/",
    )

    return session


def request_detail_html(
        session: requests.Session,
        resume_no: int,
) -> str:
    response = session.get(
        DETAIL_URL,
        params={"no": resume_no},
        timeout=REQUEST_TIMEOUT_SEC,
    )

    if response.status_code != 200:
        raise requests.HTTPError(
            f"HTTP {response.status_code}",
            response=response,
        )

    return decode_response_html(response)


# ============================================================
# 상세 HTML 파싱
# ============================================================
def find_exact_label_node(
        soup: BeautifulSoup,
        labels: Iterable[str],
) -> Tag | None:
    normalized_labels = {
        normalize_text(label)
        for label in labels
    }

    # 실제 HTML은 <td><b>성명</b></td> 형태이므로 b/strong을 먼저 확인합니다.
    for tag_name in ("b", "strong"):
        for node in soup.find_all(tag_name):
            if clean_tag_text(node) in normalized_labels:
                return node

    # 일부 페이지의 구조가 다를 경우 td 자체도 확인합니다.
    for node in soup.find_all("td"):
        if clean_tag_text(node) in normalized_labels:
            return node

    return None


def extract_labeled_value(
        soup: BeautifulSoup,
        labels: str | Iterable[str],
) -> str:
    """
    라벨이 포함된 tr에서 라벨 td 오른쪽의 첫 번째 실제 문자열 값을 반환합니다.

    예:
        성명 | 구분 이미지 td | 조미애 (여,1982년생)
    """
    if isinstance(labels, str):
        label_list = [labels]
    else:
        label_list = list(labels)

    label_node = find_exact_label_node(soup, label_list)
    if label_node is None:
        return ""

    label_td = label_node if label_node.name == "td" else label_node.find_parent("td")
    row = label_node.find_parent("tr")

    if label_td is None or row is None:
        return ""

    direct_tds = row.find_all("td", recursive=False)
    if not direct_tds:
        direct_tds = row.find_all("td")

    try:
        label_index = direct_tds.index(label_td)
    except ValueError:
        return ""

    label_texts = {
        normalize_text(label)
        for label in label_list
    }

    # 라벨 오른쪽에는 세로선 이미지만 있는 td가 있으므로 텍스트가 있는 td까지 이동합니다.
    for value_td in direct_tds[label_index + 1:]:
        value = clean_tag_text(value_td)
        if value and value not in label_texts:
            return value

    return ""


def extract_certificates(soup: BeautifulSoup) -> list[dict[str, str]]:
    """
    class="tb_tb2 link_line table_line" 테이블 중
    취득일 / 자격증명 / 발행기관 헤더를 모두 가진 테이블만 선택합니다.
    """
    required_headers = ["취득일", "자격증명", "발행기관"]

    for table in soup.find_all("table"):
        classes = set(table.get("class", []))

        if not {"tb_tb2", "link_line", "table_line"}.issubset(classes):
            continue

        rows = table.find_all("tr")
        if not rows:
            continue

        header_cells = rows[0].find_all(["td", "th"], recursive=False)
        if not header_cells:
            header_cells = rows[0].find_all(["td", "th"])

        header_values = [clean_tag_text(cell) for cell in header_cells]

        if not all(required in header_values for required in required_headers):
            continue

        header_index = {
            header: header_values.index(header)
            for header in required_headers
        }

        certificates: list[dict[str, str]] = []

        for row in rows[1:]:
            cells = row.find_all("td", recursive=False)
            if not cells:
                cells = row.find_all("td")

            if len(cells) <= max(header_index.values()):
                continue

            acquired_date = clean_tag_text(cells[header_index["취득일"]])
            certificate_name = clean_tag_text(cells[header_index["자격증명"]])
            issuer = clean_tag_text(cells[header_index["발행기관"]])

            if not any([acquired_date, certificate_name, issuer]):
                continue

            certificates.append(
                {
                    "취득일": acquired_date,
                    "자격증명": certificate_name,
                    "발행기관": issuer,
                }
            )

        return certificates

    return []


def extract_last_modified(soup: BeautifulSoup) -> str:
    """
    "최종수정일 : yyyy년 MM월 dd일" 문구에서 날짜만 추출합니다.

    바깥쪽 td의 전체 하위 텍스트를 합치지 않고,
    실제로 "최종수정일"이 포함된 문자열 노드만 검사합니다.
    """
    pattern = re.compile(
        r"최종수정일\s*[:：]\s*"
        r"(\d{4}년\s*\d{1,2}월\s*\d{1,2}일)"
    )

    for text_node in soup.find_all(string=pattern):
        parent = text_node.parent

        # 예: <td><img ...> 최종수정일 : 2026년 08월 03일</td>
        if parent is None or parent.name != "td":
            continue

        match = pattern.search(
            normalize_text(str(text_node))
        )

        if match:
            return normalize_text(match.group(1))

    return ""


def looks_like_access_problem(html: str) -> bool:
    """성명 라벨이 없을 때 세션/접근권한 문제 가능성을 판별합니다."""
    lower_html = html.lower()

    access_markers = [
        "로그인 후",
        "회원만",
        "접근 권한",
        "권한이 없습니다",
        "이용권",
        "결제 후",
    ]

    return (
            any(marker in html for marker in access_markers)
            or "name=\"login" in lower_html
            or "login.html" in lower_html
    )


def parse_resume_detail(
        html: str,
        resume_no: int,
        preferred_yn: str,
) -> dict[str, str]:
    soup = BeautifulSoup(html, "html.parser")

    full_name = extract_labeled_value(soup, "성명")

    if not full_name:
        if looks_like_access_problem(html):
            raise SessionAccessError(
                "성명 라벨을 찾지 못했고 로그인/접근 제한 페이지로 판단됩니다. "
                "PHPSESSID를 새 값으로 교체하세요."
            )

        raise DetailParseError("성명 라벨 또는 상세정보 표를 찾지 못했습니다.")

    certificates = extract_certificates(soup)

    return {
        "이력서번호": str(resume_no),
        "우대여부": preferred_yn,
        # 사이트에서 라벨이 지역/직종 또는 희망근무지역/희망직종으로 표시되는 경우를 모두 지원
        "지역": extract_labeled_value(soup, ["지역", "희망근무지역"]),
        "직종": extract_labeled_value(soup, ["직종", "희망직종"]),
        "성명(성별,나이)": full_name,
        "핸드폰": extract_labeled_value(soup, "핸드폰"),
        "전화번호": extract_labeled_value(soup, "전화번호"),
        "이메일": extract_labeled_value(soup, "이메일"),
        "주소": extract_labeled_value(soup, "주소"),
        "보유자격증JSON": json.dumps(
            certificates,
            ensure_ascii=False,
            separators=(",", ":"),
        ),
        "최종수정일": extract_last_modified(soup),
        "상세URL": f"{DETAIL_URL}?no={resume_no}",
    }


# ============================================================
# CSV 체크포인트
# ============================================================
def load_completed_resume_nos(csv_path: Path) -> set[int]:
    if not csv_path.exists() or csv_path.stat().st_size == 0:
        return set()

    completed: set[int] = set()

    with csv_path.open(
            "r",
            encoding="utf-8-sig",
            newline="",
    ) as file:
        reader = csv.DictReader(file)

        for row in reader:
            try:
                completed.add(int(row.get("이력서번호", "")))
            except (TypeError, ValueError):
                continue

    return completed


def open_checkpoint_writer() -> tuple[Any, csv.DictWriter]:
    is_new_file = (
            not DETAIL_CHECKPOINT_CSV.exists()
            or DETAIL_CHECKPOINT_CSV.stat().st_size == 0
    )

    file = DETAIL_CHECKPOINT_CSV.open(
        "a",
        encoding="utf-8-sig",
        newline="",
    )

    writer = csv.DictWriter(
        file,
        fieldnames=CSV_HEADERS,
        extrasaction="ignore",
    )

    if is_new_file:
        writer.writeheader()
        file.flush()
        os.fsync(file.fileno())

    return file, writer


def append_checkpoint_row(
        file: Any,
        writer: csv.DictWriter,
        row: dict[str, str],
) -> None:
    writer.writerow(row)
    file.flush()

    # 프로그램이 갑자기 종료되어도 현재 행이 디스크에 남을 가능성을 높입니다.
    os.fsync(file.fileno())


def append_failed_row(
        resume_no: int,
        error: Exception,
) -> None:
    is_new_file = (
            not DETAIL_FAILED_CSV.exists()
            or DETAIL_FAILED_CSV.stat().st_size == 0
    )

    with DETAIL_FAILED_CSV.open(
            "a",
            encoding="utf-8-sig",
            newline="",
    ) as file:
        writer = csv.DictWriter(
            file,
            fieldnames=FAILED_HEADERS,
            extrasaction="ignore",
        )

        if is_new_file:
            writer.writeheader()

        writer.writerow(
            {
                "처리일시": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "이력서번호": resume_no,
                "오류유형": type(error).__name__,
                "오류내용": normalize_text(str(error)),
                "상세URL": f"{DETAIL_URL}?no={resume_no}",
            }
        )


def save_debug_html(
        resume_no: int,
        html: str,
) -> Path:
    DEBUG_DIR.mkdir(parents=True, exist_ok=True)
    debug_path = DEBUG_DIR / f"resume_detail_{resume_no}.html"

    # 브라우저에서 한글이 정상적으로 보이도록 UTF-8로 저장합니다.
    debug_path.write_text(html, encoding="utf-8")
    return debug_path


def load_checkpoint_rows() -> list[dict[str, str]]:
    if not DETAIL_CHECKPOINT_CSV.exists():
        return []

    rows_by_no: dict[int, dict[str, str]] = {}

    with DETAIL_CHECKPOINT_CSV.open(
            "r",
            encoding="utf-8-sig",
            newline="",
    ) as file:
        reader = csv.DictReader(file)

        for row in reader:
            try:
                resume_no = int(row.get("이력서번호", ""))
            except (TypeError, ValueError):
                continue

            # 혹시 중복 행이 생겼다면 마지막 정상 행을 사용합니다.
            rows_by_no[resume_no] = {
                header: row.get(header, "")
                for header in CSV_HEADERS
            }

    return list(rows_by_no.values())


# ============================================================
# XLSX 생성
# ============================================================
def parse_certificates_json(value: str) -> list[dict[str, str]]:
    if not value:
        return []

    try:
        parsed = json.loads(value)
    except json.JSONDecodeError:
        return []

    if not isinstance(parsed, list):
        return []

    result: list[dict[str, str]] = []

    for item in parsed:
        if not isinstance(item, dict):
            continue

        result.append(
            {
                "취득일": normalize_text(str(item.get("취득일", ""))),
                "자격증명": normalize_text(str(item.get("자격증명", ""))),
                "발행기관": normalize_text(str(item.get("발행기관", ""))),
            }
        )

    return result


def apply_excel_style(
        worksheet: Any,
        column_widths: dict[str, float],
) -> None:
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )
    header_font = Font(
        name="맑은 고딕",
        size=10,
        bold=True,
        color="FFFFFF",
    )
    body_font = Font(
        name="맑은 고딕",
        size=10,
    )
    thin_gray = Side(
        style="thin",
        color="D9E2F3",
    )
    border = Border(
        left=thin_gray,
        right=thin_gray,
        top=thin_gray,
        bottom=thin_gray,
    )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.sheet_view.showGridLines = False
    worksheet.row_dimensions[1].height = 26

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = border

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )
            cell.border = border

    for column_name, width in column_widths.items():
        worksheet.column_dimensions[column_name].width = width

    # 전화번호, 핸드폰, 이메일은 문자열 형식으로 고정합니다.
    header_to_column = {
        worksheet.cell(row=1, column=column).value: column
        for column in range(1, worksheet.max_column + 1)
    }

    for header in ["핸드폰", "전화번호", "이메일"]:
        column_index = header_to_column.get(header)
        if column_index is None:
            continue

        for row_index in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row_index, column=column_index).number_format = "@"

    # 상세URL을 클릭 가능한 링크로 설정합니다.
    url_column = header_to_column.get("상세URL")
    if url_column is not None:
        for row_index in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_index, column=url_column)
            if cell.value:
                cell.hyperlink = str(cell.value)
                cell.style = "Hyperlink"


def save_workbook_atomically(
        workbook: Workbook,
        output_path: Path,
) -> None:
    temp_path = output_path.with_name(
        f"{output_path.stem}.tmp.xlsx"
    )

    workbook.save(temp_path)
    temp_path.replace(output_path)


def build_certificates_one_cell_xlsx(
        checkpoint_rows: list[dict[str, str]],
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "이력서 상세"

    headers = [
        "이력서번호",
        "우대여부",
        "지역",
        "직종",
        "성명(성별,나이)",
        "핸드폰",
        "전화번호",
        "이메일",
        "주소",
        "보유자격증",
        "최종수정일",
        "상세URL",
    ]

    worksheet.append(headers)

    for source_row in checkpoint_rows:
        worksheet.append(
            [
                source_row.get("이력서번호", ""),
                source_row.get("우대여부", ""),
                source_row.get("지역", ""),
                source_row.get("직종", ""),
                source_row.get("성명(성별,나이)", ""),
                source_row.get("핸드폰", ""),
                source_row.get("전화번호", ""),
                source_row.get("이메일", ""),
                source_row.get("주소", ""),
                source_row.get("보유자격증JSON", "[]"),
                source_row.get("최종수정일", ""),
                source_row.get("상세URL", ""),
            ]
        )

    widths = {
        "A": 12,
        "B": 10,
        "C": 34,
        "D": 24,
        "E": 24,
        "F": 18,
        "G": 18,
        "H": 32,
        "I": 48,
        "J": 65,
        "K": 20,
        "L": 48,
    }

    apply_excel_style(worksheet, widths)
    save_workbook_atomically(workbook, XLSX_CERTIFICATES_ONE_CELL)


def build_certificates_rows_xlsx(
        checkpoint_rows: list[dict[str, str]],
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "이력서 자격증 행분리"

    headers = [
        "이력서번호",
        "우대여부",
        "지역",
        "직종",
        "성명(성별,나이)",
        "핸드폰",
        "전화번호",
        "이메일",
        "주소",
        "자격증순번",
        "취득일",
        "자격증명",
        "발행기관",
        "최종수정일",
        "상세URL",
    ]

    worksheet.append(headers)

    for source_row in checkpoint_rows:
        certificates = parse_certificates_json(
            source_row.get("보유자격증JSON", "")
        )

        # 자격증이 없어도 이력서 자체는 한 행으로 남깁니다.
        if not certificates:
            certificates = [
                {
                    "취득일": "",
                    "자격증명": "",
                    "발행기관": "",
                }
            ]

        for sequence, certificate in enumerate(certificates, start=1):
            worksheet.append(
                [
                    source_row.get("이력서번호", ""),
                    source_row.get("우대여부", ""),
                    source_row.get("지역", ""),
                    source_row.get("직종", ""),
                    source_row.get("성명(성별,나이)", ""),
                    source_row.get("핸드폰", ""),
                    source_row.get("전화번호", ""),
                    source_row.get("이메일", ""),
                    source_row.get("주소", ""),
                    sequence if certificate.get("자격증명") else "",
                    certificate.get("취득일", ""),
                    certificate.get("자격증명", ""),
                    certificate.get("발행기관", ""),
                    source_row.get("최종수정일", ""),
                    source_row.get("상세URL", ""),
                ]
            )

    widths = {
        "A": 12,
        "B": 10,
        "C": 34,
        "D": 24,
        "E": 24,
        "F": 18,
        "G": 18,
        "H": 32,
        "I": 48,
        "J": 12,
        "K": 16,
        "L": 32,
        "M": 28,
        "N": 20,
        "O": 48,
    }

    apply_excel_style(worksheet, widths)
    save_workbook_atomically(workbook, XLSX_CERTIFICATES_ROWS)


def build_xlsx_files() -> None:
    checkpoint_rows = load_checkpoint_rows()

    if not checkpoint_rows:
        print("[XLSX 건너뜀] CSV에 변환할 정상 데이터가 없습니다.")
        return

    build_certificates_one_cell_xlsx(checkpoint_rows)
    build_certificates_rows_xlsx(checkpoint_rows)

    print(f"[XLSX 완료] {XLSX_CERTIFICATES_ONE_CELL.resolve()}")
    print(f"[XLSX 완료] {XLSX_CERTIFICATES_ROWS.resolve()}")


# ============================================================
# 메인 수집
# ============================================================
def crawl_details(
        limit: int | None = None,
        no_sleep: bool = False,
) -> None:
    preferred_no_file = resolve_preferred_no_file()

    normal_nos = load_resume_nos(NORMAL_NO_FILE)
    preferred_nos = load_resume_nos(preferred_no_file)

    all_nos = merge_resume_nos(normal_nos, preferred_nos)
    preferred_no_set = set(preferred_nos)
    completed_nos = load_completed_resume_nos(DETAIL_CHECKPOINT_CSV)

    pending_nos = [
        resume_no
        for resume_no in all_nos
        if resume_no not in completed_nos
    ]

    if limit is not None:
        pending_nos = pending_nos[:limit]

    print(f"[일반 번호] {len(normal_nos):,}개")
    print(f"[우대 번호] {len(preferred_nos):,}개")
    print(f"[합계/중복제거] {len(all_nos):,}개")
    print(f"[기존 CSV 완료] {len(completed_nos):,}개")
    print(f"[이번 실행 대상] {len(pending_nos):,}개")
    print(f"[우대 번호 파일] {preferred_no_file}")

    if not pending_nos:
        print("[수집 완료] 새로 처리할 번호가 없습니다.")
        build_xlsx_files()
        return

    session = create_session(PHPSESSID)
    checkpoint_file, checkpoint_writer = open_checkpoint_writer()

    try:
        for index, resume_no in enumerate(pending_nos, start=1):
            preferred_yn = "Y" if resume_no in preferred_no_set else "N"

            try:
                html = request_detail_html(session, resume_no)
                row = parse_resume_detail(
                    html=html,
                    resume_no=resume_no,
                    preferred_yn=preferred_yn,
                )

                append_checkpoint_row(
                    file=checkpoint_file,
                    writer=checkpoint_writer,
                    row=row,
                )

                certificate_count = len(
                    parse_certificates_json(
                        row.get("보유자격증JSON", "")
                    )
                )

                print(
                    f"[{index:05d}/{len(pending_nos):05d}] "
                    f"no={resume_no} 성공 "
                    f"| 우대={preferred_yn} "
                    f"| 자격증={certificate_count}개 "
                    f"| 성명={row.get('성명(성별,나이)', '')}"
                )

            except SessionAccessError as error:
                debug_path = save_debug_html(resume_no, html)
                append_failed_row(resume_no, error)

                print(
                    f"[세션/권한 오류] no={resume_no} | {error}\n"
                    f"[HTML 저장] {debug_path.resolve()}"
                )
                print("[중단] PHPSESSID를 교체한 뒤 다시 실행하세요.")
                break

            except DetailParseError as error:
                debug_path = save_debug_html(resume_no, html)
                append_failed_row(resume_no, error)

                print(
                    f"[{index:05d}/{len(pending_nos):05d}] "
                    f"no={resume_no} 파싱 실패 | {error} "
                    f"| HTML={debug_path}"
                )

            except requests.RequestException as error:
                append_failed_row(resume_no, error)

                print(
                    f"[{index:05d}/{len(pending_nos):05d}] "
                    f"no={resume_no} 요청 실패 | {error}"
                )

            except Exception as error:
                append_failed_row(resume_no, error)

                print(
                    f"[{index:05d}/{len(pending_nos):05d}] "
                    f"no={resume_no} 처리 실패 "
                    f"| {type(error).__name__}: {error}"
                )

            if not no_sleep and index < len(pending_nos):
                time.sleep(
                    random.uniform(
                        MIN_DELAY_SEC,
                        MAX_DELAY_SEC,
                    )
                )

    except KeyboardInterrupt:
        print("\n[사용자 중단] 현재까지 성공한 행은 CSV에 저장되어 있습니다.")

    finally:
        checkpoint_file.close()
        session.close()

        # 정상 종료/사용자 중단 모두 현재 CSV 기준으로 엑셀 2개를 갱신합니다.
        build_xlsx_files()


# ============================================================
# 실행 옵션
# ============================================================
def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "두 이력서 번호 JSON을 합쳐 상세정보를 CSV 체크포인트로 수집하고 "
            "자격증 1셀형/행분리형 XLSX 두 개를 생성합니다."
        )
    )

    parser.add_argument(
        "--xlsx-only",
        action="store_true",
        help="사이트 요청 없이 기존 CSV만 읽어서 XLSX 두 개를 다시 생성합니다.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="테스트를 위해 이번 실행에서 처리할 신규 번호 개수를 제한합니다.",
    )
    parser.add_argument(
        "--no-sleep",
        action="store_true",
        help="페이지 사이 대기를 제거합니다. 차단 위험이 있으므로 테스트 외에는 권장하지 않습니다.",
    )

    return parser.parse_args()


def main() -> None:
    args = parse_arguments()

    if args.xlsx_only:
        build_xlsx_files()
        return

    crawl_details(
        limit=args.limit,
        no_sleep=args.no_sleep,
    )


if __name__ == "__main__":
    try:
        main()
    except Exception as error:
        print(
            f"[실행 오류] {type(error).__name__}: {error}",
            file=sys.stderr,
        )
        sys.exit(1)