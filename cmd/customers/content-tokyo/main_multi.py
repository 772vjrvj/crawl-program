from __future__ import annotations

import csv
import json
import re
import sys
import threading
import time
import uuid
from concurrent.futures import Future, ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry


# =========================================================
# 입력/출력 파일
# =========================================================
INPUT_HITS_JSON_NAME = "content_tokyo_exhibitors_hits.json"

OUTPUT_DETAIL_JSON_NAME = "content_tokyo_exhibitors_details.json"
OUTPUT_CSV_NAME = "content_tokyo_exhibitors_with_address.csv"
OUTPUT_XLSX_NAME = "content_tokyo_exhibitors_with_address.xlsx"


# =========================================================
# 상세 API 설정
# =========================================================
GRAPHQL_URL = "https://api.reedexpo.com/graphql/"

X_CLIENT_ID = "uhQVcmxLwXAjVtVpTvoerERiZSsNz0om"

DEFAULT_EVENT_EDITION_ID = (
    "eve-2b972fc3-9919-4475-983b-a08253dfd7d1"
)

TARGET_LOCALE = "ja-jp"


# =========================================================
# 멀티스레드 설정
# =========================================================

# 동시에 요청할 개수
# 처음에는 5로 실행하는 것을 권장
MAX_WORKERS = 8

# 각 스레드가 요청 완료 후 대기하는 시간
# API 부하 및 429 오류 방지용
REQUEST_DELAY_SECONDS = 0.2

# 상세 JSON 중간 저장 주기
CHECKPOINT_INTERVAL = 20


# =========================================================
# CSV/XLSX 컬럼
# =========================================================
HEADERS = [
    "회사명",
    "포인트",
    "사업내용",
    "홈페이지",
    "이메일",
    "연락처",
    "addressLine1",
    "addressLine2",
    "stateProvince",
    "city",
    "countryCode",
    "postcode",
    "country",
    "additionalAddresses",
    "카테고리",
]


# =========================================================
# GraphQL Query
# =========================================================
GRAPHQL_QUERY_TEMPLATE = """
{
  exhibitingOrganisation(
    eventEditionId: __EVENT_EDITION_ID__,
    organisationId: __ORGANISATION_ID__
  ) {
    id
    productsAndServices {
      id
    }
    socialMedia {
      url
      name
    }
    isNew
    organisationId
    accompanyingWebsiteUrl
    packageId
    companyName
    contactEmail
    website
    phone
    hideAddress
    extraCharacteristics {
      id
    }
    multilingual {
      logoUrl
      coverImageUrl
      locale
      displayName
      description
      showObjective
      sortAlias
      representedBrands {
        name
      }
      addressLine1
      addressLine2
      stateProvince
      city
      countryCode
      postcode
      country
      additionalAddresses {
        addressLine1
        addressLine2
        stateProvince
        city
        countryCode
        postcode
      }
    }
    exhibitingProducers {
      id
      firstName
      lastName
    }
    representedProducers {
      id
      firstName
      lastName
    }
    directors
    filterCategories {
      multilingual {
        name
        locale
      }
      responses {
        answerId
        parentId
        taxonomyId
        multilingual {
          name
          locale
        }
      }
      id
    }
    products(includeUnpublished: false) {
      id
      exhibitingOrganisationId
      imageUrl
      isInnovative
      isNew
      videoId
      video {
        id
        defaultThumbnailUrl
        customThumbnailUrl
        status
      }
      isPublished
      lastUpdatedAt
      producerDetails {
        id
        firstName
        lastName
      }
      multilingual {
        title
        description
        locale
      }
      filterAttributes {
        id
        questionText {
          locale
          name
        }
        textualResponses {
          locale
          textResponse
        }
        responses {
          answerId
          answerText {
            locale
            name
          }
        }
        sortOrder
      }
      attributes {
        id
        questionText {
          locale
          name
        }
        responses {
          answerId
          answerText {
            locale
            name
          }
        }
        textualResponses {
          locale
          textResponse
        }
        sortOrder
      }
    }
    stands {
      name
    }
    sharers {
      id
      organisationGuid
      name
      isActive
      packageId
      multilingual {
        name
        locale
      }
    }
    mainStandHolder {
      organisationGuid
      id
      packageId
      multilingual {
        name
        locale
      }
      sharers {
        id
        organisationGuid
        name
        isActive
        packageId
        multilingual {
          name
          locale
        }
      }
    }
  }
}
"""


# 스레드마다 개별 Session을 보관
_thread_local = threading.local()


def get_execution_directory() -> Path:
    """
    Python 파일 또는 exe 파일이 있는 실행 경로를 반환한다.
    """
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent

    return Path(__file__).resolve().parent


def clean_text(value: Any) -> str:
    """
    값을 CSV/XLSX에 안전하게 저장할 문자열로 변환한다.

    list와 dict는 JSON 문자열로 변환한다.
    """
    if value is None:
        return ""

    if isinstance(value, (list, dict)):
        value = json.dumps(
            value,
            ensure_ascii=False,
            separators=(",", ":"),
        )

    text = str(value)

    # Excel에서 허용하지 않는 제어문자 제거
    text = ILLEGAL_CHARACTERS_RE.sub("", text)

    text = text.replace("\r\n", "\n")
    text = text.replace("\r", "\n")

    return text.strip()


def create_session() -> requests.Session:
    """
    429 및 일시적인 서버 오류 발생 시 자동 재시도하는 세션을 생성한다.
    """
    retry = Retry(
        total=5,
        connect=5,
        read=5,
        status=5,
        backoff_factor=1,
        status_forcelist=[
            429,
            500,
            502,
            503,
            504,
        ],
        allowed_methods=frozenset(["POST"]),
        respect_retry_after_header=True,
    )

    adapter = HTTPAdapter(
        max_retries=retry,
        pool_connections=5,
        pool_maxsize=5,
    )

    session = requests.Session()
    session.mount("https://", adapter)

    return session


def get_thread_session() -> requests.Session:
    """
    각 작업 스레드마다 독립적인 Session을 생성한다.

    하나의 requests.Session을 여러 스레드가 동시에 공유하지 않도록 한다.
    """
    session = getattr(
        _thread_local,
        "session",
        None,
    )

    if session is None:
        session = create_session()
        _thread_local.session = session

    return session


def load_hits(
        input_path: Path,
) -> list[dict[str, Any]]:
    """
    content_tokyo_exhibitors_hits.json에서 hits 배열을 읽는다.
    """
    if not input_path.exists():
        raise FileNotFoundError(
            f"입력 JSON 파일을 찾을 수 없습니다: {input_path}"
        )

    with input_path.open(
            mode="r",
            encoding="utf-8-sig",
    ) as file:
        json_data = json.load(file)

    if not isinstance(json_data, dict):
        raise ValueError(
            "입력 JSON의 최상위 데이터가 객체가 아닙니다."
        )

    hits = json_data.get("hits")

    if not isinstance(hits, list):
        raise ValueError(
            "입력 JSON에서 hits 배열을 찾을 수 없습니다."
        )

    valid_hits: list[dict[str, Any]] = []

    for index, item in enumerate(hits):
        if not isinstance(item, dict):
            print(
                f"[경고] hits[{index}]은 객체가 아니므로 제외합니다."
            )
            continue

        valid_hits.append(item)

    return valid_hits


def build_graphql_query(
        event_edition_id: str,
        organisation_id: str,
) -> str:
    """
    업체 상세조회 GraphQL Query를 생성한다.
    """
    event_literal = json.dumps(
        event_edition_id,
        ensure_ascii=False,
    )

    organisation_literal = json.dumps(
        organisation_id,
        ensure_ascii=False,
    )

    query = GRAPHQL_QUERY_TEMPLATE.replace(
        "__EVENT_EDITION_ID__",
        event_literal,
    )

    query = query.replace(
        "__ORGANISATION_ID__",
        organisation_literal,
    )

    return query


def request_detail(
        session: requests.Session,
        event_edition_id: str,
        organisation_id: str,
) -> dict[str, Any]:
    """
    업체 한 건의 상세정보를 GraphQL API로 조회한다.
    """
    headers = {
        "Accept": "application/json",
        "Accept-Language": (
            "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7"
        ),
        "Content-Type": "application/json",
        "Origin": "https://www.content-tokyo.jp",
        "Referer": "https://www.content-tokyo.jp/",
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/149.0.0.0 Safari/537.36"
        ),
        "x-clientid": X_CLIENT_ID,
        "x-correlationid": str(uuid.uuid4()),
    }

    payload = {
        "query": build_graphql_query(
            event_edition_id=event_edition_id,
            organisation_id=organisation_id,
        )
    }

    response = session.post(
        GRAPHQL_URL,
        headers=headers,
        json=payload,
        timeout=(10, 60),
    )

    response.raise_for_status()

    result = response.json()

    if not isinstance(result, dict):
        raise ValueError(
            "GraphQL 응답이 JSON 객체가 아닙니다."
        )

    graphql_errors = result.get("errors")

    if graphql_errors:
        raise ValueError(
            "GraphQL 오류: "
            + json.dumps(
                graphql_errors,
                ensure_ascii=False,
            )
        )

    data = result.get("data")

    if not isinstance(data, dict):
        raise ValueError(
            "GraphQL 응답에 data 객체가 없습니다."
        )

    organisation = data.get(
        "exhibitingOrganisation"
    )

    if organisation is None:
        raise ValueError(
            "상세 응답의 exhibitingOrganisation이 null입니다."
        )

    if not isinstance(organisation, dict):
        raise ValueError(
            "exhibitingOrganisation 형식이 올바르지 않습니다."
        )

    return data


def load_existing_detail_map(
        detail_json_path: Path,
) -> dict[str, dict[str, Any]]:
    """
    기존 상세 JSON 파일이 있으면 읽는다.

    정상 수집된 업체는 다음 실행에서 재요청하지 않는다.
    실패한 업체는 다시 요청한다.
    """
    if not detail_json_path.exists():
        return {}

    try:
        with detail_json_path.open(
                mode="r",
                encoding="utf-8-sig",
        ) as file:
            detail_json = json.load(file)

        if not isinstance(detail_json, dict):
            return {}

        details = detail_json.get("details", [])

        if not isinstance(details, list):
            return {}

        detail_map: dict[str, dict[str, Any]] = {}

        for detail in details:
            if not isinstance(detail, dict):
                continue

            organisation_id = clean_text(
                detail.get("organisationId")
            )

            if not organisation_id:
                continue

            detail_map[organisation_id] = detail

        return detail_map

    except Exception as exception:
        print(
            f"[경고] 기존 상세 JSON을 읽지 못했습니다: "
            f"{exception}"
        )

        return {}


def has_valid_detail(
        detail_entry: dict[str, Any] | None,
) -> bool:
    """
    상세조회가 정상적으로 완료된 데이터인지 확인한다.
    """
    if not isinstance(detail_entry, dict):
        return False

    data = detail_entry.get("data")

    if not isinstance(data, dict):
        return False

    organisation = data.get(
        "exhibitingOrganisation"
    )

    return isinstance(organisation, dict)


def save_detail_json(
        detail_map: dict[str, dict[str, Any]],
        output_path: Path,
) -> None:
    """
    상세 API 응답 data 전체를 JSON 파일로 저장한다.

    임시 파일에 저장한 뒤 원본 파일로 교체하여
    저장 중 파일 손상 가능성을 줄인다.
    """
    details = list(detail_map.values())

    output_data = {
        "count": len(details),
        "details": details,
    }

    temp_path = output_path.with_suffix(
        output_path.suffix + ".tmp"
    )

    with temp_path.open(
            mode="w",
            encoding="utf-8",
    ) as file:
        json.dump(
            output_data,
            file,
            ensure_ascii=False,
            indent=2,
        )

    temp_path.replace(output_path)


def fetch_detail_task(
        hit: dict[str, Any],
) -> dict[str, Any]:
    """
    멀티스레드에서 업체 한 건의 상세정보를 조회한다.

    detail_map이나 JSON 파일은 수정하지 않고
    조회 결과만 메인 스레드에 반환한다.
    """
    organisation_id = clean_text(
        hit.get("organisationGuid")
    )

    event_edition_id = (
            clean_text(hit.get("eventEditionId"))
            or DEFAULT_EVENT_EDITION_ID
    )

    company_name = (
            clean_text(hit.get("companyName"))
            or clean_text(hit.get("exhibitorName"))
    )

    source_hit_id = hit.get("id")
    source_object_id = hit.get("objectID")

    try:
        session = get_thread_session()

        data = request_detail(
            session=session,
            event_edition_id=event_edition_id,
            organisation_id=organisation_id,
        )

        detail_entry = {
            "organisationId": organisation_id,
            "eventEditionId": event_edition_id,
            "sourceHitId": source_hit_id,
            "sourceObjectID": source_object_id,
            "data": data,
        }

        result = {
            "success": True,
            "organisationId": organisation_id,
            "companyName": company_name,
            "detailEntry": detail_entry,
            "error": "",
        }

    except Exception as exception:
        detail_entry = {
            "organisationId": organisation_id,
            "eventEditionId": event_edition_id,
            "sourceHitId": source_hit_id,
            "sourceObjectID": source_object_id,
            "data": None,
            "error": str(exception),
        }

        result = {
            "success": False,
            "organisationId": organisation_id,
            "companyName": company_name,
            "detailEntry": detail_entry,
            "error": str(exception),
        }

    # 같은 스레드가 너무 빠르게 연속 요청하는 것을 완화
    if REQUEST_DELAY_SECONDS > 0:
        time.sleep(REQUEST_DELAY_SECONDS)

    return result


def collect_details(
        hits: list[dict[str, Any]],
        detail_json_path: Path,
) -> dict[str, dict[str, Any]]:
    """
    업체 상세정보를 멀티스레드로 병렬 조회한다.

    1. 기존 detail JSON 읽기
    2. 정상 조회된 업체 제외
    3. organisationGuid 중복 제거
    4. MAX_WORKERS 개수만큼 병렬 요청
    5. 메인 스레드에서 detail_map 갱신
    6. 일정 건수마다 JSON 중간 저장
    """
    detail_map = load_existing_detail_map(
        detail_json_path
    )

    if detail_map:
        print(
            f"[이어하기] 기존 상세 데이터 "
            f"{len(detail_map):,}건을 읽었습니다."
        )

    # organisationGuid를 key로 사용하여 중복 요청 방지
    pending_hits: dict[str, dict[str, Any]] = {}

    cache_count = 0
    missing_count = 0
    duplicate_count = 0

    for hit in hits:
        organisation_id = clean_text(
            hit.get("organisationGuid")
        )

        if not organisation_id:
            missing_count += 1
            continue

        existing_detail = detail_map.get(
            organisation_id
        )

        # 정상 상세 데이터가 있으면 재요청하지 않음
        if has_valid_detail(existing_detail):
            cache_count += 1
            continue

        # 같은 organisationGuid는 한 번만 요청
        if organisation_id in pending_hits:
            duplicate_count += 1
            continue

        pending_hits[organisation_id] = hit

    request_hits = list(pending_hits.values())
    total_request_count = len(request_hits)

    print()
    print("=" * 70)
    print("상세조회 준비")
    print("=" * 70)
    print(f"입력 hits          : {len(hits):,}건")
    print(f"기존 정상 데이터   : {cache_count:,}건")
    print(f"organisationGuid 없음: {missing_count:,}건")
    print(f"중복 요청 제외     : {duplicate_count:,}건")
    print(f"신규 요청          : {total_request_count:,}건")
    print(f"병렬 스레드        : {MAX_WORKERS:,}개")
    print("=" * 70)
    print()

    if total_request_count == 0:
        save_detail_json(
            detail_map=detail_map,
            output_path=detail_json_path,
        )

        print(
            "[상세조회] 새로 요청할 데이터가 없습니다."
        )

        return detail_map

    success_count = 0
    failure_count = 0
    completed_count = 0

    future_map: dict[
        Future[dict[str, Any]],
        str,
    ] = {}

    executor = ThreadPoolExecutor(
        max_workers=MAX_WORKERS,
        thread_name_prefix="detail-worker",
    )

    try:
        # 작업 등록
        for hit in request_hits:
            organisation_id = clean_text(
                hit.get("organisationGuid")
            )

            future = executor.submit(
                fetch_detail_task,
                hit,
            )

            future_map[future] = organisation_id

        # 완료되는 순서대로 결과 처리
        for future in as_completed(future_map):
            completed_count += 1

            organisation_id = future_map[future]

            try:
                result = future.result()

                detail_entry = result.get(
                    "detailEntry"
                )

                if isinstance(detail_entry, dict):
                    detail_map[organisation_id] = (
                        detail_entry
                    )

                company_name = clean_text(
                    result.get("companyName")
                )

                if result.get("success"):
                    success_count += 1

                    print(
                        f"[성공] "
                        f"{completed_count:,}/"
                        f"{total_request_count:,} "
                        f"{company_name}"
                    )

                else:
                    failure_count += 1

                    print(
                        f"[실패] "
                        f"{completed_count:,}/"
                        f"{total_request_count:,} "
                        f"{company_name}"
                    )
                    print(
                        f"       {result.get('error', '')}"
                    )

            except Exception as exception:
                failure_count += 1

                detail_map[organisation_id] = {
                    "organisationId": organisation_id,
                    "data": None,
                    "error": str(exception),
                }

                print(
                    f"[스레드 오류] "
                    f"{completed_count:,}/"
                    f"{total_request_count:,} "
                    f"{organisation_id}"
                )
                print(f"              {exception}")

            # 일정 건수마다 중간 저장
            if (
                    completed_count % CHECKPOINT_INTERVAL == 0
                    or completed_count == total_request_count
            ):
                save_detail_json(
                    detail_map=detail_map,
                    output_path=detail_json_path,
                )

                print(
                    f"[중간 저장] "
                    f"{completed_count:,}/"
                    f"{total_request_count:,}, "
                    f"detail JSON {len(detail_map):,}건"
                )

    except KeyboardInterrupt:
        print()
        print(
            "[사용자 중단] 대기 중인 요청을 취소하고 "
            "현재까지 결과를 저장합니다."
        )

        for future in future_map:
            if not future.done():
                future.cancel()

        raise

    finally:
        executor.shutdown(
            wait=True,
            cancel_futures=True,
        )

        save_detail_json(
            detail_map=detail_map,
            output_path=detail_json_path,
        )

    print()
    print("=" * 70)
    print("상세조회 완료")
    print("=" * 70)
    print(f"신규 성공        : {success_count:,}건")
    print(f"신규 실패        : {failure_count:,}건")
    print(f"기존 데이터 사용 : {cache_count:,}건")
    print(f"ID 없음          : {missing_count:,}건")
    print(f"중복 요청 제외   : {duplicate_count:,}건")
    print("=" * 70)

    return detail_map


def get_detail_organisation(
        detail_entry: dict[str, Any] | None,
) -> dict[str, Any]:
    """
    detail entry에서 exhibitingOrganisation 객체를 반환한다.
    """
    if not isinstance(detail_entry, dict):
        return {}

    data = detail_entry.get("data")

    if not isinstance(data, dict):
        return {}

    organisation = data.get(
        "exhibitingOrganisation"
    )

    if not isinstance(organisation, dict):
        return {}

    return organisation


def has_address_value(
        multilingual_item: dict[str, Any],
) -> bool:
    """
    multilingual 객체에 주소 관련 값이 하나라도 있는지 확인한다.
    """
    address_fields = [
        "addressLine1",
        "addressLine2",
        "stateProvince",
        "city",
        "countryCode",
        "postcode",
        "country",
    ]

    for field_name in address_fields:
        if clean_text(
                multilingual_item.get(field_name)
        ):
            return True

    additional_addresses = multilingual_item.get(
        "additionalAddresses"
    )

    return (
            isinstance(additional_addresses, list)
            and len(additional_addresses) > 0
    )


def get_multilingual_data(
        detail_entry: dict[str, Any] | None,
        locale: str,
) -> dict[str, Any]:
    """
    상세 응답의 multilingual 배열에서 사용할 객체를 찾는다.

    우선순위:
    1. hit의 locale과 일치하는 항목
    2. 주소 정보가 존재하는 첫 번째 항목
    3. multilingual 첫 번째 객체
    """
    organisation = get_detail_organisation(
        detail_entry
    )

    multilingual = organisation.get(
        "multilingual"
    )

    if not isinstance(multilingual, list):
        return {}

    target_locale = clean_text(locale).lower()

    # 지정 locale 우선
    for item in multilingual:
        if not isinstance(item, dict):
            continue

        item_locale = clean_text(
            item.get("locale")
        ).lower()

        if item_locale == target_locale:
            return item

    # 주소가 있는 항목
    for item in multilingual:
        if not isinstance(item, dict):
            continue

        if has_address_value(item):
            return item

    # 첫 번째 객체
    for item in multilingual:
        if isinstance(item, dict):
            return item

    return {}


def get_address_fields(
        multilingual_data: dict[str, Any],
) -> dict[str, str]:
    """
    multilingual 객체의 주소 관련 8개 값을 그대로 반환한다.

    additionalAddresses 배열은 JSON 문자열로 저장한다.
    """
    additional_addresses = multilingual_data.get(
        "additionalAddresses"
    )

    if not isinstance(additional_addresses, list):
        additional_addresses = []

    return {
        "addressLine1": clean_text(
            multilingual_data.get("addressLine1")
        ),
        "addressLine2": clean_text(
            multilingual_data.get("addressLine2")
        ),
        "stateProvince": clean_text(
            multilingual_data.get("stateProvince")
        ),
        "city": clean_text(
            multilingual_data.get("city")
        ),
        "countryCode": clean_text(
            multilingual_data.get("countryCode")
        ),
        "postcode": clean_text(
            multilingual_data.get("postcode")
        ),
        "country": clean_text(
            multilingual_data.get("country")
        ),
        "additionalAddresses": clean_text(
            additional_addresses
        ),
    }


def remove_category_code(
        value: Any,
) -> str:
    """
    카테고리 앞쪽 내부 관리번호를 제거한다.

    예:
    948439:2: XR・イマーシブ体験

    결과:
    XR・イマーシブ体験
    """
    text = clean_text(value)

    return re.sub(
        r"^\d+\s*:\s*\d+\s*:\s*",
        "",
        text,
    ).strip()


def get_categories(
        hit: dict[str, Any],
) -> list[str]:
    """
    hit의 제품 카테고리를 가져온다.
    """
    exhibitor_filters = hit.get(
        "exhibitorFilters"
    )

    if isinstance(exhibitor_filters, dict):
        product_category = exhibitor_filters.get(
            "製品カテゴリー"
        )

        if isinstance(product_category, dict):
            level1_categories = product_category.get(
                "lvl1"
            )

            if isinstance(level1_categories, list):
                categories: list[str] = []

                for value in level1_categories:
                    category = remove_category_code(
                        value
                    )

                    if (
                            category
                            and category not in categories
                    ):
                        categories.append(category)

                if categories:
                    return categories

    # 製品カテゴリー가 없으면 ppsAnswers 앞쪽 값 사용
    pps_answers = hit.get("ppsAnswers")

    if not isinstance(pps_answers, list):
        return []

    categories: list[str] = []

    for value in pps_answers[:3]:
        category = clean_text(value)

        if (
                category
                and category not in categories
        ):
            categories.append(category)

    return categories


def convert_hit_to_row(
        hit: dict[str, Any],
        detail_entry: dict[str, Any] | None,
) -> dict[str, str]:
    """
    hit와 상세 응답을 결합해 CSV/XLSX 한 행으로 변환한다.
    """
    organisation = get_detail_organisation(
        detail_entry
    )

    locale = (
            clean_text(hit.get("locale"))
            or TARGET_LOCALE
    )

    multilingual_data = get_multilingual_data(
        detail_entry=detail_entry,
        locale=locale,
    )

    company_name = (
            clean_text(hit.get("companyName"))
            or clean_text(hit.get("exhibitorName"))
            or clean_text(organisation.get("companyName"))
            or clean_text(
        multilingual_data.get("displayName")
    )
    )

    show_objective = (
            clean_text(hit.get("showObjective"))
            or clean_text(
        multilingual_data.get("showObjective")
    )
    )

    description = (
            clean_text(
                hit.get("exhibitorDescription")
            )
            or clean_text(
        multilingual_data.get("description")
    )
    )

    website = (
            clean_text(hit.get("website"))
            or clean_text(organisation.get("website"))
            or clean_text(
        organisation.get(
            "accompanyingWebsiteUrl"
        )
    )
    )

    email = (
            clean_text(hit.get("email"))
            or clean_text(
        organisation.get("contactEmail")
    )
    )

    phone = (
            clean_text(hit.get("phone"))
            or clean_text(organisation.get("phone"))
    )

    address_fields = get_address_fields(
        multilingual_data
    )

    categories = get_categories(hit)

    return {
        "회사명": company_name,
        "포인트": show_objective,
        "사업내용": description,
        "홈페이지": website,
        "이메일": email,
        "연락처": phone,
        "addressLine1": address_fields[
            "addressLine1"
        ],
        "addressLine2": address_fields[
            "addressLine2"
        ],
        "stateProvince": address_fields[
            "stateProvince"
        ],
        "city": address_fields["city"],
        "countryCode": address_fields[
            "countryCode"
        ],
        "postcode": address_fields["postcode"],
        "country": address_fields["country"],
        "additionalAddresses": address_fields[
            "additionalAddresses"
        ],
        "카테고리": " | ".join(categories),
    }


def create_rows(
        hits: list[dict[str, Any]],
        detail_map: dict[str, dict[str, Any]],
) -> list[dict[str, str]]:
    """
    전체 hits를 CSV/XLSX 행으로 변환한다.
    """
    rows: list[dict[str, str]] = []

    total_count = len(hits)

    for index, hit in enumerate(
            hits,
            start=1,
    ):
        organisation_id = clean_text(
            hit.get("organisationGuid")
        )

        detail_entry = detail_map.get(
            organisation_id
        )

        row = convert_hit_to_row(
            hit=hit,
            detail_entry=detail_entry,
        )

        rows.append(row)

        if (
                index % 100 == 0
                or index == total_count
        ):
            print(
                f"[변환] {index:,}/{total_count:,}"
            )

    return rows


def save_csv(
        rows: list[dict[str, str]],
        output_path: Path,
) -> None:
    """
    CSV 파일을 UTF-8 BOM 형식으로 저장한다.
    일본어가 Excel에서 깨지지 않는다.
    """
    with output_path.open(
            mode="w",
            encoding="utf-8-sig",
            newline="",
    ) as file:
        writer = csv.DictWriter(
            file,
            fieldnames=HEADERS,
            quoting=csv.QUOTE_MINIMAL,
        )

        writer.writeheader()
        writer.writerows(rows)


def save_xlsx(
        rows: list[dict[str, str]],
        output_path: Path,
) -> None:
    """
    XLSX 파일을 저장한다.
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "업체목록"

    # 헤더
    for column_index, header in enumerate(
            HEADERS,
            start=1,
    ):
        cell = worksheet.cell(
            row=1,
            column=column_index,
            value=header,
        )

        cell.font = Font(
            bold=True,
            color="FFFFFF",
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="4472C4",
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    # 데이터
    for row_index, row_data in enumerate(
            rows,
            start=2,
    ):
        for column_index, header in enumerate(
                HEADERS,
                start=1,
        ):
            value = clean_text(
                row_data.get(header, "")
            )

            # Excel 셀 문자열 최대 길이
            if len(value) > 32767:
                value = value[:32767]

            cell = worksheet.cell(
                row=row_index,
                column=column_index,
                value=value,
            )

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    column_widths = {
        "회사명": 30,
        "포인트": 60,
        "사업내용": 80,
        "홈페이지": 40,
        "이메일": 35,
        "연락처": 22,
        "addressLine1": 70,
        "addressLine2": 45,
        "stateProvince": 25,
        "city": 25,
        "countryCode": 15,
        "postcode": 18,
        "country": 20,
        "additionalAddresses": 80,
        "카테고리": 70,
    }

    for column_index, header in enumerate(
            HEADERS,
            start=1,
    ):
        column_letter = get_column_letter(
            column_index
        )

        worksheet.column_dimensions[
            column_letter
        ].width = column_widths.get(
            header,
            25,
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.row_dimensions[1].height = 25

    workbook.save(output_path)


def main() -> None:
    execution_directory = get_execution_directory()

    input_json_path = (
            execution_directory
            / INPUT_HITS_JSON_NAME
    )

    detail_json_path = (
            execution_directory
            / OUTPUT_DETAIL_JSON_NAME
    )

    csv_path = (
            execution_directory
            / OUTPUT_CSV_NAME
    )

    xlsx_path = (
            execution_directory
            / OUTPUT_XLSX_NAME
    )

    print("=" * 70)
    print("Content Tokyo 상세정보 멀티스레드 수집")
    print("=" * 70)
    print(f"실행 경로 : {execution_directory}")
    print(f"입력 JSON : {input_json_path}")
    print(f"상세 JSON : {detail_json_path}")
    print(f"스레드 수 : {MAX_WORKERS}")
    print()

    try:
        # 1. hits JSON 읽기
        hits = load_hits(input_json_path)

        print(
            f"[입력 완료] hits {len(hits):,}건"
        )

        # 2. 상세정보 멀티스레드 조회
        detail_map = collect_details(
            hits=hits,
            detail_json_path=detail_json_path,
        )

        print()
        print(
            "[변환] CSV/XLSX 데이터 생성 중..."
        )

        # 3. hits + 상세정보 결합
        rows = create_rows(
            hits=hits,
            detail_map=detail_map,
        )

        # 4. CSV 저장
        print("[저장] CSV 생성 중...")

        save_csv(
            rows=rows,
            output_path=csv_path,
        )

        # 5. XLSX 저장
        print("[저장] XLSX 생성 중...")

        save_xlsx(
            rows=rows,
            output_path=xlsx_path,
        )

        address_fields = [
            "addressLine1",
            "addressLine2",
            "stateProvince",
            "city",
            "countryCode",
            "postcode",
            "country",
            "additionalAddresses",
        ]

        address_data_count = sum(
            1
            for row in rows
            if any(
                clean_text(row.get(field_name))
                for field_name in address_fields
            )
        )

        print()
        print("=" * 70)
        print("작업 완료")
        print("=" * 70)
        print(f"상세 JSON : {detail_json_path}")
        print(f"CSV       : {csv_path}")
        print(f"XLSX      : {xlsx_path}")
        print(f"전체 업체 : {len(rows):,}건")
        print(
            f"주소 데이터 있음 : "
            f"{address_data_count:,}건"
        )
        print(
            f"주소 데이터 없음 : "
            f"{len(rows) - address_data_count:,}건"
        )
        print("=" * 70)

    except FileNotFoundError as exception:
        print()
        print(f"[파일 오류] {exception}")

    except json.JSONDecodeError as exception:
        print()
        print(
            "[JSON 오류] 입력 JSON 형식이 "
            "올바르지 않습니다."
        )
        print(
            f"오류 위치: {exception.lineno}행 "
            f"{exception.colno}열"
        )
        print(f"상세 오류: {exception}")

    except requests.HTTPError as exception:
        response = exception.response

        print()
        print(
            "[HTTP 오류] 상세 API 요청에 실패했습니다."
        )

        if response is not None:
            print(
                f"상태 코드: {response.status_code}"
            )
            print(
                f"응답 내용: {response.text[:2000]}"
            )

    except requests.RequestException as exception:
        print()
        print(f"[통신 오류] {exception}")

    except PermissionError as exception:
        print()
        print(
            "[파일 오류] JSON, CSV 또는 XLSX 파일이 "
            "열려 있거나 쓰기 권한이 없습니다."
        )
        print(f"상세 오류: {exception}")

    except KeyboardInterrupt:
        print()
        print(
            "[사용자 중단] 현재까지 수집한 상세정보는 "
            "detail JSON에 저장했습니다."
        )

    except Exception as exception:
        print()
        print(
            f"[오류] {type(exception).__name__}: "
            f"{exception}"
        )

    input("\n엔터를 누르면 종료합니다...")


if __name__ == "__main__":
    main()