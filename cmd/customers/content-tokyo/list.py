from __future__ import annotations

import csv
import json
import re
import sys
from pathlib import Path
from typing import Any
from urllib.parse import quote, urlencode

import requests
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry


# =========================================================
# API 설정
# =========================================================
API_URL = (
    "https://xd0u5m6y4r-dsn.algolia.net/"
    "1/indexes/"
    "evt-57bc0de5-45b7-477f-aa0d-1d4cf08fc4c5-index/"
    "query"
)

ALGOLIA_APPLICATION_ID = "XD0U5M6Y4R"
ALGOLIA_API_KEY = "d5cd7d4ec26134ff4a34d736a7f9ad47"

EVENT_EDITION_ID = "eve-2b972fc3-9919-4475-983b-a08253dfd7d1"
LOCALE = "ja-jp"
HITS_PER_PAGE = 100


# =========================================================
# 출력 파일 설정
# =========================================================
OUTPUT_JSON_NAME = "content_tokyo_exhibitors_hits.json"
OUTPUT_CSV_NAME = "content_tokyo_exhibitors.csv"
OUTPUT_XLSX_NAME = "content_tokyo_exhibitors.xlsx"

HEADERS = [
    "회사명",
    "포인트",
    "사업내용",
    "홈페이지",
    "이메일",
    "연락처",
    "카테고리",
]


def get_execution_directory() -> Path:
    """
    실행 중인 Python 파일 또는 exe 파일이 있는 경로를 반환한다.
    """
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent

    return Path(__file__).resolve().parent


def create_session() -> requests.Session:
    """
    요청 실패 시 자동 재시도하는 Session을 생성한다.
    """
    retry = Retry(
        total=5,
        connect=5,
        read=5,
        status=5,
        backoff_factor=1,
        status_forcelist=[
            429,
            500,
            502,
            503,
            504,
        ],
        allowed_methods=frozenset(["POST"]),
        respect_retry_after_header=True,
    )

    adapter = HTTPAdapter(
        max_retries=retry,
        pool_connections=10,
        pool_maxsize=10,
    )

    session = requests.Session()
    session.mount("https://", adapter)

    return session


def build_algolia_params(page: int) -> str:
    """
    Algolia 요청 body에 들어갈 params 문자열을 생성한다.
    """
    filters = (
        "recordType:exhibitor "
        f"AND locale:{LOCALE} "
        f"AND eventEditionId:{EVENT_EDITION_ID}"
    )

    params = {
        "query": "",
        "page": str(page),
        "hitsPerPage": str(HITS_PER_PAGE),
        "filters": filters,
        "facetFilters": "",
        "optionalFilters": "[]",
    }

    return urlencode(
        params,
        quote_via=quote,
    )


def request_page(
        session: requests.Session,
        page: int,
) -> dict[str, Any]:
    """
    지정한 페이지의 데이터를 요청한다.
    """
    query_parameters = {
        "x-algolia-agent": "Algolia for JavaScript (3.35.1); Browser",
        "x-algolia-application-id": ALGOLIA_APPLICATION_ID,
        "x-algolia-api-key": ALGOLIA_API_KEY,
    }

    headers = {
        "Accept": "application/json",
        "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
        "Content-Type": "application/x-www-form-urlencoded",
        "Origin": "https://www.content-tokyo.jp",
        "Referer": "https://www.content-tokyo.jp/",
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/149.0.0.0 Safari/537.36"
        ),
    }

    payload = {
        "params": build_algolia_params(page)
    }

    # 중요:
    # data=payload로 보내면 form-urlencoded 형태가 되어 400 오류 발생
    # JSON 문자열로 변환해서 보내야 함
    response = session.post(
        API_URL,
        params=query_parameters,
        headers=headers,
        data=json.dumps(
            payload,
            ensure_ascii=False,
        ).encode("utf-8"),
        timeout=(10, 60),
    )

    response.raise_for_status()

    result = response.json()

    if not isinstance(result, dict):
        raise ValueError(
            f"페이지 {page} 응답이 JSON 객체가 아닙니다."
        )

    hits = result.get("hits")

    if not isinstance(hits, list):
        raise ValueError(
            f"페이지 {page} 응답에서 hits 배열을 찾을 수 없습니다."
        )

    return result


def collect_all_hits() -> list[dict[str, Any]]:
    """
    첫 페이지에서 nbPages를 확인한 뒤 전체 페이지를 요청한다.
    """
    session = create_session()

    try:
        print("[조회] page=0 요청 중...")

        first_result = request_page(
            session=session,
            page=0,
        )

        total_pages = int(first_result.get("nbPages", 1))
        total_hits = int(first_result.get("nbHits", 0))

        print(
            f"[정보] 전체 검색 건수: {total_hits:,}건"
        )
        print(
            f"[정보] 전체 페이지 수: {total_pages:,}페이지"
        )

        all_hits: list[dict[str, Any]] = []

        first_hits = first_result.get("hits", [])

        for item in first_hits:
            if isinstance(item, dict):
                all_hits.append(item)

        print(
            f"[완료] 1 / {total_pages} 페이지, "
            f"{len(first_hits):,}건"
        )

        # nbPages가 10이면 실제 page 값은 0~9
        for page in range(1, total_pages):
            print(
                f"[조회] page={page} 요청 중..."
            )

            result = request_page(
                session=session,
                page=page,
            )

            page_hits = result.get("hits", [])
            valid_count = 0

            for item in page_hits:
                if not isinstance(item, dict):
                    continue

                all_hits.append(item)
                valid_count += 1

            print(
                f"[완료] {page + 1} / {total_pages} 페이지, "
                f"{valid_count:,}건"
            )

        return all_hits

    finally:
        session.close()


def make_deduplication_key(
        item: dict[str, Any],
) -> str:
    """
    중복 제거 기준값을 만든다.

    우선순위:
    1. id
    2. objectID
    3. 전체 JSON 내용
    """
    item_id = item.get("id")

    if item_id is not None:
        item_id_text = str(item_id).strip()

        if item_id_text:
            return f"id:{item_id_text}"

    object_id = item.get("objectID")

    if object_id is not None:
        object_id_text = str(object_id).strip()

        if object_id_text:
            return f"objectID:{object_id_text}"

    return "json:" + json.dumps(
        item,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )


def remove_duplicate_hits(
        hits: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """
    전체 hits에서 중복을 제거한다.

    같은 키가 여러 번 나오면 처음 조회된 데이터를 유지한다.
    """
    unique_hits: list[dict[str, Any]] = []
    seen_keys: set[str] = set()

    for item in hits:
        key = make_deduplication_key(item)

        if key in seen_keys:
            continue

        seen_keys.add(key)
        unique_hits.append(item)

    return unique_hits


def clean_text(value: Any) -> str:
    """
    CSV 및 XLSX에 저장할 문자열을 정리한다.
    """
    if value is None:
        return ""

    if isinstance(value, (list, dict)):
        value = json.dumps(
            value,
            ensure_ascii=False,
        )

    text = str(value)

    # Excel에서 허용하지 않는 제어문자 제거
    text = ILLEGAL_CHARACTERS_RE.sub("", text)

    text = text.replace("\r\n", "\n")
    text = text.replace("\r", "\n")

    return text.strip()


def remove_category_code(value: Any) -> str:
    """
    카테고리 앞쪽 관리번호를 제거한다.

    예:
    948439:2: XR・イマーシブ体験 > XR／メタバースソリューション

    결과:
    XR・イマーシブ体験 > XR／メタバースソリューション
    """
    text = clean_text(value)

    return re.sub(
        r"^\d+\s*:\s*\d+\s*:\s*",
        "",
        text,
    ).strip()


def get_categories(
        item: dict[str, Any],
) -> list[str]:
    """
    exhibitorFilters > 製品カテゴリー > lvl1에서
    카테고리를 추출한다.
    """
    exhibitor_filters = item.get("exhibitorFilters")

    if isinstance(exhibitor_filters, dict):
        product_category = exhibitor_filters.get(
            "製品カテゴリー"
        )

        if isinstance(product_category, dict):
            level1_categories = product_category.get(
                "lvl1"
            )

            if isinstance(level1_categories, list):
                categories: list[str] = []

                for value in level1_categories:
                    category = remove_category_code(value)

                    if (
                            category
                            and category not in categories
                    ):
                        categories.append(category)

                if categories:
                    return categories

    # 製品カテゴリー가 없는 경우 ppsAnswers 앞의 3개를 보조 사용
    pps_answers = item.get("ppsAnswers")

    if not isinstance(pps_answers, list):
        return []

    categories: list[str] = []

    for value in pps_answers[:3]:
        category = clean_text(value)

        if category and category not in categories:
            categories.append(category)

    return categories


def convert_hit_to_row(
        item: dict[str, Any],
) -> dict[str, str]:
    """
    원본 hit 한 건을 CSV/XLSX용 데이터로 변환한다.
    """
    company_name = (
            clean_text(item.get("companyName"))
            or clean_text(item.get("exhibitorName"))
    )

    categories = get_categories(item)

    return {
        "회사명": company_name,
        "포인트": clean_text(
            item.get("showObjective")
        ),
        "사업내용": clean_text(
            item.get("exhibitorDescription")
        ),
        "홈페이지": clean_text(
            item.get("website")
        ),
        "이메일": clean_text(
            item.get("email")
        ),
        "연락처": clean_text(
            item.get("phone")
        ),
        "카테고리": " | ".join(categories),
    }


def save_hits_json(
        hits: list[dict[str, Any]],
        output_path: Path,
) -> None:
    """
    중복 제거된 원본 hits 전체 데이터를 JSON으로 저장한다.

    저장 구조:
    {
        "hits": [
            {...},
            {...}
        ]
    }
    """
    output_data = {
        "hits": hits
    }

    with output_path.open(
            mode="w",
            encoding="utf-8",
    ) as file:
        json.dump(
            output_data,
            file,
            ensure_ascii=False,
            indent=2,
        )


def save_csv(
        rows: list[dict[str, str]],
        output_path: Path,
) -> None:
    """
    CSV 파일을 저장한다.

    utf-8-sig를 사용하므로 Excel에서 일본어가 깨지지 않는다.
    """
    with output_path.open(
            mode="w",
            encoding="utf-8-sig",
            newline="",
    ) as file:
        writer = csv.DictWriter(
            file,
            fieldnames=HEADERS,
            quoting=csv.QUOTE_MINIMAL,
        )

        writer.writeheader()
        writer.writerows(rows)


def save_xlsx(
        rows: list[dict[str, str]],
        output_path: Path,
) -> None:
    """
    XLSX 파일을 저장한다.
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "업체목록"

    # 헤더 생성
    for column_index, header in enumerate(
            HEADERS,
            start=1,
    ):
        cell = worksheet.cell(
            row=1,
            column=column_index,
            value=header,
        )

        cell.font = Font(
            bold=True,
            color="FFFFFF",
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="4472C4",
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    # 데이터 생성
    for row_index, row_data in enumerate(
            rows,
            start=2,
    ):
        for column_index, header in enumerate(
                HEADERS,
                start=1,
        ):
            value = clean_text(
                row_data.get(header, "")
            )

            # Excel 셀 문자열 최대 길이 제한
            if len(value) > 32767:
                value = value[:32767]

            cell = worksheet.cell(
                row=row_index,
                column=column_index,
                value=value,
            )

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    column_widths = {
        "회사명": 30,
        "포인트": 60,
        "사업내용": 80,
        "홈페이지": 40,
        "이메일": 35,
        "연락처": 22,
        "카테고리": 70,
    }

    for column_index, header in enumerate(
            HEADERS,
            start=1,
    ):
        column_letter = get_column_letter(
            column_index
        )

        worksheet.column_dimensions[
            column_letter
        ].width = column_widths.get(
            header,
            25,
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.row_dimensions[1].height = 25

    workbook.save(output_path)


def main() -> None:
    execution_directory = get_execution_directory()

    json_path = (
            execution_directory
            / OUTPUT_JSON_NAME
    )

    csv_path = (
            execution_directory
            / OUTPUT_CSV_NAME
    )

    xlsx_path = (
            execution_directory
            / OUTPUT_XLSX_NAME
    )

    print("=" * 70)
    print("Content Tokyo 전시업체 데이터 수집")
    print("=" * 70)
    print(f"실행 경로: {execution_directory}")
    print()

    try:
        # 1. 전체 페이지의 hits 수집
        all_hits = collect_all_hits()

        print()
        print(
            f"[수집 완료] 원본 hits: "
            f"{len(all_hits):,}건"
        )

        # 2. 원본 hits 중복 제거
        unique_hits = remove_duplicate_hits(
            all_hits
        )

        duplicate_count = (
                len(all_hits)
                - len(unique_hits)
        )

        print(
            f"[중복 제거] "
            f"{duplicate_count:,}건 제거"
        )
        print(
            f"[최종 hits] "
            f"{len(unique_hits):,}건"
        )

        # 3. 중복 제거된 원본 hits 전체를 JSON으로 저장
        print()
        print("[저장] 원본 hits JSON 생성 중...")

        save_hits_json(
            hits=unique_hits,
            output_path=json_path,
        )

        # 4. CSV/XLSX용 데이터 변환
        print("[변환] CSV/XLSX 데이터 생성 중...")

        rows: list[dict[str, str]] = []

        for index, item in enumerate(
                unique_hits,
                start=1,
        ):
            rows.append(
                convert_hit_to_row(item)
            )

            if (
                    index % 100 == 0
                    or index == len(unique_hits)
            ):
                print(
                    f"[변환] {index:,} / "
                    f"{len(unique_hits):,}"
                )

        # 5. CSV 저장
        print("[저장] CSV 생성 중...")

        save_csv(
            rows=rows,
            output_path=csv_path,
        )

        # 6. XLSX 저장
        print("[저장] XLSX 생성 중...")

        save_xlsx(
            rows=rows,
            output_path=xlsx_path,
        )

        print()
        print("=" * 70)
        print("작업 완료")
        print("=" * 70)
        print(
            f"원본 JSON : {json_path}"
        )
        print(
            f"CSV       : {csv_path}"
        )
        print(
            f"XLSX      : {xlsx_path}"
        )
        print(
            f"최종 저장 건수: "
            f"{len(unique_hits):,}건"
        )
        print("=" * 70)

    except requests.HTTPError as exception:
        response = exception.response

        print()
        print(
            "[HTTP 오류] "
            "API 요청에 실패했습니다."
        )

        if response is not None:
            print(
                f"상태 코드: "
                f"{response.status_code}"
            )
            print(
                f"응답 내용: "
                f"{response.text[:1000]}"
            )

    except requests.RequestException as exception:
        print()
        print(
            f"[통신 오류] {exception}"
        )

    except PermissionError as exception:
        print()
        print(
            "[파일 오류] 출력 파일이 열려 있거나 "
            "쓰기 권한이 없습니다."
        )
        print(
            "CSV, XLSX, JSON 파일을 닫은 뒤 "
            "다시 실행해 주세요."
        )
        print(
            f"상세 오류: {exception}"
        )

    except json.JSONDecodeError as exception:
        print()
        print(
            "[JSON 오류] API 응답을 "
            "JSON으로 변환할 수 없습니다."
        )
        print(
            f"상세 오류: {exception}"
        )

    except Exception as exception:
        print()
        print(
            f"[오류] "
            f"{type(exception).__name__}: "
            f"{exception}"
        )

    input("\n엔터를 누르면 종료합니다...")


if __name__ == "__main__":
    main()