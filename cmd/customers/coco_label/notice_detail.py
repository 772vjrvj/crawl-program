import csv
import io
import json
import os
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import urlencode

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from PIL import Image, ImageOps


# =========================
# 설정
# =========================
LIST_JSON_FILE = "coco_posts.json"
RESULT_CSV_FILE = "coco_notice_result.csv"
OUTPUT_EXCEL_FILE = "coco_notice_result.xlsx"
NOTICE_ROOT_DIR = "notice"

BASE_URL = "https://coco-label.com/admin/ajax/contents/v2/api.cm"

COMMON_QUERY_PARAMS = {
    "site_code": "S202410211a92d560f8f0e",
    "unit_code": "u202410216715ffa21a17f",
    "board_code": "",
    "endpoint": "post",
    "type": "post",
}

HEADERS = {
    "accept": "application/json, text/plain, */*",
    "accept-encoding": "gzip, deflate, br, zstd",
    "accept-language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
    "cache-control": "no-cache",
    "content-length": "0",
    "content-type": "application/x-www-form-urlencoded",
    # === 쿠키 직접 넣으세요 ===
    "cookie": "al=KR; _fwb=153Fgpdb68AF6wtU0X73tQL.1774768427504; _fbp=fb.1.1774768429368.806216781255867927; __fs_imweb=%7B%22deviceId%22%3A%22mnbfcju9-0430ae7c4b36b520ae04b4544cb4be37-454gbpp%22%2C%22useSubDomain%22%3A%22Y%22%7D; _dd_s=aid=a4e0643f-4dfa-4b70-ad58-a8b3d777aa9d&rum=2&id=090e1d54-79c0-4615-bc55-dd67bba97bc4&created=1775739387350&expire=1775740666071; __bs_imweb=%7B%22deviceId%22%3A%22019d387118fb7a158d09c8b1123c61b0%22%2C%22deviceIdCreatedAt%22%3A%222025-02-15T18%3A30%3A00%22%2C%22siteCode%22%3A%22S202410211a92d560f8f0e%22%2C%22unitCode%22%3A%22u202410216715ffa21a17f%22%2C%22platform%22%3A%22DESKTOP%22%2C%22browserSessionId%22%3A%22019d7525774f7118a76f1c7fd0aa4ebe%22%2C%22sdkJwt%22%3A%22eyJhbGciOiJFUzI1NiIsImtpZCI6bnVsbH0.eyJzdWIiOiJtMjAyNTA1MjAzMjE1Mjk1NTlkZTJmIiwic2l0ZUNvZGUiOiJTMjAyNDEwMjExYTkyZDU2MGY4ZjBlIiwidW5pdENvZGUiOiJ1MjAyNDEwMjE2NzE1ZmZhMjFhMTdmIiwiY2hlY2tPZmZpY2UiOmZhbHNlLCJpYXQiOjE3NzU3ODY5MjAsImV4cCI6MTc3NTc4NzUyMH0.G47myWFkCyQnlgSss_fi2Gko4aIuCnJc5aUCW4dN4hM%22%2C%22referrer%22%3A%22https%3A%2F%2Fcoco-label.com%2F%22%2C%22initialReferrer%22%3A%22%40direct%22%2C%22initialReferrerDomain%22%3A%22%40direct%22%2C%22utmSource%22%3Anull%2C%22utmMedium%22%3Anull%2C%22utmCampaign%22%3Anull%2C%22utmTerm%22%3Anull%2C%22utmContent%22%3Anull%2C%22utmLandingUrl%22%3Anull%2C%22utmUpdatedTime%22%3Anull%2C%22updatedAt%22%3A%222026-04-10T02%3A08%3A51.347Z%22%2C%22commonSessionId%22%3A%22sc_019d752577517bb1a5f8a1048f282778%22%2C%22commonSessionUpdatedAt%22%3A%222026-04-10T02%3A08%3A41.061Z%22%2C%22customSessionId%22%3A%22cs_019d7525775276e291c8906215a43ad9%22%2C%22customSessionUpdatedAt%22%3A%222026-04-10T02%3A08%3A41.062Z%22%2C%22browser_session_id%22%3A%22019d7525774f7118a76f1c7fd0aa4ebe%22%2C%22sdk_jwt%22%3A%22eyJhbGciOiJFUzI1NiIsImtpZCI6bnVsbH0.eyJzdWIiOiJtMjAyNTA1MjAzMjE1Mjk1NTlkZTJmIiwic2l0ZUNvZGUiOiJTMjAyNDEwMjExYTkyZDU2MGY4ZjBlIiwidW5pdENvZGUiOiJ1MjAyNDEwMjE2NzE1ZmZhMjFhMTdmIiwiY2hlY2tPZmZpY2UiOmZhbHNlLCJpYXQiOjE3NzU3ODY5MjAsImV4cCI6MTc3NTc4NzUyMH0.G47myWFkCyQnlgSss_fi2Gko4aIuCnJc5aUCW4dN4hM%22%2C%22initial_referrer%22%3A%22%40direct%22%2C%22initial_referrer_domain%22%3A%22%40direct%22%2C%22utm_source%22%3Anull%2C%22utm_medium%22%3Anull%2C%22utm_campaign%22%3Anull%2C%22utm_term%22%3Anull%2C%22utm_content%22%3Anull%2C%22utm_landing_url%22%3Anull%2C%22utm_updated_time%22%3Anull%2C%22updated_at%22%3A%222026-04-10T02%3A08%3A51.347Z%22%2C%22common_session_id%22%3A%22sc_019d752577517bb1a5f8a1048f282778%22%2C%22common_session_updated_at%22%3A%222026-04-10T02%3A08%3A41.061Z%22%2C%22custom_session_id%22%3A%22cs_019d7525775276e291c8906215a43ad9%22%2C%22custom_session_updated_at%22%3A%222026-04-10T02%3A08%3A41.062Z%22%7D; IMWEBVSSID=2372eadpe4lfn41nsnkc597oggcda7hvh96sgpu09vnjuvmsqge9k4ths3kkvv23i9ipfd3jhgp00p75brvn8jfkgiamjb63ctktv00; ISDID=69db984ebf2cf; ilc=%2BptAWpJ8GLxIVCY4yyewr0sI3mA2nt6gsm4mOiTKFXA%3D; ial=c81c5d84a43a12e659a0fa45c685e65ed4820ef001221e91c6eec9061d71f124; _clck=1r1oz16%5E2%5Eg55%5E0%5E2279; IMWEB_REFRESH_TOKEN=8298d8c4-2f70-436e-9d46-2b1821abe5d5; _clsk=1f0pg02%5E1776002260657%5E39%5E1%5Eq.clarity.ms%2Fcollect; IMWEB_ACCESS_TOKEN=eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJob3N0IjoiY29jby1sYWJlbC5jb20iLCJzaXRlQ29kZSI6IlMyMDI0MTAyMTFhOTJkNTYwZjhmMGUiLCJ1bml0Q29kZSI6InUyMDI0MTAyMTY3MTVmZmEyMWExN2YiLCJtZW1iZXJDb2RlIjoibTIwMjUwNTIwMzIxNTI5NTU5ZGUyZiIsInJvbGUiOiJvd25lciIsImlhdCI6MTc3NjAwMzE3NiwiZXhwIjoxNzc2MDAzNDc2LCJpc3MiOiJpbXdlYi1jb3JlLWF1dGgtc2l0ZSJ9.R_qVrWP8xbmmweNuW_1fNNVU852WaShrw9SB3dJxMpo; mp_a4939111ea54962dbf95fe89a992eab3_mixpanel=%7B%22distinct_id%22%3A%22%22%2C%22%24device_id%22%3A%22920327e0-74a9-4609-90f6-77825e72978e%22%2C%22from_imweb_office%22%3Afalse%2C%22%24initial_referrer%22%3A%22https%3A%2F%2Fcoco-label.com%2Fbackpg%2Flogin.cm%3Fback_url%3DaHR0cHM6Ly9jb2NvLWxhYmVsLmNvbS9hZG1pbg%253D%253D%22%2C%22%24initial_referring_domain%22%3A%22coco-label.com%22%2C%22__mps%22%3A%7B%7D%2C%22__mpso%22%3A%7B%7D%2C%22__mpus%22%3A%7B%7D%2C%22__mpa%22%3A%7B%7D%2C%22__mpu%22%3A%7B%7D%2C%22__mpr%22%3A%5B%5D%2C%22__mpap%22%3A%5B%5D%7D",
    "origin": "https://coco-label.com",
    "pragma": "no-cache",
    "priority": "u=1, i",
    "referer": "https://coco-label.com/_/site-content/",
    "sec-ch-ua": '"Google Chrome";v="147", "Not.A/Brand";v="8", "Chromium";v="147"',
    "sec-ch-ua-mobile": "?0",
    "sec-ch-ua-platform": '"Windows"',
    "sec-fetch-dest": "empty",
    "sec-fetch-mode": "cors",
    "sec-fetch-site": "same-origin",
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/147.0.0.0 Safari/537.36",
}

IMAGE_HEADERS = {
    "user-agent": HEADERS["user-agent"],
    "referer": "https://coco-label.com/",
}

MAX_IMAGE_COUNT = 5
MAX_IMAGE_WIDTH = 600
MAX_IMAGE_WORKERS = 5


# =========================
# CSV 컬럼
# =========================
RESULT_COLUMNS = [
    "wr_id",
    "code",
    "ex_1",
    "wr_subject",
    "wr_content",
    "wr_seo_title",
    "wr_file",
    "success_yn",
]


# =========================
# 엑셀 컬럼
# =========================
WRITE_CHECK_COLUMNS = [
    "wr_id", "wr_num", "wr_reply", "wr_parent", "wr_is_comment", "wr_comment",
    "wr_comment_reply", "ca_name", "wr_option", "wr_subject", "wr_content",
    "wr_seo_title", "wr_link1", "wr_link2", "wr_link1_hit", "wr_link2_hit",
    "wr_hit", "wr_anonymous", "wr_good", "wr_nogood", "mb_id", "wr_password",
    "wr_name", "wr_email", "wr_homepage", "wr_datetime", "wr_file", "wr_last",
    "wr_ip", "wr_facebook_user", "wr_twitter_user", "wr_1", "wr_2", "wr_3",
    "wr_4", "wr_5", "wr_6", "wr_7", "wr_8", "wr_9", "wr_10", "eb_1", "eb_2",
    "eb_3", "eb_4", "eb_5", "eb_6", "eb_7", "eb_8", "eb_9", "eb_10", "ex_1", "ex_2"
]

BOARD_FILE_COLUMNS = [
    "bo_table", "wr_id", "bf_no", "bf_source", "bf_file",
    "bf_path_origin", "bf_path_now",
    "bf_download", "bf_content", "bf_fileurl", "bf_thumburl",
    "bf_storage", "bf_filesize", "bf_width", "bf_height", "bf_type", "bf_datetime"
]


# =========================
# 유틸
# =========================
def ensure_dir(path: str):
    os.makedirs(path, exist_ok=True)


def safe_int(value, default=0):
    try:
        return int(value)
    except Exception:
        return default


def safe_folder_name(text: str) -> str:
    text = str(text or "").strip()
    text = re.sub(r'[\\/:*?"<>|]+', "_", text)
    return text


def normalize_image_url(url: str) -> str:
    url = (url or "").strip()
    if url.startswith("//"):
        return "https:" + url
    if url.startswith("/"):
        return "https://coco-label.com" + url
    return url


def read_json_file(file_path: str) -> list[dict]:
    with open(file_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    if type(data) is not list:
        raise ValueError(f"{file_path} 는 객체 배열(list) 형식이어야 합니다.")

    return data


def empty_result_row(wr_id: int, code: str) -> dict:
    return {
        "wr_id": str(wr_id),
        "code": code,
        "ex_1": "",
        "wr_subject": "",
        "wr_content": "",
        "wr_seo_title": "",
        "wr_file": "0",
        "success_yn": "N",
    }


def load_result_csv(file_path: str, valid_codes: set[str]) -> dict[str, dict]:
    result_map = {}

    if not os.path.exists(file_path):
        return result_map

    with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            code = str(row.get("code", "")).strip()
            if not code:
                continue
            if code not in valid_codes:
                continue

            normalized = {}
            for col in RESULT_COLUMNS:
                normalized[col] = str(row.get(col, "") or "").strip()

            result_map[code] = normalized

    return result_map


def save_result_csv(file_path: str, result_map: dict[str, dict], ordered_codes: list[str]):
    with open(file_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=RESULT_COLUMNS)
        writer.writeheader()

        for code in ordered_codes:
            row = result_map.get(code)
            if not row:
                continue

            write_row = {}
            for col in RESULT_COLUMNS:
                write_row[col] = row.get(col, "")
            writer.writerow(write_row)


def build_detail_url(post_code: str) -> str:
    params = COMMON_QUERY_PARAMS.copy()
    params["code"] = post_code
    return f"{BASE_URL}?{urlencode(params)}"


def request_post_detail(session: requests.Session, post_code: str) -> dict:
    url = build_detail_url(post_code)

    payload = {
        "site_code": "S202410211a92d560f8f0e",
        "unit_code": "u202410216715ffa21a17f",
        "board_code": "",
        "code": post_code,
        "endpoint": "post",
        "type": "post",
    }

    response = session.post(
        url,
        headers=HEADERS,
        data=payload,
        timeout=30,
    )
    response.raise_for_status()
    return response.json()


def parse_subject(subject: str) -> tuple[str, str]:
    subject = (subject or "").strip()
    match = re.match(r"^\[(.*?)\](.*)$", subject)

    if match:
        ex_1 = match.group(1).strip()
        clean_subject = match.group(2).strip()
        return ex_1, clean_subject

    return "", subject


def make_seo_title(subject: str) -> str:
    subject = (subject or "").strip()
    return re.sub(r"\s+", "-", subject)


def extract_first_text_from_content(html: str) -> str:
    html = html or ""
    soup = BeautifulSoup(html, "html.parser")

    for p_tag in soup.find_all("p"):
        text = p_tag.get_text(" ", strip=True)
        if text:
            return text

    return soup.get_text(" ", strip=True)


def extract_image_urls(html: str, max_count: int = 5) -> list[str]:
    html = html or ""
    soup = BeautifulSoup(html, "html.parser")

    image_urls = []
    for img_tag in soup.find_all("img"):
        src = normalize_image_url(img_tag.get("src", ""))
        if not src:
            continue
        image_urls.append(src)
        if len(image_urls) >= max_count:
            break

    return image_urls


def convert_to_rgb(image: Image.Image) -> Image.Image:
    if image.mode in ("RGBA", "LA"):
        background = Image.new("RGB", image.size, (255, 255, 255))
        alpha = image.getchannel("A")
        background.paste(image.convert("RGBA"), mask=alpha)
        return background

    if image.mode != "RGB":
        return image.convert("RGB")

    return image


def resize_image_keep_ratio(image: Image.Image, max_width: int) -> Image.Image:
    width, height = image.size

    if width <= max_width:
        return image

    new_width = max_width
    new_height = int(height * (new_width / width))
    return image.resize((new_width, new_height), Image.LANCZOS)


def download_and_save_image(image_url: str, save_path: str, max_width: int = 600):
    response = requests.get(image_url, headers=IMAGE_HEADERS, timeout=30)
    response.raise_for_status()

    image = Image.open(io.BytesIO(response.content))
    image = ImageOps.exif_transpose(image)
    image = convert_to_rgb(image)
    image = resize_image_keep_ratio(image, max_width)
    image.save(save_path, format="JPEG", quality=95)


def download_image_worker(task: dict) -> dict:
    image_url = task["image_url"]
    save_path = task["save_path"]
    file_name = task["file_name"]
    path_now = task["path_now"]
    bf_no = task["bf_no"]

    try:
        if os.path.exists(save_path):
            return {
                "ok": True,
                "bf_no": bf_no,
                "file_name": file_name,
                "path_origin": os.path.abspath(save_path),
                "path_now": path_now,
                "skipped": True,
            }

        download_and_save_image(
            image_url=image_url,
            save_path=save_path,
            max_width=MAX_IMAGE_WIDTH,
        )

        return {
            "ok": True,
            "bf_no": bf_no,
            "file_name": file_name,
            "path_origin": os.path.abspath(save_path),
            "path_now": path_now,
            "skipped": False,
        }

    except Exception as e:
        return {
            "ok": False,
            "bf_no": bf_no,
            "file_name": file_name,
            "path_origin": os.path.abspath(save_path),
            "path_now": path_now,
            "error": str(e),
        }


def build_image_tasks(wr_id: int, ex_1: str, image_urls: list[str]) -> list[dict]:
    folder_name = safe_folder_name(f"{wr_id}_{ex_1}") if ex_1 else safe_folder_name(str(wr_id))
    folder_path = os.path.join(NOTICE_ROOT_DIR, folder_name)
    ensure_dir(folder_path)

    tasks = []
    for idx, image_url in enumerate(image_urls):
        file_index = idx + 1
        bf_no = idx

        if ex_1:
            file_name = f"{wr_id}_{ex_1}({file_index}).jpg"
        else:
            file_name = f"{wr_id}({file_index}).jpg"

        save_path = os.path.join(folder_path, file_name)
        path_now = f"{NOTICE_ROOT_DIR}/{folder_name}/{file_name}".replace("\\", "/")

        tasks.append({
            "bf_no": bf_no,
            "image_url": image_url,
            "file_name": file_name,
            "save_path": save_path,
            "path_now": path_now,
        })

    return tasks


def process_images_in_threads(tasks: list[dict]) -> tuple[bool, list[dict]]:
    if not tasks:
        return True, []

    results = []
    max_workers = min(MAX_IMAGE_WORKERS, len(tasks))

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_map = {
            executor.submit(download_image_worker, task): task
            for task in tasks
        }

        for future in as_completed(future_map):
            result = future.result()
            results.append(result)

    results.sort(key=lambda x: x["bf_no"])

    all_success = all(item.get("ok") for item in results)
    return all_success, results


def make_write_check_row_from_result(result_row: dict) -> list:
    row_map = {col: "" for col in WRITE_CHECK_COLUMNS}

    wr_id = safe_int(result_row.get("wr_id", 0))
    row_map["wr_id"] = wr_id
    row_map["wr_num"] = f"-{wr_id}"
    row_map["wr_parent"] = wr_id
    row_map["wr_subject"] = result_row.get("wr_subject", "")
    row_map["wr_content"] = result_row.get("wr_content", "")
    row_map["wr_seo_title"] = result_row.get("wr_seo_title", "")
    row_map["wr_file"] = safe_int(result_row.get("wr_file", 0))
    row_map["ex_1"] = result_row.get("ex_1", "")

    return [row_map[col] for col in WRITE_CHECK_COLUMNS]


def make_board_file_row(
        wr_id: int,
        bf_no: int,
        file_name: str,
        path_origin: str,
        path_now: str,
) -> list:
    row_map = {col: "" for col in BOARD_FILE_COLUMNS}

    row_map["bo_table"] = "check"
    row_map["wr_id"] = wr_id
    row_map["bf_no"] = bf_no
    row_map["bf_source"] = file_name
    row_map["bf_file"] = file_name
    row_map["bf_path_origin"] = path_origin
    row_map["bf_path_now"] = path_now

    return [row_map[col] for col in BOARD_FILE_COLUMNS]


def build_excel_from_result_csv(result_map: dict[str, dict], ordered_codes: list[str], output_excel_file: str):
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    ws_write_check = workbook.create_sheet("ey_write_check")
    ws_board_file = workbook.create_sheet("ey_board_file")

    ws_write_check.append(WRITE_CHECK_COLUMNS)
    ws_board_file.append(BOARD_FILE_COLUMNS)

    for code in ordered_codes:
        result_row = result_map.get(code)
        if not result_row:
            continue
        if result_row.get("success_yn") != "Y":
            continue

        ws_write_check.append(make_write_check_row_from_result(result_row))

        wr_id = safe_int(result_row.get("wr_id", 0))
        ex_1 = result_row.get("ex_1", "")
        wr_file = safe_int(result_row.get("wr_file", 0))

        folder_name = safe_folder_name(f"{wr_id}_{ex_1}") if ex_1 else safe_folder_name(str(wr_id))

        for idx in range(wr_file):
            file_index = idx + 1
            bf_no = idx

            if ex_1:
                file_name = f"{wr_id}_{ex_1}({file_index}).jpg"
            else:
                file_name = f"{wr_id}({file_index}).jpg"

            path_now = f"{NOTICE_ROOT_DIR}/{folder_name}/{file_name}".replace("\\", "/")
            path_origin = os.path.abspath(os.path.join(NOTICE_ROOT_DIR, folder_name, file_name))

            ws_board_file.append(
                make_board_file_row(
                    wr_id=wr_id,
                    bf_no=bf_no,
                    file_name=file_name,
                    path_origin=path_origin,
                    path_now=path_now,
                )
            )

    workbook.save(output_excel_file)


def process_one_post(session: requests.Session, wr_id: int, code: str) -> dict:
    detail_json = request_post_detail(session, code)
    data = detail_json.get("data", {})

    subject_raw = data.get("subject", "") or ""
    content_html = data.get("content", "") or ""

    ex_1, wr_subject = parse_subject(subject_raw)
    wr_content = extract_first_text_from_content(content_html)
    wr_seo_title = make_seo_title(wr_subject)

    image_urls = extract_image_urls(content_html, MAX_IMAGE_COUNT)
    image_tasks = build_image_tasks(wr_id, ex_1, image_urls)

    image_success, image_results = process_images_in_threads(image_tasks)

    for image_result in image_results:
        if image_result.get("ok"):
            if image_result.get("skipped"):
                print(f"  [이미지스킵] {image_result.get('path_origin')}")
            else:
                print(f"  [이미지완료] {image_result.get('path_origin')}")
        else:
            print(
                f"  [이미지실패] bf_no={image_result.get('bf_no')} "
                f"file={image_result.get('file_name')} "
                f"error={image_result.get('error')}"
            )

    if not image_success:
        raise RuntimeError("이미지 다운로드 실패")

    result_row = {
        "wr_id": str(wr_id),
        "code": code,
        "ex_1": ex_1,
        "wr_subject": wr_subject,
        "wr_content": wr_content,
        "wr_seo_title": wr_seo_title,
        "wr_file": str(len(image_results)),
        "success_yn": "Y",
    }
    return result_row


def main():
    ensure_dir(NOTICE_ROOT_DIR)

    post_list = read_json_file(LIST_JSON_FILE)
    ordered_posts = []

    for item in post_list:
        code = str(item.get("code", "")).strip()
        if code:
            ordered_posts.append(code)

    valid_codes = set(ordered_posts)
    result_map = load_result_csv(RESULT_CSV_FILE, valid_codes)

    session = requests.Session()

    total_count = len(ordered_posts)
    success_skip_count = 0
    processed_count = 0

    current_wr_id = 0

    for post_item in post_list:
        code = str(post_item.get("code", "")).strip()
        if not code:
            continue

        current_wr_id += 1
        existing_row = result_map.get(code)

        if existing_row and existing_row.get("success_yn") == "Y":
            existing_row["wr_id"] = str(current_wr_id)
            existing_row["code"] = code
            result_map[code] = existing_row
            success_skip_count += 1
            print(f"[스킵] wr_id={current_wr_id}, code={code}, 이미 성공")
            continue

        print(f"[처리시작] wr_id={current_wr_id}, code={code}")

        try:
            result_row = process_one_post(
                session=session,
                wr_id=current_wr_id,
                code=code,
            )
            result_map[code] = result_row
            print(
                f"[성공] wr_id={current_wr_id}, code={code}, "
                f"subject={result_row.get('wr_subject')}, "
                f"wr_file={result_row.get('wr_file')}"
            )

        except Exception as e:
            fail_row = empty_result_row(current_wr_id, code)
            if existing_row:
                for col in RESULT_COLUMNS:
                    if col in existing_row and existing_row.get(col):
                        fail_row[col] = existing_row.get(col, "")
                fail_row["wr_id"] = str(current_wr_id)
                fail_row["code"] = code
                fail_row["success_yn"] = "N"

            result_map[code] = fail_row
            print(f"[실패] wr_id={current_wr_id}, code={code}, error={e}")

        save_result_csv(
            file_path=RESULT_CSV_FILE,
            result_map=result_map,
            ordered_codes=ordered_posts,
        )
        processed_count += 1
        print(f"[CSV저장] {RESULT_CSV_FILE} / 진행={processed_count} / 전체={total_count}")

    build_excel_from_result_csv(
        result_map=result_map,
        ordered_codes=ordered_posts,
        output_excel_file=OUTPUT_EXCEL_FILE,
    )

    success_count = 0
    fail_count = 0
    for code in ordered_posts:
        row = result_map.get(code)
        if not row:
            continue
        if row.get("success_yn") == "Y":
            success_count += 1
        else:
            fail_count += 1

    print(f"[엑셀저장완료] {OUTPUT_EXCEL_FILE}")
    print(
        f"[최종] 전체={total_count}, "
        f"기존성공스킵={success_skip_count}, "
        f"성공={success_count}, 실패={fail_count}"
    )


if __name__ == "__main__":
    main()