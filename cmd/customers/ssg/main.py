# -*- coding: utf-8 -*-
"""
SSG/이마트 상품 상세 수집

입력
- 실행 경로의 엑셀샘플.xlsx 또는 엑셀셈플.xlsx
- A열에 상품 URL
- 헤더가 없어도 됨

출력
- 엑셀샘플_수집결과.xlsx
- 컬럼: 원본 URL / 상품명 / 판매가 / 할인금액 / 행사기간

수집 규칙
1) 일반 할인 상품
   - 판매가: 가격 안내 팝업의 '판매가'
   - 할인금액: 가격 안내 팝업의 '즉시할인가' 또는 '행사판매가'
   - 행사기간: 가격 안내 팝업의 '기간'

2) 1+1 상품
   - 판매가: 상품 판매가
   - 할인금액: '1+1'
   - 행사기간: 1+1 구매혜택 영역의 기간
"""

import re
import sys
import time
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from selenium.common.exceptions import TimeoutException, WebDriverException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

from src.utils.selenium_utils import SeleniumUtils


# =============================================================================
# 설정
# =============================================================================

INPUT_FILE_NAMES = (
    "엑셀샘플.xlsx",
    "엑셀셈플.xlsx",  # 파일명이 이렇게 되어 있어도 동작하도록 처리
)
OUTPUT_FILE = "엑셀샘플_수집결과.xlsx"

# 디버깅할 때는 False 권장.
# 안정적으로 동작 확인 후 True로 바꿔도 됩니다.
HEADLESS = False

PAGE_LOAD_TIMEOUT = 30
ELEMENT_TIMEOUT = 12

# URL 사이에 너무 빠르게 연속 요청하지 않도록 짧게 대기
REQUEST_INTERVAL_SEC = 0.4

# 중간 저장 주기
SAVE_EVERY = 10


# =============================================================================
# 공통 함수
# =============================================================================

def normalize_text(value: Optional[str]) -> str:
    """줄바꿈/연속 공백을 한 칸으로 정리한다."""
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def normalize_price(value: Optional[str]) -> str:
    """
    가격 문자열에서 숫자를 추출하여 '8,980원' 형태로 반환한다.
    값이 없으면 빈 문자열.
    """
    text = normalize_text(value)
    if not text:
        return ""

    match = re.search(r"([0-9][0-9,]*)\s*원?", text)
    if not match:
        return text

    return f"{match.group(1)}원"


def get_text_content(element) -> str:
    """display:none 요소도 읽을 수 있도록 textContent를 사용한다."""
    try:
        return normalize_text(element.get_attribute("textContent"))
    except Exception:
        return ""


def find_input_file() -> Path:
    """실행 경로에서 입력 엑셀 파일을 찾는다."""
    for name in INPUT_FILE_NAMES:
        path = Path(name)
        if path.exists():
            return path

    names = ", ".join(INPUT_FILE_NAMES)
    raise FileNotFoundError(
        f"입력 엑셀 파일을 찾을 수 없습니다. 실행 경로에 다음 중 하나를 넣어주세요: {names}"
    )


def read_urls_from_excel(path: Path) -> List[str]:
    """
    엑셀 A열에서 URL을 읽는다.
    - 헤더가 없어도 됨
    - http:// 또는 https:// 로 시작하는 값만 사용
    - 원래 순서를 유지
    """
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active

    urls: List[str] = []

    try:
        for row in sheet.iter_rows(min_col=1, max_col=1):
            value = row[0].value
            if value is None:
                continue

            url = str(value).strip()
            if url.startswith("http://") or url.startswith("https://"):
                urls.append(url)
    finally:
        workbook.close()

    return urls


# =============================================================================
# 상품 상세 수집
# =============================================================================

def wait_product_name(driver) -> str:
    """상품명을 기다린 뒤 수집한다."""
    element = WebDriverWait(driver, ELEMENT_TIMEOUT).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, ".cdtl_info_tit_txt"))
    )
    return get_text_content(element)


def read_price_guide_from_dom(driver) -> Dict[str, str]:
    """
    가격 안내 레이어를 DOM에서 읽는다.

    실제 화면에서는 ? 버튼을 눌러 팝업이 열리지만,
    제공된 HTML처럼 레이어 내용이 display:none 상태로 DOM에 이미 존재하면
    클릭하지 않고 textContent로 바로 읽는다.

    반환값
    {
        "sale_price": "8,980원",
        "discount_price": "6,286원",
        "period": "2026.08.13까지"
    }
    """
    result = {
        "sale_price": "",
        "discount_price": "",
        "period": "",
    }

    # 가격 안내 tooltip만 정확히 범위를 제한한다.
    wrappers = driver.find_elements(
        By.CSS_SELECTOR,
        ".ssg-tooltip-wrap.cdtl_ly_wrap.ty_sale_detail",
    )

    for wrapper in wrappers:
        try:
            titles = wrapper.find_elements(By.CSS_SELECTOR, ".cdtl_ly_price_guide_tit")
            items = wrapper.find_elements(By.CSS_SELECTOR, ".cdtl_ly_price_guide_item")

            for element in titles:
                text = get_text_content(element)
                if not text:
                    continue

                # 판매가 8,980원
                match = re.match(r"^판매가\s*([0-9][0-9,]*)\s*원", text)
                if match:
                    result["sale_price"] = f"{match.group(1)}원"
                    continue

                # 즉시할인가 6,286원
                # 혹시 행사판매가라는 문구로 나오는 상품도 같이 처리
                match = re.match(
                    r"^(?:즉시할인가|행사판매가)\s*([0-9][0-9,]*)\s*원",
                    text,
                )
                if match:
                    result["discount_price"] = f"{match.group(1)}원"
                    continue

            for element in items:
                text = get_text_content(element)
                if not text:
                    continue

                # 기간 : 2026.08.13까지
                match = re.search(r"기간\s*:\s*(.+)", text)
                if match:
                    result["period"] = normalize_text(match.group(1))
                    break

            if any(result.values()):
                return result

        except Exception:
            continue

    return result


def click_price_guide_and_read(driver) -> Dict[str, str]:
    """
    DOM에서 가격 안내 내용을 바로 읽지 못한 경우의 fallback.
    실제 ? 버튼을 클릭한 뒤 다시 읽는다.
    """
    result = read_price_guide_from_dom(driver)
    if any(result.values()):
        return result

    buttons = driver.find_elements(
        By.CSS_SELECTOR,
        ".ssg-tooltip-wrap.cdtl_ly_wrap.ty_sale_detail "
        "a.ssg-tooltip.cdtl_tooltip",
    )

    for button in buttons:
        try:
            driver.execute_script(
                "arguments[0].scrollIntoView({block: 'center'});",
                button,
            )
            time.sleep(0.1)
            driver.execute_script("arguments[0].click();", button)

            WebDriverWait(driver, 3).until(
                lambda d: any(
                    get_text_content(layer)
                    for layer in d.find_elements(
                        By.CSS_SELECTOR,
                        ".ssg-tooltip-wrap.cdtl_ly_wrap.ty_sale_detail "
                        ".cdtl_ly_price_guide_cont",
                    )
                )
            )

            result = read_price_guide_from_dom(driver)
            if any(result.values()):
                return result

        except Exception:
            continue

    return result


def extract_one_plus_one(driver) -> Tuple[bool, str]:
    """
    구매혜택 영역에서 1+1을 찾는다.

    예)
    <span class="cdtl_benefit">1+1</span>
    ...
    <span class="desc">2026.07.30 ~ 2026.08.13</span>
    """
    benefit_rows = driver.find_elements(By.CSS_SELECTOR, ".cdtl_bene li")

    for row in benefit_rows:
        try:
            badges = row.find_elements(By.CSS_SELECTOR, ".cdtl_benefit")
            if not badges:
                continue

            benefit_text = get_text_content(badges[0])
            if benefit_text != "1+1":
                continue

            descs = row.find_elements(By.CSS_SELECTOR, ".cdtl_benefit_info .desc")
            period = get_text_content(descs[0]) if descs else ""
            return True, period

        except Exception:
            continue

    return False, ""


def extract_page_price(driver) -> str:
    """
    가격 안내 레이어에 '판매가'가 없을 때
    상품 상세 화면에 표시된 실제 판매가를 찾는다.

    현재 SSG/이마트 상세 페이지의 최적가 영역 예시:
    <div class="cdtl_prd_new_price_wrap">
        <span class="cdtl_new_price notranslate">
            <em class="ssg_price">9,980</em>
            <span class="ssg_tx">원</span>
        </span>
    </div>
    """
    selectors = (
        # 현재 SSG/이마트 상품 상세 가격 구조
        ".cdtl_prd_new_price_wrap .cdtl_new_price .ssg_price",
        ".cdtl_prd_new_price_wrap .ssg_price",
        ".cdtl_new_price .ssg_price",

        # 이전 구조 fallback
        ".cdtl_prd_cur_row .ssg_price",
        ".cdtl_prd_price .ssg_price",
        ".cdtl_prd_cur_row em.ssg_price",
        ".cdtl_prd_price em.ssg_price",
    )

    for selector in selectors:
        elements = driver.find_elements(By.CSS_SELECTOR, selector)

        for element in elements:
            text = get_text_content(element)
            if not text:
                continue

            match = re.search(r"([0-9][0-9,]*)", text)
            if match:
                return f"{match.group(1)}원"

    # 마지막 fallback:
    # 최적가 영역 전체 텍스트에서 '원' 앞 숫자를 탐색한다.
    fallback_selectors = (
        ".cdtl_prd_new_price_wrap",
        ".cdtl_optprice_wrap",
        ".cdtl_prd_cur_row",
        ".cdtl_prd_price",
    )

    for selector in fallback_selectors:
        rows = driver.find_elements(By.CSS_SELECTOR, selector)

        for row in rows:
            text = get_text_content(row)
            if not text:
                continue

            matches = re.findall(r"([0-9][0-9,]*)\s*원", text)
            if matches:
                return f"{matches[0]}원"

    return ""


def collect_product(driver, selenium_utils: SeleniumUtils, url: str) -> Dict[str, str]:
    """상품 URL 1개를 수집한다."""
    driver.get(url)

    # pageLoadTimeout이 걸리지 않은 정상 페이지라면 readyState를 짧게 확인
    selenium_utils.wait_ready_state_complete(timeout_sec=8)

    product_name = wait_product_name(driver)

    # 1+1 여부는 구매혜택에서 확인
    is_one_plus_one, one_plus_one_period = extract_one_plus_one(driver)

    # 가격 안내 레이어
    price_guide = click_price_guide_and_read(driver)

    # 판매가는 가격안내의 판매가를 우선 사용
    sale_price = price_guide.get("sale_price", "")
    if not sale_price:
        sale_price = extract_page_price(driver)

    if is_one_plus_one:
        # 합의한 규칙: 1+1 상품이면 할인금액 칸에 '1+1'
        discount_value = "1+1"
        event_period = one_plus_one_period
    else:
        discount_value = price_guide.get("discount_price", "")
        event_period = price_guide.get("period", "")

    return {
        "원본 URL": url,
        "상품명": product_name,
        "판매가": sale_price,
        "할인금액": discount_value,
        "행사기간": event_period,
    }


# =============================================================================
# 결과 엑셀
# =============================================================================

def create_result_workbook() -> Tuple[Workbook, object]:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "수집결과"

    headers = ["원본 URL", "상품명", "판매가", "할인금액", "행사기간"]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="FFD966")
    header_font = Font(bold=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet.freeze_panes = "A2"

    # 보기 편한 기본 너비
    sheet.column_dimensions["A"].width = 95
    sheet.column_dimensions["B"].width = 48
    sheet.column_dimensions["C"].width = 16
    sheet.column_dimensions["D"].width = 16
    sheet.column_dimensions["E"].width = 28

    return workbook, sheet


def append_result(sheet, result: Dict[str, str]) -> None:
    sheet.append(
        [
            result.get("원본 URL", ""),
            result.get("상품명", ""),
            result.get("판매가", ""),
            result.get("할인금액", ""),
            result.get("행사기간", ""),
        ]
    )

    row_no = sheet.max_row
    for cell in sheet[row_no]:
        cell.alignment = Alignment(vertical="top")

    sheet.cell(row=row_no, column=1).alignment = Alignment(
        vertical="top",
        wrap_text=True,
    )
    sheet.cell(row=row_no, column=2).alignment = Alignment(
        vertical="top",
        wrap_text=True,
    )


def save_workbook(workbook: Workbook, sheet, output_path: Path) -> None:
    if sheet.max_row >= 1:
        sheet.auto_filter.ref = f"A1:E{sheet.max_row}"
    workbook.save(output_path)


# =============================================================================
# main
# =============================================================================

def main() -> int:
    input_path = find_input_file()
    output_path = Path(OUTPUT_FILE)

    urls = read_urls_from_excel(input_path)
    if not urls:
        print(f"[종료] {input_path} A열에서 URL을 찾지 못했습니다.")
        return 1

    print(f"[입력] {input_path.resolve()}")
    print(f"[URL 개수] {len(urls)}")
    print(f"[출력] {output_path.resolve()}")
    print()

    workbook, sheet = create_result_workbook()

    selenium_utils = SeleniumUtils(
        headless=HEADLESS,
        debug=False,
    )

    # 네트워크 캡처는 필요 없고, 이미지 로딩은 막아 속도를 높인다.
    selenium_utils.set_capture_options(
        enabled=False,
        block_images=True,
    )

    driver = None

    try:
        driver = selenium_utils.start_driver(
            timeout=PAGE_LOAD_TIMEOUT,
            view_mode="browser",
            window_size=(1500, 950),
        )

        total = len(urls)

        for index, url in enumerate(urls, start=1):
            print(f"[{index}/{total}] {url}")

            try:
                result = collect_product(driver, selenium_utils, url)

                print(f"  상품명   : {result['상품명']}")
                print(f"  판매가   : {result['판매가']}")
                print(f"  할인금액 : {result['할인금액']}")
                print(f"  행사기간 : {result['행사기간']}")

            except TimeoutException as e:
                print(f"  [TIMEOUT] {e}")
                try:
                    driver.execute_script("window.stop();")
                except Exception:
                    pass

                result = {
                    "원본 URL": url,
                    "상품명": "",
                    "판매가": "",
                    "할인금액": "",
                    "행사기간": "",
                }

            except WebDriverException as e:
                print(f"  [WEBDRIVER ERROR] {e}")
                result = {
                    "원본 URL": url,
                    "상품명": "",
                    "판매가": "",
                    "할인금액": "",
                    "행사기간": "",
                }

            except Exception as e:
                print(f"  [ERROR] {type(e).__name__}: {e}")
                result = {
                    "원본 URL": url,
                    "상품명": "",
                    "판매가": "",
                    "할인금액": "",
                    "행사기간": "",
                }

            append_result(sheet, result)

            # 작업 중간에 종료되어도 최대한 결과가 남도록 주기적으로 저장
            if index % SAVE_EVERY == 0:
                try:
                    save_workbook(workbook, sheet, output_path)
                    print(f"  -> 중간 저장 완료 ({index}건)")
                except PermissionError:
                    print(
                        f"  [저장 실패] {OUTPUT_FILE} 파일이 Excel에서 열려 있습니다. "
                        "파일을 닫아주세요."
                    )

            print()
            time.sleep(REQUEST_INTERVAL_SEC)

        save_workbook(workbook, sheet, output_path)

        print("=" * 80)
        print(f"[완료] 총 {total}건 처리")
        print(f"[결과 파일] {output_path.resolve()}")
        print("=" * 80)
        return 0

    finally:
        selenium_utils.quit()


if __name__ == "__main__":
    try:
        sys.exit(main())
    except FileNotFoundError as e:
        print(f"[ERROR] {e}")
        sys.exit(1)
    except PermissionError:
        print(
            f"[ERROR] {OUTPUT_FILE} 파일이 Excel에서 열려 있습니다. "
            "파일을 닫고 다시 실행해주세요."
        )
        sys.exit(1)
