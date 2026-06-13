import csv
import json
import re
import time
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE


# ==================================================
# 설정
# ==================================================
INPUT_CSV  = "yanolja_local_accommodation.csv"
OUTPUT_CSV = "yanolja_local_accommodation_with_seller.csv"
OUTPUT_XLSX = "yanolja_local_accommodation_with_seller.xlsx"

TRPC_URL = "https://nol.yanolja.com/stay/api/trpc/searchHome.home,stay.properties.getFavorite,stay.properties.getSellerInfo"

_print_lock = threading.Lock()


def safe_print(msg):
    with _print_lock:
        print(msg, flush=True)


def create_options():
    return {
        "maxWorkers": 8,      # 멀티쓰레드 개수
        "sleepSec": 0.05,     # 너무 빡세면 0.1~0.3 추천
        "retry": 2,           # 실패시 재시도 횟수
        "timeout": 20,
        "logValueLimit": 0,   # 0이면 전체, 예: 80이면 값이 길 때 80자까지만 로그
    }


# ==================================================
# 최종 결과 컬럼
# ==================================================
FINAL_FIELDNAMES = [
    "아이디",
    "키워드",
    "카테고리",
    "시도",
    "시군구",
    "이름",
    "상호",
    "대표자명",
    "사업자번호",
    "주소",
    "이메일",
    "전화번호",
    "통신판매업 신고번호",
]


# ==================================================
# CSV I/O
# ==================================================
def read_csv_rows(path):
    rows = []
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for r in reader:
            rows.append(r)
    return rows


def write_csv_rows(path, rows, fieldnames):
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for r in rows:
            w.writerow(r)


# ==================================================
# Excel 변환
# ==================================================
def safe_excel_value(value):
    if value is None:
        return ""

    s = str(value)

    # Excel에서 허용하지 않는 제어문자 제거
    s = ILLEGAL_CHARACTERS_RE.sub("", s)

    # Excel 셀 1개 최대 글자 수 제한: 32,767자
    if len(s) > 32767:
        s = s[:32767]

    return s


def csv_to_xlsx(csv_path, xlsx_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "data"

    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)

        row_no = 1
        for row in reader:
            col_no = 1
            for value in row:
                cell = ws.cell(
                    row=row_no,
                    column=col_no,
                    value=safe_excel_value(value)
                )

                # 연락처 010, 사업자번호, ID 등 앞자리 0 보존
                cell.number_format = "@"

                col_no += 1
            row_no += 1

    # 첫 행 고정
    ws.freeze_panes = "A2"

    # 필터 적용
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions

    # 컬럼 너비 자동 조정
    max_width = 60

    for col_idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_idx)
        width = 10

        for row_idx in range(1, min(ws.max_row, 200) + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                width = max(width, len(str(cell_value)) + 2)

        if width > max_width:
            width = max_width

        ws.column_dimensions[column_letter].width = width

    wb.save(xlsx_path)


# ==================================================
# tRPC 요청 준비
# ==================================================
def build_headers(stay_id):
    return {
        "accept": "*/*",
        "accept-language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
        "platform": "Web",
        "referer": "https://nol.yanolja.com/stay/domestic/{}".format(stay_id),
        "user-agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/143.0.0.0 Safari/537.36"
        ),
    }


def build_trpc_params(stay_id):
    payload = {
        "0": {"json": {"verticalCategory": "LOCAL_ACCOMMODATION"}},
        "1": {"json": {"stayId": int(stay_id)}},
        "2": {"json": {"stayId": int(stay_id)}},
    }
    return {
        "batch": 1,
        "input": json.dumps(payload, ensure_ascii=False, separators=(",", ":")),
    }


# ==================================================
# 응답 파싱: tableComponent -> dict(title -> value)
# ==================================================
def normalize_title(title):
    if title is None:
        return ""
    return str(title).replace("\n", " ").strip()


def parse_seller_table(resp_json):
    out = {}

    if not resp_json or not isinstance(resp_json, list):
        return out

    if len(resp_json) < 3:
        return out

    third = resp_json[2] or {}
    result = third.get("result") or {}
    data = result.get("data") or {}
    j = data.get("json")

    if not j or not isinstance(j, list):
        return out

    i = 0
    while i < len(j):
        block = j[i] or {}
        if block.get("type") == "table":
            comps = block.get("tableComponent") or []

            k = 0
            while k < len(comps):
                c = comps[k] or {}
                title = normalize_title(c.get("title"))
                bodys = c.get("bodys") or []

                if isinstance(bodys, list):
                    val = " | ".join([str(x) for x in bodys if x is not None])
                else:
                    val = str(bodys)

                if title:
                    out[title] = val

                k += 1

            break

        i += 1

    return out


# ==================================================
# 네트워크 호출 (숙소 1개)
# ==================================================
def fetch_seller_info(stay_id, opt):
    headers = build_headers(stay_id)
    params = build_trpc_params(stay_id)

    tries = 0
    last_err = None

    with requests.Session() as session:
        while tries <= opt["retry"]:
            try:
                r = session.get(
                    TRPC_URL,
                    headers=headers,
                    params=params,
                    timeout=opt["timeout"]
                )

                if r.status_code != 200:
                    raise RuntimeError("HTTP {}".format(r.status_code))

                data = r.json()
                table_map = parse_seller_table(data)

                return {
                    "ok": True,
                    "stayId": str(stay_id),
                    "table": table_map,
                    "err": ""
                }

            except Exception as e:
                last_err = str(e)
                tries += 1

                if tries <= opt["retry"]:
                    time.sleep(0.3 * tries)

    return {
        "ok": False,
        "stayId": str(stay_id),
        "table": {},
        "err": last_err or "unknown error"
    }


# ==================================================
# 로그
# ==================================================
def shorten_value(v, limit):
    if v is None:
        return ""

    s = str(v)

    if limit and limit > 0 and len(s) > limit:
        return s[:limit] + "...(+" + str(len(s) - limit) + ")"

    return s


def log_done(done_cnt, total_cnt, ok_cnt, stay_id, res, opt):
    table = res.get("table") or {}
    keys = list(table.keys())

    safe_print(
        "[DONE] {}/{} | ok={}/{} | stayId={} | fields={}{}".format(
            done_cnt,
            total_cnt,
            ok_cnt,
            done_cnt,
            stay_id,
            len(keys),
            "" if res.get("ok") else " | err=" + str(res.get("err"))
        )
    )

    if table:
        limit = opt.get("logValueLimit", 0) or 0

        for k in table:
            v = table.get(k, "")
            safe_print("    - {} : {}".format(k, shorten_value(v, limit)))


# ==================================================
# 메인 처리: 멀티스레드
# ==================================================
def merge_rows_with_seller(rows, opt):
    tasks = []
    idx = 0

    while idx < len(rows):
        rid = rows[idx].get("id", "")
        rid = str(rid).strip()

        if rid:
            tasks.append({
                "index": idx,
                "stayId": rid
            })

        idx += 1

    safe_print("[START] input_rows={} / tasks(stayId)={}".format(len(rows), len(tasks)))

    seller_map_by_index = {}

    with ThreadPoolExecutor(max_workers=opt["maxWorkers"]) as ex:
        future_map = {}
        t = 0

        while t < len(tasks):
            it = tasks[t]
            fut = ex.submit(fetch_seller_info, it["stayId"], opt)
            future_map[fut] = it
            t += 1

        done_cnt = 0
        ok_cnt = 0

        for fut in as_completed(future_map):
            meta = future_map[fut]
            stay_id = meta["stayId"]
            index = meta["index"]

            try:
                res = fut.result()
            except Exception as e:
                res = {
                    "ok": False,
                    "stayId": stay_id,
                    "table": {},
                    "err": str(e)
                }

            done_cnt += 1

            if res.get("ok"):
                ok_cnt += 1

            seller_map_by_index[index] = res

            log_done(done_cnt, len(tasks), ok_cnt, stay_id, res, opt)

            if opt["sleepSec"] and opt["sleepSec"] > 0:
                time.sleep(opt["sleepSec"])

    out_rows = []
    all_new_cols = set()

    i = 0
    while i < len(rows):
        base = dict(rows[i])
        res = seller_map_by_index.get(i)

        if res and res.get("table"):
            table = res.get("table") or {}

            for k in table.keys():
                all_new_cols.add(k)
                base[k] = table.get(k, "")

        out_rows.append(base)
        i += 1

    return out_rows, sorted(list(all_new_cols))


# ==================================================
# 최종 한글 컬럼 변환
# ==================================================
def normalize_key(key):
    if key is None:
        return ""

    s = str(key)
    s = s.replace("\r", " ")
    s = s.replace("\n", " ")
    s = re.sub(r"\s+", " ", s)
    s = s.strip()

    return s


def get_by_key(row, key):
    if key in row:
        return row.get(key, "")

    target = normalize_key(key)

    for k in row.keys():
        if normalize_key(k) == target:
            return row.get(k, "")

    return ""


def get_first_value(row, keys):
    for key in keys:
        value = get_by_key(row, key)
        if value is not None and str(value).strip() != "":
            return value

    return ""


def collect_texts_from_json(obj, out):
    if obj is None:
        return

    if isinstance(obj, str):
        s = obj.strip()
        if s:
            out.append(s)
        return

    if isinstance(obj, int) or isinstance(obj, float):
        return

    if isinstance(obj, list):
        for item in obj:
            collect_texts_from_json(item, out)
        return

    if isinstance(obj, dict):
        for v in obj.values():
            collect_texts_from_json(v, out)
        return


def extract_sigungu(location_details_json, sido):
    if not location_details_json:
        return ""

    try:
        obj = json.loads(location_details_json)
    except Exception:
        text = str(location_details_json).strip()

        if sido and text.startswith(sido):
            text = text.replace(sido, "", 1).strip()

        return text

    texts = []
    collect_texts_from_json(obj, texts)

    cleaned = []
    sido_text = str(sido or "").strip()

    for t in texts:
        t = str(t).strip()

        if not t:
            continue

        if sido_text and t == sido_text:
            continue

        if sido_text and t.startswith(sido_text):
            t = t.replace(sido_text, "", 1).strip()

        if t and t not in cleaned:
            cleaned.append(t)

    if not cleaned:
        return ""

    # 보통 첫 번째 상세 지역을 시군구로 사용
    return cleaned[0]


def build_final_output_rows(rows):
    final_rows = []

    for row in rows:
        sido = get_first_value(row, ["regionName"])
        sigungu = extract_sigungu(get_first_value(row, ["locationDetails_json"]), sido)

        final_row = {
            "아이디": get_first_value(row, ["id"]),
            "키워드": get_first_value(row, ["topCategory"]),
            "카테고리": get_first_value(row, ["searchType", "pageName"]),
            "시도": sido,
            "시군구": sigungu,
            "이름": get_first_value(row, ["title"]),
            "상호": get_first_value(row, ["상호명"]),

            "대표자명": get_first_value(row, ["대표자명"]),
            "사업자번호": get_first_value(row, ["사업자등록번호", "사업자번호"]),
            "주소": get_first_value(row, ["사업자주소", "주소"]),
            "이메일": get_first_value(row, ["전자우편주소", "이메일"]),
            "전화번호": get_first_value(row, ["연락처", "전화번호"]),
            "통신판매업 신고번호": get_first_value(row, [
                "통신판매업자 신고번호",
                "통신판매업 신고번호",
                "통신판매업자\n신고번호",
                "통신판매업\n신고번호",
            ]),
        }

        final_rows.append(final_row)

    return final_rows


# ==================================================
# main
# ==================================================
def main():
    opt = create_options()

    rows = read_csv_rows(INPUT_CSV)

    if not rows:
        safe_print("[EXIT] input csv empty: " + INPUT_CSV)
        return

    merged_rows, new_cols = merge_rows_with_seller(rows, opt)

    # 기존 영문 컬럼 + API 원본 컬럼을 최종 한글 컬럼으로 변환
    final_rows = build_final_output_rows(merged_rows)

    write_csv_rows(OUTPUT_CSV, final_rows, FINAL_FIELDNAMES)

    # CSV 저장 후 XLSX도 생성
    csv_to_xlsx(OUTPUT_CSV, OUTPUT_XLSX)

    safe_print("[OK] output_rows={} -> {}".format(len(final_rows), OUTPUT_CSV))
    safe_print("[OK] output_xlsx={} -> {}".format(len(final_rows), OUTPUT_XLSX))
    safe_print("[OK] final_cols={} -> {}".format(len(FINAL_FIELDNAMES), ", ".join(FINAL_FIELDNAMES)))
    safe_print("[OK] original_added_cols={} -> {}".format(len(new_cols), ", ".join(new_cols)))


if __name__ == "__main__":
    main()