# -*- coding: utf-8 -*-

import os
import time
from datetime import datetime
from pprint import pprint
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from selenium.common.exceptions import (
    ElementClickInterceptedException,
    ElementNotInteractableException,
    JavascriptException,
    NoSuchElementException,
    StaleElementReferenceException,
    TimeoutException,
    WebDriverException,
)
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.remote.webdriver import WebDriver
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait

from src.utils.selenium_utils import SeleniumUtils


# =============================================================================
# 기본 설정
# =============================================================================

# 상세 페이지 사이 대기 시간
REQUEST_INTERVAL_SECONDS = 2.0

# URL 한 개당 최대 시도 횟수
MAX_ATTEMPTS = 2

# 페이지 로드 제한 시간(초)
PAGE_LOAD_TIMEOUT = 60

# 첫 번째 시도의 판매자 정보 탐색 시간(초)
FAST_SEARCH_TIMEOUT = 15

# 두 번째 시도의 판매자 정보 탐색 시간(초)
SLOW_SEARCH_TIMEOUT = 35

# 첫 번째 시도의 스크롤 간격(초)
FAST_SCROLL_WAIT = 0.15

# 두 번째 시도의 스크롤 간격(초)
SLOW_SCROLL_WAIT = 0.45

# 모달 대기 시간(초)
FAST_MODAL_TIMEOUT = 10
SLOW_MODAL_TIMEOUT = 20

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ERROR_SCREENSHOT_DIR = os.path.join(BASE_DIR, "error_screenshots")
EXCEL_OUTPUT_DIR = os.path.join(BASE_DIR, "output")

TARGET_FIELDS = {
    "상호",
    "대표자명",
    "주소",
    "전화번호",
    "이메일",
}


# =============================================================================
# URL 목록
# 이전 실행에서 실패했던 URL을 위쪽에 배치
# =============================================================================

URLS: List[str] = [
    # 이전 실행 실패 URL
    "https://www.yeogi.com/domestic-accommodations/4449",
    "https://www.yeogi.com/domestic-accommodations/51530",
    "https://www.yeogi.com/domestic-accommodations/1883",
    "https://www.yeogi.com/domestic-accommodations/5359",

    # 기존 성공 URL
    "https://www.yeogi.com/domestic-accommodations/895",
    "https://www.yeogi.com/domestic-accommodations/55239",
    "https://www.yeogi.com/domestic-accommodations/49680",
    "https://www.yeogi.com/domestic-accommodations/700",
    "https://www.yeogi.com/domestic-accommodations/1695",
    "https://www.yeogi.com/domestic-accommodations/2556",
    "https://www.yeogi.com/domestic-accommodations/2158",
    "https://www.yeogi.com/domestic-accommodations/49848",
    "https://www.yeogi.com/domestic-accommodations/67703",
    "https://www.yeogi.com/domestic-accommodations/49802",
    "https://www.yeogi.com/domestic-accommodations/55127",
    "https://www.yeogi.com/domestic-accommodations/3950",
    "https://www.yeogi.com/domestic-accommodations/67397",
    "https://www.yeogi.com/domestic-accommodations/48916",
    "https://www.yeogi.com/domestic-accommodations/888",
    "https://www.yeogi.com/domestic-accommodations/1887",
]


# =============================================================================
# 사용자 정의 예외
# =============================================================================

class AccessBlockedError(Exception):
    """사이트에서 접속을 제한한 경우."""


class SellerButtonTimeoutError(Exception):
    """판매자 정보 버튼을 제한 시간 안에 찾지 못한 경우."""


class SellerModalTimeoutError(Exception):
    """판매자 정보 모달 또는 표 데이터를 찾지 못한 경우."""


# =============================================================================
# 결과 데이터
# =============================================================================

def create_empty_result(url: str) -> Dict[str, str]:
    return {
        "url": url,
        "상호": "",
        "대표자명": "",
        "주소": "",
        "전화번호": "",
        "이메일": "",
        "수집상태": "FAIL",
        "시도횟수": "0",
        "오류내용": "",
    }


# =============================================================================
# 공통 대기 및 스크롤 정보
# =============================================================================

def wait_document_ready(driver: WebDriver, timeout: int = 15) -> None:
    """DOM이 interactive 또는 complete 상태가 될 때까지 기다린다."""

    try:
        WebDriverWait(driver, timeout).until(
            lambda current_driver: current_driver.execute_script(
                "return document.readyState"
            ) in ("interactive", "complete")
        )

        WebDriverWait(driver, timeout).until(
            lambda current_driver: current_driver.execute_script(
                "return document.body !== null"
            )
        )
    except TimeoutException:
        # 일부 DOM만 생성된 경우에도 이후 탐색을 계속한다.
        pass


def get_page_height(driver: WebDriver) -> int:
    try:
        height = driver.execute_script(
            """
            return Math.max(
                document.body ? document.body.scrollHeight : 0,
                document.documentElement ? document.documentElement.scrollHeight : 0
            );
            """
        )
        return int(height or 0)
    except Exception:
        return 0


def get_scroll_position(driver: WebDriver) -> int:
    try:
        position = driver.execute_script(
            """
            return window.pageYOffset
                || document.documentElement.scrollTop
                || document.body.scrollTop
                || 0;
            """
        )
        return int(position or 0)
    except Exception:
        return 0


# =============================================================================
# 접속 제한 확인
# =============================================================================

def is_access_blocked(driver: WebDriver) -> bool:
    """제목, 본문 및 HTML에서 접속 제한 문구를 검사한다."""

    title = ""
    body_text = ""
    page_source = ""

    try:
        title = driver.title.lower()
    except Exception:
        pass

    try:
        body_text = driver.find_element(By.TAG_NAME, "body").text.lower()
    except Exception:
        pass

    try:
        page_source = driver.page_source.lower()
    except Exception:
        pass

    combined_text = title + "\n" + body_text + "\n" + page_source[:30_000]

    blocked_keywords = [
        "403 forbidden",
        "http 403",
        "access denied",
        "request blocked",
        "temporarily blocked",
        "접근이 제한",
        "접속이 제한",
        "요청이 차단",
        "비정상적인 접근",
        "서비스 이용이 제한",
    ]

    return any(keyword in combined_text for keyword in blocked_keywords)


# =============================================================================
# 판매자 정보 버튼 탐색
# =============================================================================

def get_visible_element(elements: List[WebElement]) -> Optional[WebElement]:
    """중복 요소 중 실제 화면에 표시되는 요소를 반환한다."""

    for element in reversed(elements):
        try:
            if element.is_displayed():
                return element
        except (StaleElementReferenceException, WebDriverException):
            continue

    return None


def find_seller_button_with_selectors(driver: WebDriver) -> Optional[WebElement]:
    selectors = [
        (By.CSS_SELECTOR, '[aria-label="판매자 정보"][role="button"]'),
        (By.CSS_SELECTOR, '[aria-label="판매자 정보"]'),
        (
            By.XPATH,
            "//h2[normalize-space()='판매자 정보']/ancestor::*[@role='button'][1]",
        ),
        (
            By.XPATH,
            "//*[normalize-space()='판매자 정보']/ancestor::*[@role='button'][1]",
        ),
    ]

    for by, selector in selectors:
        try:
            visible_element = get_visible_element(driver.find_elements(by, selector))
            if visible_element is not None:
                return visible_element
        except Exception:
            continue

    return None


def find_seller_button_with_javascript(driver: WebDriver) -> Optional[WebElement]:
    """보이는 판매자 정보 버튼을 JavaScript로 보조 탐색한다."""

    try:
        element = driver.execute_script(
            """
            const candidates = Array.from(
                document.querySelectorAll('[role="button"], [aria-label="판매자 정보"]')
            );

            for (let i = candidates.length - 1; i >= 0; i--) {
                const element = candidates[i];
                const ariaLabel = (element.getAttribute('aria-label') || '').trim();
                const text = (element.innerText || element.textContent || '').trim();
                const rect = element.getBoundingClientRect();
                const style = window.getComputedStyle(element);

                const visible = (
                    rect.width > 0
                    && rect.height > 0
                    && style.display !== 'none'
                    && style.visibility !== 'hidden'
                    && Number(style.opacity || '1') > 0
                );

                if (
                    visible
                    && (
                        ariaLabel === '판매자 정보'
                        || text === '판매자 정보'
                        || text.startsWith('판매자 정보')
                    )
                ) {
                    return element;
                }
            }

            return null;
            """
        )

        if element is not None:
            return element
    except Exception:
        pass

    return None


def find_visible_seller_button(driver: WebDriver) -> Optional[WebElement]:
    button = find_seller_button_with_selectors(driver)
    if button is not None:
        return button

    return find_seller_button_with_javascript(driver)


def scroll_to_bottom_with_lazy_loading(
        driver: WebDriver,
        slow_mode: bool,
) -> Optional[WebElement]:
    """페이지 최하단 이동을 반복하며 지연 렌더링을 유도한다."""

    wait_seconds = 0.8 if slow_mode else 0.35
    max_cycles = 8 if slow_mode else 4
    previous_height = -1
    same_height_count = 0

    for _ in range(max_cycles):
        button = find_visible_seller_button(driver)
        if button is not None:
            return button

        try:
            driver.execute_script(
                """
                window.scrollTo(
                    0,
                    Math.max(
                        document.body ? document.body.scrollHeight : 0,
                        document.documentElement ? document.documentElement.scrollHeight : 0
                    )
                );
                """
            )
        except JavascriptException:
            pass

        time.sleep(wait_seconds)

        button = find_visible_seller_button(driver)
        if button is not None:
            return button

        new_height = get_page_height(driver)

        if new_height == previous_height:
            same_height_count += 1
        else:
            same_height_count = 0

        previous_height = new_height

        if same_height_count >= 2:
            break

    return None


def scroll_page_step_by_step(
        driver: WebDriver,
        slow_mode: bool,
) -> Optional[WebElement]:
    """페이지를 위에서부터 조금씩 내려가며 버튼을 탐색한다."""

    search_timeout = SLOW_SEARCH_TIMEOUT if slow_mode else FAST_SEARCH_TIMEOUT
    scroll_wait = SLOW_SCROLL_WAIT if slow_mode else FAST_SCROLL_WAIT
    scroll_step = 500 if slow_mode else 900
    start_time = time.monotonic()

    driver.execute_script("window.scrollTo(0, 0);")
    time.sleep(0.8 if slow_mode else 0.2)

    while time.monotonic() - start_time < search_timeout:
        button = find_visible_seller_button(driver)
        if button is not None:
            return button

        if is_access_blocked(driver):
            raise AccessBlockedError("사이트에서 접속을 제한했습니다.")

        current_position = get_scroll_position(driver)
        page_height = get_page_height(driver)
        viewport_height = int(
            driver.execute_script("return window.innerHeight || 900;") or 900
        )

        driver.execute_script(
            "window.scrollTo(0, arguments[0]);",
            current_position + scroll_step,
            )
        time.sleep(scroll_wait)

        new_position = get_scroll_position(driver)

        if new_position + viewport_height >= page_height - 100:
            # 최하단에서 지연 렌더링 대기
            time.sleep(1.5 if slow_mode else 0.5)

            button = find_visible_seller_button(driver)
            if button is not None:
                return button

            # IntersectionObserver 반응을 위해 약간 올렸다가 다시 내린다.
            driver.execute_script("window.scrollBy(0, -500);")
            time.sleep(0.5 if slow_mode else 0.2)

            driver.execute_script(
                """
                window.scrollTo(
                    0,
                    Math.max(
                        document.body ? document.body.scrollHeight : 0,
                        document.documentElement ? document.documentElement.scrollHeight : 0
                    )
                );
                """
            )
            time.sleep(1.5 if slow_mode else 0.5)

            button = find_visible_seller_button(driver)
            if button is not None:
                return button

            if get_page_height(driver) <= page_height:
                break

    return None


def find_seller_button(driver: WebDriver, slow_mode: bool) -> WebElement:
    button = find_visible_seller_button(driver)
    if button is not None:
        return button

    button = scroll_to_bottom_with_lazy_loading(driver, slow_mode)
    if button is not None:
        return button

    button = scroll_page_step_by_step(driver, slow_mode)
    if button is not None:
        return button

    driver.execute_script(
        """
        window.scrollTo(
            0,
            Math.max(
                document.body ? document.body.scrollHeight : 0,
                document.documentElement ? document.documentElement.scrollHeight : 0
            )
        );
        """
    )
    time.sleep(2.0 if slow_mode else 0.7)

    button = find_visible_seller_button(driver)
    if button is not None:
        return button

    if is_access_blocked(driver):
        raise AccessBlockedError("사이트에서 접속을 제한했습니다.")

    raise SellerButtonTimeoutError(
        "페이지 하단까지 스크롤했지만 판매자 정보 버튼을 찾지 못했습니다."
    )


# =============================================================================
# 버튼 클릭 및 모달 처리
# =============================================================================

def click_seller_button(driver: WebDriver, button: WebElement) -> None:
    try:
        driver.execute_script(
            """
            arguments[0].scrollIntoView({
                block: 'center',
                inline: 'center'
            });
            """,
            button,
        )
    except Exception:
        pass

    time.sleep(0.3)

    try:
        ActionChains(driver).move_to_element(button).pause(0.1).click().perform()
        return
    except (
            ElementClickInterceptedException,
            ElementNotInteractableException,
            StaleElementReferenceException,
            WebDriverException,
    ):
        pass

    try:
        driver.execute_script("arguments[0].click();", button)
    except Exception as error:
        raise RuntimeError(f"판매자 정보 버튼 클릭 실패: {error}") from error


def get_visible_seller_dialog(driver: WebDriver) -> Optional[WebElement]:
    for selector in ['#modal-wrapper [role="dialog"]', '[role="dialog"]']:
        try:
            dialogs = driver.find_elements(By.CSS_SELECTOR, selector)

            for dialog in reversed(dialogs):
                try:
                    if dialog.is_displayed() and "판매자 정보" in dialog.text.strip():
                        return dialog
                except (StaleElementReferenceException, WebDriverException):
                    continue
        except Exception:
            continue

    return None


def wait_seller_dialog(driver: WebDriver, slow_mode: bool) -> WebElement:
    timeout = SLOW_MODAL_TIMEOUT if slow_mode else FAST_MODAL_TIMEOUT

    try:
        return WebDriverWait(driver, timeout).until(
            lambda current_driver: get_visible_seller_dialog(current_driver)
        )
    except TimeoutException as error:
        raise SellerModalTimeoutError(
            "판매자 정보 버튼은 클릭했지만 모달이 나타나지 않았습니다."
        ) from error


def open_seller_dialog(driver: WebDriver, slow_mode: bool) -> WebElement:
    seller_button = find_seller_button(driver, slow_mode)
    click_seller_button(driver, seller_button)

    try:
        return wait_seller_dialog(driver, slow_mode)
    except SellerModalTimeoutError:
        # 모달이 열리지 않으면 버튼을 다시 찾아 JavaScript 클릭한다.
        time.sleep(1.5 if slow_mode else 0.5)
        seller_button = find_seller_button(driver, slow_mode)
        driver.execute_script("arguments[0].click();", seller_button)
        return wait_seller_dialog(driver, True)


# =============================================================================
# 판매자 정보 추출
# =============================================================================

def extract_seller_info(driver: WebDriver, dialog: WebElement) -> Dict[str, str]:
    def table_rows_loaded(_current_driver: WebDriver) -> bool:
        try:
            return len(dialog.find_elements(By.CSS_SELECTOR, "table tbody tr")) > 0
        except StaleElementReferenceException:
            return False

    try:
        WebDriverWait(driver, 10).until(table_rows_loaded)
    except TimeoutException as error:
        raise SellerModalTimeoutError(
            "판매자 정보 모달은 열렸지만 표 데이터가 로드되지 않았습니다."
        ) from error

    result: Dict[str, str] = {
        "상호": "",
        "대표자명": "",
        "주소": "",
        "전화번호": "",
        "이메일": "",
    }

    for row in dialog.find_elements(By.CSS_SELECTOR, "table tbody tr"):
        try:
            header = row.find_element(By.TAG_NAME, "th").text.strip()
            value = row.find_element(By.TAG_NAME, "td").text.strip()

            if header in TARGET_FIELDS:
                result[header] = value
        except (NoSuchElementException, StaleElementReferenceException):
            continue

    if not result["상호"]:
        raise RuntimeError("판매자 정보 표는 찾았지만 상호 값이 비어 있습니다.")

    return result


# =============================================================================
# URL 한 건 수집 및 재시도
# =============================================================================

def collect_one(
        driver: WebDriver,
        url: str,
        slow_mode: bool,
) -> Dict[str, str]:
    try:
        driver.get(url)
    except TimeoutException:
        try:
            driver.execute_script("window.stop();")
        except Exception:
            pass

    wait_document_ready(driver, timeout=20 if slow_mode else 10)
    time.sleep(2.0 if slow_mode else 0.5)

    if is_access_blocked(driver):
        raise AccessBlockedError("사이트에서 접속을 제한했습니다.")

    dialog = open_seller_dialog(driver, slow_mode)
    return extract_seller_info(driver, dialog)


def collect_with_retry(
        driver: WebDriver,
        url: str,
        current: int,
        total: int,
) -> Dict[str, str]:
    result = create_empty_result(url)
    last_error = ""

    for attempt in range(1, MAX_ATTEMPTS + 1):
        slow_mode = attempt >= 2
        result["시도횟수"] = str(attempt)

        print(
            f"[수집 시도] {current}/{total} | "
            f"{attempt}/{MAX_ATTEMPTS} | "
            f"{'느린 재시도' if slow_mode else '일반 시도'}"
        )

        try:
            result.update(collect_one(driver, url, slow_mode))
            result["수집상태"] = "SUCCESS"
            result["오류내용"] = ""
            return result
        except AccessBlockedError:
            raise
        except Exception as error:
            last_error = f"{type(error).__name__}: {error}"
            print(f"[시도 실패] {current}/{total} | {last_error}")

            if attempt < MAX_ATTEMPTS:
                try:
                    driver.get("about:blank")
                except Exception:
                    pass
                time.sleep(2.0)

    result["오류내용"] = last_error
    return result


# =============================================================================
# 실패 화면 저장
# =============================================================================

def save_error_screenshot(
        driver: WebDriver,
        index: int,
        url: str,
) -> None:
    try:
        os.makedirs(ERROR_SCREENSHOT_DIR, exist_ok=True)
        accommodation_id = url.rstrip("/").split("/")[-1]
        file_path = os.path.join(
            ERROR_SCREENSHOT_DIR,
            f"yeogi_error_{index}_{accommodation_id}.png",
        )
        driver.save_screenshot(file_path)
        print(f"[오류 화면 저장] {file_path}")
    except Exception as error:
        print(f"[오류 화면 저장 실패] {error}")


# =============================================================================
# 콘솔 출력
# =============================================================================

def print_result(
        result: Dict[str, str],
        current: int,
        total: int,
) -> None:
    print("=" * 90)
    print(f"진행      : {current}/{total}")
    print(f"수집상태  : {result.get('수집상태', '')}")
    print(f"시도횟수  : {result.get('시도횟수', '')}")
    print(f"URL       : {result.get('url', '')}")
    print(f"상호      : {result.get('상호', '')}")
    print(f"대표자명  : {result.get('대표자명', '')}")
    print(f"주소      : {result.get('주소', '')}")
    print(f"전화번호  : {result.get('전화번호', '')}")
    print(f"이메일    : {result.get('이메일', '')}")

    if result.get("오류내용"):
        print(f"오류내용  : {result['오류내용']}")

    print("=" * 90)


# =============================================================================
# 고객 전달용 엑셀 저장
# =============================================================================

def save_results_to_excel(results: List[Dict[str, str]]) -> str:
    os.makedirs(EXCEL_OUTPUT_DIR, exist_ok=True)

    now = datetime.now()
    file_path = os.path.join(
        EXCEL_OUTPUT_DIR,
        f"여기어때_판매자정보_{now.strftime('%Y%m%d_%H%M%S')}.xlsx",
    )

    workbook = Workbook()

    # 공통 스타일
    title_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_fill = PatternFill(fill_type="solid", fgColor="5B9BD5")
    summary_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    success_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    fail_fill = PatternFill(fill_type="solid", fgColor="FCE4D6")

    title_font = Font(name="맑은 고딕", size=16, bold=True, color="FFFFFF")
    white_bold_font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
    normal_font = Font(name="맑은 고딕", size=10, color="000000")
    bold_font = Font(name="맑은 고딕", size=10, bold=True)
    link_font = Font(
        name="맑은 고딕",
        size=10,
        color="0563C1",
        underline="single",
    )

    thin_side = Side(style="thin", color="D9E2F3")
    thin_border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side,
    )

    center_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")
    wrap_alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True,
    )

    success_results = [
        item for item in results if item.get("수집상태") == "SUCCESS"
    ]
    fail_results = [
        item for item in results if item.get("수집상태") != "SUCCESS"
    ]

    # -------------------------------------------------------------------------
    # 판매자 정보 시트
    # -------------------------------------------------------------------------
    seller_sheet = workbook.active
    seller_sheet.title = "판매자 정보"

    seller_headers = [
        "번호",
        "상호",
        "대표자명",
        "주소",
        "전화번호",
        "이메일",
        "상세 URL",
    ]
    last_column = len(seller_headers)

    seller_sheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=last_column,
    )
    title_cell = seller_sheet.cell(row=1, column=1, value="여기어때 숙소 판매자 정보")
    title_cell.fill = title_fill
    title_cell.font = title_font
    title_cell.alignment = center_alignment
    seller_sheet.row_dimensions[1].height = 30

    summary_values = [
        ("전체", len(results)),
        ("성공", len(success_results)),
        ("실패", len(fail_results)),
    ]

    summary_column = 1
    for label, value in summary_values:
        label_cell = seller_sheet.cell(row=2, column=summary_column, value=label)
        value_cell = seller_sheet.cell(row=2, column=summary_column + 1, value=value)

        label_cell.fill = summary_fill
        label_cell.font = bold_font
        label_cell.alignment = center_alignment
        label_cell.border = thin_border

        value_cell.font = bold_font
        value_cell.alignment = center_alignment
        value_cell.border = thin_border
        summary_column += 2

    created_cell = seller_sheet.cell(
        row=2,
        column=7,
        value=f"생성일시: {now.strftime('%Y-%m-%d %H:%M:%S')}",
    )
    created_cell.font = normal_font
    created_cell.alignment = center_alignment

    header_row = 4
    for column_index, header in enumerate(seller_headers, start=1):
        cell = seller_sheet.cell(row=header_row, column=column_index, value=header)
        cell.fill = header_fill
        cell.font = white_bold_font
        cell.alignment = center_alignment
        cell.border = thin_border

    data_start_row = header_row + 1

    for row_index, item in enumerate(success_results, start=data_start_row):
        values = [
            row_index - data_start_row + 1,
            item.get("상호", ""),
            item.get("대표자명", ""),
            item.get("주소", ""),
            str(item.get("전화번호", "")),
            item.get("이메일", ""),
            item.get("url", ""),
            ]

        for column_index, value in enumerate(values, start=1):
            cell = seller_sheet.cell(row=row_index, column=column_index, value=value)
            cell.font = normal_font
            cell.border = thin_border
            cell.alignment = left_alignment
            cell.fill = success_fill

        seller_sheet.cell(row=row_index, column=1).alignment = center_alignment

        phone_cell = seller_sheet.cell(row=row_index, column=5)
        phone_cell.number_format = "@"
        phone_cell.alignment = center_alignment

        email_cell = seller_sheet.cell(row=row_index, column=6)
        email_cell.number_format = "@"

        url = str(item.get("url", ""))
        url_cell = seller_sheet.cell(row=row_index, column=7)
        if url:
            url_cell.hyperlink = url
            url_cell.font = link_font

        seller_sheet.cell(row=row_index, column=4).alignment = wrap_alignment
        url_cell.alignment = wrap_alignment
        seller_sheet.row_dimensions[row_index].height = 30

    seller_last_row = max(
        header_row,
        data_start_row + len(success_results) - 1,
        )
    seller_sheet.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(last_column)}{seller_last_row}"
    )
    seller_sheet.freeze_panes = "A5"
    seller_sheet.sheet_view.showGridLines = False

    seller_sheet.column_dimensions["A"].width = 8
    seller_sheet.column_dimensions["B"].width = 24
    seller_sheet.column_dimensions["C"].width = 16
    seller_sheet.column_dimensions["D"].width = 43
    seller_sheet.column_dimensions["E"].width = 18
    seller_sheet.column_dimensions["F"].width = 31
    seller_sheet.column_dimensions["G"].width = 52

    # -------------------------------------------------------------------------
    # 수집 실패 시트
    # -------------------------------------------------------------------------
    fail_sheet = workbook.create_sheet(title="수집 실패")
    fail_headers = ["번호", "상세 URL", "시도횟수", "오류 내용"]
    fail_last_column = len(fail_headers)

    fail_sheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=fail_last_column,
    )
    fail_title_cell = fail_sheet.cell(
        row=1,
        column=1,
        value="판매자 정보 수집 실패 목록",
    )
    fail_title_cell.fill = title_fill
    fail_title_cell.font = title_font
    fail_title_cell.alignment = center_alignment
    fail_sheet.row_dimensions[1].height = 30

    fail_sheet.cell(row=2, column=1, value="실패 건수")
    fail_sheet.cell(row=2, column=2, value=len(fail_results))
    fail_sheet.cell(row=2, column=1).fill = summary_fill
    fail_sheet.cell(row=2, column=1).font = bold_font
    fail_sheet.cell(row=2, column=1).alignment = center_alignment
    fail_sheet.cell(row=2, column=2).font = bold_font
    fail_sheet.cell(row=2, column=2).alignment = center_alignment

    fail_header_row = 4
    for column_index, header in enumerate(fail_headers, start=1):
        cell = fail_sheet.cell(row=fail_header_row, column=column_index, value=header)
        cell.fill = header_fill
        cell.font = white_bold_font
        cell.alignment = center_alignment
        cell.border = thin_border

    fail_data_start_row = fail_header_row + 1

    for row_index, item in enumerate(fail_results, start=fail_data_start_row):
        values = [
            row_index - fail_data_start_row + 1,
            item.get("url", ""),
            item.get("시도횟수", ""),
            item.get("오류내용", ""),
            ]

        for column_index, value in enumerate(values, start=1):
            cell = fail_sheet.cell(row=row_index, column=column_index, value=value)
            cell.font = normal_font
            cell.fill = fail_fill
            cell.border = thin_border
            cell.alignment = wrap_alignment

        fail_sheet.cell(row=row_index, column=1).alignment = center_alignment
        fail_sheet.cell(row=row_index, column=3).alignment = center_alignment

        url = str(item.get("url", ""))
        url_cell = fail_sheet.cell(row=row_index, column=2)
        if url:
            url_cell.hyperlink = url
            url_cell.font = link_font

        fail_sheet.row_dimensions[row_index].height = 42

    fail_last_row = max(
        fail_header_row,
        fail_data_start_row + len(fail_results) - 1,
        )
    fail_sheet.auto_filter.ref = (
        f"A{fail_header_row}:{get_column_letter(fail_last_column)}{fail_last_row}"
    )
    fail_sheet.freeze_panes = "A5"
    fail_sheet.sheet_view.showGridLines = False

    fail_sheet.column_dimensions["A"].width = 8
    fail_sheet.column_dimensions["B"].width = 55
    fail_sheet.column_dimensions["C"].width = 12
    fail_sheet.column_dimensions["D"].width = 70

    if not fail_results:
        no_fail_cell = fail_sheet.cell(
            row=5,
            column=1,
            value="수집 실패 데이터가 없습니다.",
        )
        fail_sheet.merge_cells(
            start_row=5,
            start_column=1,
            end_row=5,
            end_column=fail_last_column,
        )
        no_fail_cell.font = Font(
            name="맑은 고딕",
            size=11,
            bold=True,
            color="548235",
        )
        no_fail_cell.fill = success_fill
        no_fail_cell.alignment = center_alignment
        no_fail_cell.border = thin_border
        fail_sheet.row_dimensions[5].height = 28

    workbook.active = 0
    workbook.save(file_path)
    return file_path


# =============================================================================
# main
# =============================================================================

def main() -> None:
    results: List[Dict[str, str]] = []
    total_count = len(URLS)

    selenium_utils = SeleniumUtils(
        headless=False,
        debug=True,
    )

    # 정상 브라우저와 비슷하게 이미지도 로드한다.
    selenium_utils.set_capture_options(
        enabled=False,
        block_images=False,
    )

    driver: Optional[WebDriver] = None

    print(
        f"전체 수집 시작 | 대상={total_count}개 | "
        f"실패 URL 우선 처리"
    )

    try:
        driver = selenium_utils.start_driver(
            timeout=PAGE_LOAD_TIMEOUT,
            view_mode="browser",
            window_size=(1400, 1000),
        )

        print("[초기 접속] https://www.yeogi.com/")

        try:
            driver.get("https://www.yeogi.com/")
        except TimeoutException:
            try:
                driver.execute_script("window.stop();")
            except Exception:
                pass

        wait_document_ready(driver, timeout=15)

        if is_access_blocked(driver):
            print("[초기 접속 실패] 메인 페이지부터 접속이 제한되었습니다.")
            return

        time.sleep(2.0)

        for index, url in enumerate(URLS, start=1):
            print()
            print(f"[접속 시작] {index}/{total_count} | {url}")

            try:
                result = collect_with_retry(
                    driver=driver,
                    url=url,
                    current=index,
                    total=total_count,
                )
            except AccessBlockedError as error:
                result = create_empty_result(url)
                result["오류내용"] = str(error)
                results.append(result)

                print_result(result, index, total_count)
                print("[수집 중단] 접속 제한이 확인되어 추가 요청을 중단합니다.")
                break

            results.append(result)

            if result["수집상태"] == "FAIL":
                save_error_screenshot(driver, index, url)

            print_result(result, index, total_count)

            if index < total_count:
                time.sleep(REQUEST_INTERVAL_SECONDS)

    except Exception as error:
        print(f"[프로그램 오류] {type(error).__name__}: {error}")

    finally:
        selenium_utils.quit()

    success_count = sum(
        1 for item in results if item.get("수집상태") == "SUCCESS"
    )
    fail_count = len(results) - success_count

    print()
    print("#" * 90)
    print("수집 종료")
    print(f"실행 : {len(results)}개")
    print(f"성공 : {success_count}개")
    print(f"실패 : {fail_count}개")
    print("#" * 90)

    try:
        excel_file_path = save_results_to_excel(results)
        print(f"[엑셀 저장 완료] {excel_file_path}")
    except Exception as error:
        print(f"[엑셀 저장 실패] {type(error).__name__}: {error}")

    print("\n최종 배열:")
    pprint(results, sort_dicts=False, width=220)


if __name__ == "__main__":
    main()
