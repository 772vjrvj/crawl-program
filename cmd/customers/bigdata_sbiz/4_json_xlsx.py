# -*- coding: utf-8 -*-
import json
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook


INPUT_JSON = "gu_grouped_stats.json"
TARGET_XLSX = "상권분석 서울시 치과.xlsx"


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_name(value: Any) -> str:
    return normalize_text(value).replace(" ", "")


def safe_number(value: Any) -> Optional[float]:
    if value is None:
        return None

    text = str(value).strip()
    if not text:
        return None

    text = text.replace(",", "")

    try:
        num = float(text)
        if num.is_integer():
            return int(num)
        return num
    except Exception:
        return None


def load_grouped_json(file_path: Path) -> Dict[str, List[Dict[str, Any]]]:
    with open(file_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    if not isinstance(data, dict):
        raise ValueError("gu_grouped_stats.json 형식이 잘못되었습니다. {\"강남구\": [...]} 형태여야 합니다.")

    return data


def build_dong_row_map(ws) -> Dict[str, int]:
    """
    시트의 B4부터 아래로 읽으면서
    동이름 -> row 번호 매핑
    '전체'가 나오면 종료
    """
    dong_row_map: Dict[str, int] = {}

    for row_idx in range(4, ws.max_row + 1):
        dong_name = normalize_text(ws[f"B{row_idx}"].value)

        if not dong_name:
            continue

        if dong_name == "전체":
            break

        dong_key = normalize_name(dong_name)

        if dong_key not in dong_row_map:
            dong_row_map[dong_key] = row_idx

    return dong_row_map


def calc_dental_ratio_map(items: List[Dict[str, Any]]) -> Dict[str, Optional[float]]:
    """
    해당 구의 모든 치과수 합을 기준으로
    동별 치과 비율 계산
    예: 전체 100, 신사동 10 -> 0.10
    """
    total_dental = 0.0

    for item in items:
        dental_cnt = safe_number(item.get("치과수"))
        if dental_cnt is not None:
            total_dental += float(dental_cnt)

    ratio_map: Dict[str, Optional[float]] = {}

    for item in items:
        dong_key = normalize_name(item.get("읍면동"))
        dental_cnt = safe_number(item.get("치과수"))

        if not dong_key:
            continue

        if dental_cnt is None or total_dental <= 0:
            ratio_map[dong_key] = None
        else:
            ratio_map[dong_key] = float(dental_cnt) / total_dental

    return ratio_map


def write_value(ws, cell_address: str, value: Any) -> None:
    ws[cell_address] = value


def fill_sheet(ws, gu_name: str, items: List[Dict[str, Any]]) -> None:
    dong_row_map = build_dong_row_map(ws)
    ratio_map = calc_dental_ratio_map(items)

    matched_count = 0
    unmatched_count = 0

    for item in items:
        dong_name_raw = item.get("읍면동")
        dong_key = normalize_name(dong_name_raw)

        if not dong_key:
            unmatched_count += 1
            continue

        row_idx = dong_row_map.get(dong_key)
        if row_idx is None:
            print(f"[미매칭] {gu_name} / {dong_name_raw}")
            unmatched_count += 1
            continue

        floating_pop = safe_number(item.get("유동인구"))
        worker_pop = safe_number(item.get("직장인구"))
        resident_pop = safe_number(item.get("주거인구"))
        store_cnt = safe_number(item.get("업소"))
        dental_cnt = safe_number(item.get("치과수"))
        dental_ratio = ratio_map.get(dong_key)

        # C 유동인구
        # D 직장인구
        # E 주거인구
        # F 업소수
        # G 치과수
        # I 치과 비율
        write_value(ws, f"C{row_idx}", floating_pop)
        write_value(ws, f"D{row_idx}", worker_pop)
        write_value(ws, f"E{row_idx}", resident_pop)
        write_value(ws, f"F{row_idx}", store_cnt)
        write_value(ws, f"G{row_idx}", dental_cnt)
        write_value(ws, f"H{row_idx}", dental_ratio)

        # 숫자 포맷
        ws[f"C{row_idx}"].number_format = '#,##0'
        ws[f"D{row_idx}"].number_format = '#,##0'
        ws[f"E{row_idx}"].number_format = '#,##0'
        ws[f"F{row_idx}"].number_format = '#,##0'
        ws[f"G{row_idx}"].number_format = '#,##0'
        ws[f"H{row_idx}"].number_format = '0.00%'

        matched_count += 1

    print(f"[완료] {gu_name}: 매칭 {matched_count}건 / 미매칭 {unmatched_count}건")


def main() -> None:
    base_dir = Path.cwd()

    json_path = base_dir / INPUT_JSON
    xlsx_path = base_dir / TARGET_XLSX

    if not json_path.exists():
        raise FileNotFoundError(f"JSON 파일이 없습니다: {json_path}")

    if not xlsx_path.exists():
        raise FileNotFoundError(f"엑셀 파일이 없습니다: {xlsx_path}")

    grouped_data = load_grouped_json(json_path)

    # 기존 엑셀 열기
    wb = load_workbook(xlsx_path)

    for gu_name, items in grouped_data.items():
        if gu_name not in wb.sheetnames:
            print(f"[시트없음] {gu_name}")
            continue

        ws = wb[gu_name]
        fill_sheet(ws, gu_name, items)

    # 기존 엑셀 파일에 그대로 덮어쓰기
    wb.save(xlsx_path)
    wb.close()

    print(f"\n기존 엑셀 업데이트 완료: {xlsx_path}")


if __name__ == "__main__":
    main()