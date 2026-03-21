# -*- coding: utf-8 -*-
import json
from pathlib import Path

from openpyxl import load_workbook


EXCEL_FILENAME = "상권분석 서울시 치과.xlsx"
OUTPUT_JSON = "data.json"

EXCLUDE_SHEETS = {"구별 요약", "구별 순위", "데이터 현황"}


def normalize_text(value):
    """셀 값을 문자열로 정리"""
    if value is None:
        return ""
    text = str(value).strip()
    return text


def get_merged_or_cell_value(ws, row, col):
    """
    병합셀까지 고려해서 값 가져오기
    - 현재 셀이 병합영역 안에 있으면 병합 시작셀 값을 반환
    - 아니면 일반 셀 값을 반환
    """
    cell = ws.cell(row=row, column=col)

    if cell.value is not None:
        return cell.value

    for merged_range in ws.merged_cells.ranges:
        if (
                merged_range.min_row <= row <= merged_range.max_row
                and merged_range.min_col <= col <= merged_range.max_col
        ):
            return ws.cell(merged_range.min_row, merged_range.min_col).value

    return None


def extract_dong_list_from_sheet(ws, gu_name):
    """
    각 구 시트에서 B4부터 읽다가 '전체' 나오기 전까지 읍면동 추출
    """
    result = []

    for row_idx in range(4, ws.max_row + 1):
        # 기본은 B열 기준
        b_value = normalize_text(get_merged_or_cell_value(ws, row_idx, 2))
        a_value = normalize_text(get_merged_or_cell_value(ws, row_idx, 1))

        # 병합 때문에 B가 비고 A에 값이 있을 수도 있어서 보조 체크
        current_text = b_value or a_value

        if not current_text:
            continue

        # '전체'가 나오면 종료
        if current_text == "전체":
            break

        result.append({
            "시도": "서울시",
            "시군구": gu_name,
            "읍면동": current_text
        })

    return result


def main():
    base_dir = Path.cwd()
    excel_path = base_dir / EXCEL_FILENAME
    output_path = base_dir / OUTPUT_JSON

    if not excel_path.exists():
        raise FileNotFoundError(f"엑셀 파일을 찾을 수 없습니다: {excel_path}")

    wb = load_workbook(excel_path, data_only=True)

    all_data = []

    for sheet_name in wb.sheetnames:
        if sheet_name in EXCLUDE_SHEETS:
            continue

        ws = wb[sheet_name]
        gu_name = sheet_name.strip()

        # 구 시트만 처리하고 싶으면 아래 조건 유지
        # 예: 강남구, 강동구 ... 중랑구
        if not gu_name.endswith("구"):
            continue

        sheet_data = extract_dong_list_from_sheet(ws, gu_name)
        all_data.extend(sheet_data)

        print(f"[완료] {gu_name}: {len(sheet_data)}건")

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(all_data, f, ensure_ascii=False, indent=2)

    print(f"\n총 {len(all_data)}건 저장 완료")
    print(f"저장 경로: {output_path}")


if __name__ == "__main__":
    main()