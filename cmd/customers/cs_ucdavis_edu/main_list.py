from urllib.parse import urljoin

import pandas as pd
from bs4 import BeautifulSoup
from curl_cffi import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment


# BASE_URL = "https://cs.ucdavis.edu"
# OUTPUT_FILE = "cs_ucdavis_people.xlsx"

# BASE_URL = "https://ece.ucdavis.edu"
# OUTPUT_FILE = "ece_ucdavis_people.xlsx"

BASE_URL = "https://che.engineering.ucdavis.edu/"
OUTPUT_FILE = "che_ucdavis_people.xlsx"



def join_unique(values):
    result = []
    for v in values:
        v = (v or "").strip()
        if v and v not in result:
            result.append(v)
    return ", ".join(result)


def decode_cfemail(cfemail: str) -> str:
    if not cfemail:
        return ""

    r = int(cfemail[:2], 16)
    email = ""

    for i in range(2, len(cfemail), 2):
        email += chr(int(cfemail[i:i + 2], 16) ^ r)

    return email


def make_website_text(items):
    lines = []
    for item in items:
        name = (item.get("name") or "").strip()
        url = (item.get("url") or "").strip()
        if name and url:
            lines.append(f"{name} : {url}")
        elif url:
            lines.append(url)
    return "\n".join(lines)


def parse_list_article(article):
    body = article.select_one(".vm-teaser__body")
    if not body:
        return None

    title_a = body.select_one("h3.vm-teaser__title a")
    name = title_a.get_text(" ", strip=True) if title_a else ""

    profile_url = ""
    if title_a and title_a.get("href"):
        profile_url = urljoin(BASE_URL, title_a["href"])

    position_items = [
        li.get_text(" ", strip=True)
        for li in body.select(".vm-teaser__position li")
        if li.get_text(" ", strip=True)
    ]

    role = position_items[0] if len(position_items) >= 1 else ""

    phones = [
        a.get_text(" ", strip=True)
        for a in body.select(".vm-teaser__contact li.icon--phone a")
        if a.get_text(" ", strip=True)
    ]

    emails = []
    for a in body.select(".vm-teaser__contact li.icon--envelope a"):
        href = (a.get("href") or "").strip()

        if href.startswith("mailto:"):
            emails.append(href.replace("mailto:", "").strip())
            continue

        cf_span = a.select_one("span.__cf_email__")
        if cf_span:
            cfemail = (cf_span.get("data-cfemail") or "").strip()
            decoded = decode_cfemail(cfemail)
            if decoded:
                emails.append(decoded)

    website_items = []
    for a in body.select(".vm-teaser__contact li.icon--web a"):
        href = (a.get("href") or "").strip()
        text = a.get_text(" ", strip=True)
        if href:
            website_items.append({
                "name": text,
                "url": href
            })

    location_tag = body.select_one(".text--smaller.u-space-bottom--small .icon.icon--location")
    location = location_tag.get_text(" ", strip=True) if location_tag else ""

    return {
        "Name": name,
        "Role": role,
        "Email": join_unique(emails),
        "Phone": join_unique(phones),
        "Location": location,
        "Research_Areas": "",
        "Lab_Name": "",
        "Website URL": make_website_text(website_items),
        "Profile_URL": profile_url,
    }


def fetch_tags(session, headers, url):
    if not url:
        return ""

    res = session.get(
        url,
        headers=headers,
        impersonate="chrome136",
        timeout=30,
    )

    if res.status_code != 200:
        print(f"[TAGS FAIL] status={res.status_code} url={url}")
        return ""

    soup = BeautifulSoup(res.text, "html.parser")
    tags = [
        a.get_text(" ", strip=True)
        for a in soup.select(".field--name-field-sf-tags a.tags__tag")
        if a.get_text(" ", strip=True)
    ]
    return join_unique(tags)


def crawl():
    session = requests.Session()
    rows = []
    page = 0

    headers = {
        "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
        "accept-language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
        "cache-control": "no-cache",
        "pragma": "no-cache",
        "referer": f"{BASE_URL}/people?page=0",
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/146.0.0.0 Safari/537.36",
    }

    while True:
        list_url = f"{BASE_URL}/people?page={page}"
        headers["referer"] = f"{BASE_URL}/people?page={page - 1}" if page > 0 else f"{BASE_URL}/people?page=0"

        print(f"[PAGE] {page} 요청")
        res = session.get(
            list_url,
            headers=headers,
            impersonate="chrome136",
            timeout=30,
        )

        if res.status_code != 200:
            print(f"[STOP] status={res.status_code} page={page}")
            break

        soup = BeautifulSoup(res.text, "html.parser")
        articles = soup.select("article.node.node--type-sf-person.vm-teaser--grouped.vm-teaser")

        if not articles:
            print(f"[STOP] page={page} 데이터 없음")
            break

        print(f"[PAGE] {page} article={len(articles)}")

        for idx, article in enumerate(articles, 1):
            row = parse_list_article(article)
            if not row:
                continue

            row["Research_Areas"] = fetch_tags(session, headers, row["Profile_URL"])
            rows.append(row)

            print(
                f"[DONE] total={len(rows)} / page={page} / item={idx} / "
                f"Name={row['Name']}"
            )

        page += 1

    return rows


def apply_excel_style(path):
    wb = load_workbook(path)
    ws = wb.active

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    width_map = {
        "A": 24,
        "B": 18,
        "C": 28,
        "D": 18,
        "E": 22,
        "F": 40,
        "G": 16,
        "H": 50,
        "I": 36,
    }

    for col, width in width_map.items():
        ws.column_dimensions[col].width = width

    wb.save(path)


def save_excel(rows):
    df = pd.DataFrame(
        rows,
        columns=[
            "Name",
            "Role",
            "Email",
            "Phone",
            "Location",
            "Research_Areas",
            "Lab_Name",
            "Website URL",
            "Profile_URL",
        ],
    )
    df.to_excel(OUTPUT_FILE, index=False)
    apply_excel_style(OUTPUT_FILE)
    print(f"[SAVE] {OUTPUT_FILE} / 총 {len(df)}건")


def main():
    rows = crawl()
    save_excel(rows)


if __name__ == "__main__":
    main()