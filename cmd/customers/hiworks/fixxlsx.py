# -*- coding: utf-8 -*-
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font


def normalize_excel_folder_link(path_text: str) -> str:
    path_text = (path_text or "").strip()

    # 역슬래시 → 슬래시
    path_text = path_text.replace("\\", "/")

    # hiwork_email부터 시작하는 상대경로 유지
    path_text = path_text.lstrip("./")

    # 폴더는 끝에 / 붙임
    if not path_text.endswith("/"):
        path_text += "/"

    return path_text


def escape_excel_formula_text(text: str) -> str:
    return str(text or "").replace('"', '""')


def fix_excel_links(excel_path: str):
    excel_path = Path(excel_path)

    backup_path = excel_path.with_name(excel_path.stem + "_backup" + excel_path.suffix)
    fixed_path = excel_path.with_name(excel_path.stem + "_fixed" + excel_path.suffix)

    # 원본 백업 먼저
    if not backup_path.exists():
        backup_wb = load_workbook(excel_path)
        backup_wb.save(backup_path)

    wb = load_workbook(excel_path)

    for ws in wb.worksheets:
        header_map = {}

        for cell in ws[1]:
            if cell.value:
                header_map[str(cell.value).strip()] = cell.column

        if "경로" not in header_map:
            continue

        path_col = header_map["경로"]
        fixed_count = 0

        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=path_col)
            path_text = str(cell.value or "").strip()

            if not path_text:
                continue

            # hiwork_email 로 시작하는 경로만 수정
            check_text = path_text.replace("/", "\\").lstrip(".\\")
            if not check_text.startswith("hiwork_email\\"):
                continue

            link = normalize_excel_folder_link(path_text)

            # Office365에서는 cell.hyperlink보다 HYPERLINK 수식이 더 안정적
            safe_link = escape_excel_formula_text(link)
            safe_text = escape_excel_formula_text("폴더 열기")

            cell.value = f'=HYPERLINK("{safe_link}", "{safe_text}")'
            cell.hyperlink = None
            cell.font = Font(color="0563C1", underline="single")

            fixed_count += 1

        print(f"[OK] sheet={ws.title}, 수정={fixed_count}개")

    wb.save(fixed_path)

    print(f"[DONE] 백업파일: {backup_path}")
    print(f"[DONE] 수정파일: {fixed_path}")


if __name__ == "__main__":
    fix_excel_links("hiwork_email_attachment_list.xlsx")