from pathlib import Path
from email import policy
from email.parser import BytesParser
from email.utils import parsedate_to_datetime
from openpyxl import Workbook


INVALID_CHARS = '<>:"/\\|?*'


def clean_name(name: str) -> str:
    name = (name or "").strip()
    if not name:
        return "attachment"
    for ch in INVALID_CHARS:
        name = name.replace(ch, "_")
    return name


def unique_path(path: Path) -> Path:
    if not path.exists():
        return path

    stem = path.stem
    suffix = path.suffix
    parent = path.parent
    no = 1

    while True:
        new_path = parent / f"{stem}_{no}{suffix}"
        if not new_path.exists():
            return new_path
        no += 1


def parse_id_from_stem(stem: str) -> str:
    parts = stem.split("_", 3)
    if len(parts) >= 2:
        return f"{parts[0]}_{parts[1]}"
    return stem


def parse_title_from_stem(stem: str) -> str:
    parts = stem.split("_", 3)
    if len(parts) >= 4:
        return parts[3]
    return stem


def id_to_datetime_text(mail_id: str) -> str:
    # 20240216_184716 -> 2024-02-16 18:47:16
    if len(mail_id) >= 15 and "_" in mail_id:
        y = mail_id[0:4]
        m = mail_id[4:6]
        d = mail_id[6:8]
        hh = mail_id[9:11]
        mm = mail_id[11:13]
        ss = mail_id[13:15]
        return f"{y}-{m}-{d} {hh}:{mm}:{ss}"
    return ""


def parse_mail_date(msg, mail_id: str) -> str:
    raw_date = msg.get("date")
    if raw_date:
        try:
            dt = parsedate_to_datetime(raw_date)
            if dt.tzinfo:
                dt = dt.astimezone()
            return dt.strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            pass

    return id_to_datetime_text(mail_id)


def save_attachments(eml_path: Path, mail_root: Path):
    with open(eml_path, "rb") as f:
        msg = BytesParser(policy=policy.default).parse(f)

    mail_id = parse_id_from_stem(eml_path.stem)
    subject = (msg.get("subject") or "").strip()
    if not subject:
        subject = parse_title_from_stem(eml_path.stem)

    date_text = parse_mail_date(msg, mail_id)

    year = mail_id[0:4]
    month = str(int(mail_id[4:6]))  # 02 -> 2
    save_dir = mail_root / year / month / mail_id

    saved_count = 0

    for part in msg.iter_attachments():
        if part.get_content_disposition() != "attachment":
            continue

        filename = clean_name(part.get_filename())
        data = part.get_payload(decode=True)

        if not filename or data is None:
            continue

        save_dir.mkdir(parents=True, exist_ok=True)

        file_path = unique_path(save_dir / filename)
        with open(file_path, "wb") as fw:
            fw.write(data)

        saved_count += 1

    return {
        "id": mail_id,
        "title": subject,
        "date": date_text,
        "folder_path": save_dir,
        "saved_count": saved_count,
    }


def make_excel(rows, excel_path: Path, base_dir: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "mail"

    ws.append(["ID", "제목", "날짜", "경로"])

    for row in rows:
        rel_path = row["folder_path"].relative_to(base_dir)
        rel_path_text = str(rel_path)

        ws.append([row["id"], row["title"], row["date"], rel_path_text])

        cell = ws.cell(row=ws.max_row, column=4)
        cell.hyperlink = str(row["folder_path"].resolve())
        cell.style = "Hyperlink"

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 40

    wb.save(excel_path)


def main():
    base_dir = Path(__file__).resolve().parent
    mail_root = base_dir / "mail"
    excel_path = base_dir / "mail_attachment_list.xlsx"

    rows = []

    for no in range(1, 90):
        src_dir = mail_root / str(no)
        if not src_dir.exists():
            continue

        eml_files = sorted(src_dir.glob("*.eml"))
        for eml_path in eml_files:
            try:
                info = save_attachments(eml_path, mail_root)

                # 첨부파일 있는 메일만 엑셀에 기록
                if info["saved_count"] > 0:
                    rows.append(info)
                    print(f"[OK] {eml_path.name} -> {info['folder_path']} ({info['saved_count']}개)")
                else:
                    print(f"[SKIP] 첨부파일 없음: {eml_path.name}")

            except Exception as e:
                print(f"[ERROR] {eml_path.name} / {e}")

    rows.sort(key=lambda x: x["id"])
    make_excel(rows, excel_path, base_dir)

    print(f"[DONE] 엑셀 저장 완료: {excel_path}")


if __name__ == "__main__":
    main()