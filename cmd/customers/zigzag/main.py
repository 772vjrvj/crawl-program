from __future__ import annotations

import json
import logging
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Set, Tuple
from zoneinfo import ZoneInfo

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry


LIST_API_URL = "https://api.zigzag.kr/api/2/graphql/GetSearchResult"
PRODUCT_DETAIL_URL = "https://zigzag.kr/catalog/products/{catalog_product_id}"

CATEGORY_NAME = "상의"
CATEGORY_PAGE_URL = (
    "https://zigzag.kr/categories/474"
    "?title=%EC%A0%84%EC%B2%B4&category_id=474&middle_category_id=-1"
)

MAX_PAGES = 1
DETAIL_DELAY_SECONDS = 0.3
REQUEST_TIMEOUT = (10, 30)
OUTPUT_DIR = Path("output")

# 같은 쇼핑몰 상품이 여러 개 있으면 상세 요청은 첫 상품 한 번만 실행합니다.
DEDUPLICATE_SHOPS = True

# 사용자가 전달한 payload 값을 그대로 사용합니다.
# 결과가 전체 카테고리로 섞여 나오면 아래 값을 ["474"]로 시험하세요.
# 팬츠는 일반적으로 ["547"]로 시험할 수 있습니다.
DISPLAY_CATEGORY_ID_LIST = ["-1"]

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/150.0.0.0 Safari/537.36"
)

# 2페이지 요청 시 사용하는 커서 필드 후보입니다.
# 실제 브라우저 요청에서 다른 이름이 확인되면 여기에 추가하면 됩니다.
CURSOR_FIELD_CANDIDATES = (
    "cursor",
    "end_cursor",
    "after",
    "page_cursor",
    "next_cursor",
)

SEARCH_QUERY = """
fragment ProductFields on UxGoodsCardItem {
  catalog_product_id
  shop_id
  shop_name
  title
}

query GetSearchResult($input: SearchResultInput!) {
  search_result(input: $input) {
    end_cursor
    has_next
    ui_item_list {
      __typename

      ...ProductFields

      ... on UxGoodsGroup {
        goods_carousel {
          component_list {
            ...ProductFields
          }
        }
      }

      ... on UxShopRankingCardItem {
        component_list {
          ...ProductFields
        }
      }

      ... on UxGoodsCarousel {
        component_list {
          ...ProductFields
        }
      }
    }
  }
}
"""

logger = logging.getLogger("zigzag-crawler")


def configure_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )


def create_session() -> requests.Session:
    retry = Retry(
        total=3,
        connect=3,
        read=3,
        status=3,
        backoff_factor=0.8,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=frozenset(("GET", "POST")),
        raise_on_status=False,
    )

    adapter = HTTPAdapter(
        max_retries=retry,
        pool_connections=10,
        pool_maxsize=10,
    )

    session = requests.Session()
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    session.headers.update(
        {
            "user-agent": USER_AGENT,
            "accept-language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
        }
    )
    return session


def list_headers(category_page_url: str) -> Dict[str, str]:
    return {
        "accept": "*/*",
        "content-type": "application/json",
        "origin": "https://zigzag.kr",
        "referer": category_page_url,
        "cache-control": "no-cache",
        "pragma": "no-cache",
    }


def detail_headers(category_page_url: str) -> Dict[str, str]:
    return {
        "accept": (
            "text/html,application/xhtml+xml,application/xml;q=0.9,"
            "image/avif,image/webp,image/apng,*/*;q=0.8,"
            "application/signed-exchange;v=b3;q=0.7"
        ),
        "referer": category_page_url,
        "cache-control": "no-cache",
        "pragma": "no-cache",
        "upgrade-insecure-requests": "1",
        "sec-fetch-dest": "document",
        "sec-fetch-mode": "navigate",
        "sec-fetch-site": "same-origin",
        "sec-fetch-user": "?1",
        "sec-ch-ua": (
            '"Not;A=Brand";v="8", "Chromium";v="150", '
            '"Google Chrome";v="150"'
        ),
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": '"Windows"',
    }


def warm_up_session(
        session: requests.Session,
        category_page_url: str,
) -> None:
    """
    카테고리 페이지를 먼저 열어 응답 쿠키가 있으면 Session에 저장합니다.
    실패해도 목록 API 요청은 계속 진행합니다.
    """
    try:
        response = session.get(
            category_page_url,
            headers=detail_headers(category_page_url),
            timeout=REQUEST_TIMEOUT,
        )
        logger.info(
            "세션 준비 | status=%s | cookie_count=%s",
            response.status_code,
            len(session.cookies),
        )
    except requests.RequestException as exc:
        logger.warning("세션 준비 실패, 계속 진행 | error=%s", exc)


def post_search(
        session: requests.Session,
        category_page_url: str,
        search_input: Dict[str, Any],
) -> Dict[str, Any]:
    response = session.post(
        LIST_API_URL,
        headers=list_headers(category_page_url),
        json={
            "query": SEARCH_QUERY,
            "variables": {
                "input": search_input,
            },
        },
        timeout=REQUEST_TIMEOUT,
    )
    response.raise_for_status()

    try:
        body = response.json()
    except requests.JSONDecodeError as exc:
        preview = response.text[:500].replace("\n", " ")
        raise RuntimeError(
            f"목록 응답이 JSON이 아닙니다: {preview}"
        ) from exc

    if body.get("errors"):
        raise RuntimeError(
            "GraphQL 오류: "
            + json.dumps(
                body["errors"],
                ensure_ascii=False,
                separators=(",", ":"),
            )
        )

    return body


def extract_products(value: Any) -> List[Dict[str, str]]:
    """
    ui_item_list의 직접 상품뿐 아니라 carousel/group 안의 상품도 재귀 탐색합니다.
    """
    products: List[Dict[str, str]] = []
    seen_product_ids: Set[str] = set()

    def walk(node: Any) -> None:
        if isinstance(node, dict):
            catalog_product_id = str(
                node.get("catalog_product_id") or ""
            ).strip()
            shop_name = str(node.get("shop_name") or "").strip()

            if catalog_product_id and shop_name:
                if catalog_product_id not in seen_product_ids:
                    seen_product_ids.add(catalog_product_id)
                    products.append(
                        {
                            "catalog_product_id": catalog_product_id,
                            "shop_id": str(
                                node.get("shop_id") or ""
                            ).strip(),
                            "shop_name": shop_name,
                            "title": str(
                                node.get("title") or ""
                            ).strip(),
                        }
                    )

            for child in node.values():
                walk(child)

        elif isinstance(node, list):
            for child in node:
                walk(child)

    walk(value)
    return products


def parse_search_response(
        body: Dict[str, Any],
) -> Tuple[List[Dict[str, str]], Optional[str], bool]:
    search_result = (
        body.get("data", {})
        .get("search_result")
    )

    if not isinstance(search_result, dict):
        raise RuntimeError(
            "응답에서 data.search_result를 찾지 못했습니다"
        )

    products = extract_products(
        search_result.get("ui_item_list", [])
    )
    end_cursor = search_result.get("end_cursor")
    has_next = bool(search_result.get("has_next"))

    return products, end_cursor, has_next


def fetch_first_page(
        session: requests.Session,
        category_page_url: str,
) -> Tuple[List[Dict[str, str]], Optional[str], bool]:
    search_input = {
        "display_category_id_list": DISPLAY_CATEGORY_ID_LIST,
        "page_id": "web_srp_clp_category",
    }

    body = post_search(
        session=session,
        category_page_url=category_page_url,
        search_input=search_input,
    )
    return parse_search_response(body)


def fetch_next_page(
        session: requests.Session,
        category_page_url: str,
        cursor: str,
        previous_product_ids: Set[str],
        resolved_cursor_field: Optional[str],
) -> Tuple[
    List[Dict[str, str]],
    Optional[str],
    bool,
    str,
]:
    """
    정확한 커서 입력 필드명이 제공되지 않았으므로 후보를 순서대로 시험합니다.
    성공한 필드명은 다음 페이지에서도 재사용할 수 있도록 반환합니다.
    """
    candidates: Sequence[str]

    if resolved_cursor_field:
        candidates = (resolved_cursor_field,)
    else:
        candidates = CURSOR_FIELD_CANDIDATES

    last_error: Optional[Exception] = None

    for cursor_field in candidates:
        search_input = {
            "display_category_id_list": DISPLAY_CATEGORY_ID_LIST,
            "page_id": "web_srp_clp_category",
            cursor_field: cursor,
        }

        try:
            body = post_search(
                session=session,
                category_page_url=category_page_url,
                search_input=search_input,
            )
            products, end_cursor, has_next = parse_search_response(body)

            current_product_ids = {
                item["catalog_product_id"]
                for item in products
            }

            if (
                    previous_product_ids
                    and current_product_ids
                    and current_product_ids == previous_product_ids
            ):
                logger.warning(
                    "커서 미적용 추정 | field=%s | 다음 후보 시험",
                    cursor_field,
                )
                continue

            logger.info(
                "페이지 커서 필드 확정 | field=%s",
                cursor_field,
            )
            return (
                products,
                end_cursor,
                has_next,
                cursor_field,
            )

        except Exception as exc:
            last_error = exc
            logger.warning(
                "커서 후보 실패 | field=%s | error=%s",
                cursor_field,
                exc,
            )

    raise RuntimeError(
        "2페이지 커서 필드를 찾지 못했습니다. "
        "개발자도구의 두 번째 목록 요청에서 "
        "variables.input 값을 확인해야 합니다."
    ) from last_error


def collect_product_list(
        session: requests.Session,
        category_name: str,
        category_page_url: str,
        max_pages: int,
) -> List[Dict[str, str]]:
    logger.info(
        "목록 전체 시작 | 카테고리=%s | 목표페이지=%s",
        category_name,
        max_pages,
    )

    all_products: List[Dict[str, str]] = []
    seen_product_ids: Set[str] = set()

    cursor: Optional[str] = None
    has_next = True
    cursor_field: Optional[str] = None
    previous_product_ids: Set[str] = set()

    for page_number in range(1, max_pages + 1):
        logger.info(
            "목록 시작 | page=%s/%s | cursor=%s",
            page_number,
            max_pages,
            "FIRST"
            if page_number == 1
            else f"{str(cursor)[:35]}...",
        )

        if page_number == 1:
            products, cursor, has_next = fetch_first_page(
                session=session,
                category_page_url=category_page_url,
            )
        else:
            if not cursor:
                logger.warning(
                    "end_cursor 없음 | page=%s 요청 중단",
                    page_number,
                )
                break

            (
                products,
                cursor,
                has_next,
                cursor_field,
            ) = fetch_next_page(
                session=session,
                category_page_url=category_page_url,
                cursor=cursor,
                previous_product_ids=previous_product_ids,
                resolved_cursor_field=cursor_field,
            )

        current_product_ids = {
            item["catalog_product_id"]
            for item in products
        }

        added_count = 0

        for product in products:
            product_id = product["catalog_product_id"]

            if product_id in seen_product_ids:
                continue

            seen_product_ids.add(product_id)
            all_products.append(product)
            added_count += 1

        logger.info(
            "목록 완료 | page=%s/%s | 페이지상품=%s | "
            "신규=%s | 누적=%s | has_next=%s | end_cursor=%s",
            page_number,
            max_pages,
            len(products),
            added_count,
            len(all_products),
            has_next,
            f"{str(cursor)[:45]}..." if cursor else "NONE",
        )

        previous_product_ids = current_product_ids

        if not has_next:
            logger.info("다음 페이지 없음 | page=%s", page_number)
            break

    logger.info(
        "목록 전체 완료 | 상품=%s",
        len(all_products),
    )
    return all_products


def deduplicate_by_shop(
        products: Sequence[Dict[str, str]],
) -> List[Dict[str, str]]:
    result: List[Dict[str, str]] = []
    seen_shop_keys: Set[str] = set()

    for product in products:
        shop_key = (
                product.get("shop_id")
                or product.get("shop_name")
                or product["catalog_product_id"]
        )

        if shop_key in seen_shop_keys:
            continue

        seen_shop_keys.add(shop_key)
        result.append(product)

    return result


def find_shop_node(value: Any) -> Optional[Dict[str, Any]]:
    """
    __NEXT_DATA__ 내부 깊이가 바뀌어도 main_contact/business_license를
    가진 shop 객체를 재귀적으로 찾습니다.
    """
    if isinstance(value, dict):
        if (
                isinstance(value.get("main_contact"), dict)
                or isinstance(value.get("business_license"), dict)
        ):
            return value

        for child in value.values():
            found = find_shop_node(child)

            if found is not None:
                return found

    elif isinstance(value, list):
        for child in value:
            found = find_shop_node(child)

            if found is not None:
                return found

    return None


def combine_phone_numbers(*values: Any) -> str:
    result: List[str] = []
    seen: Set[str] = set()

    for raw_value in values:
        value = str(raw_value or "").strip()

        if not value or value in seen:
            continue

        seen.add(value)
        result.append(value)

    return " / ".join(result)


def parse_seller_info(html: str) -> Dict[str, str]:
    soup = BeautifulSoup(html, "html.parser")
    script = soup.select_one("script#__NEXT_DATA__")

    if script is None:
        raise RuntimeError(
            "상세 HTML에서 script#__NEXT_DATA__를 찾지 못했습니다"
        )

    raw_json = script.string or script.get_text(strip=True)

    if not raw_json:
        raise RuntimeError("__NEXT_DATA__ 내용이 비어 있습니다")

    try:
        next_data = json.loads(raw_json)
    except json.JSONDecodeError as exc:
        raise RuntimeError(
            "__NEXT_DATA__ JSON 파싱 실패"
        ) from exc

    shop_node = find_shop_node(next_data)

    if shop_node is None:
        raise RuntimeError(
            "__NEXT_DATA__에서 판매자 정보를 찾지 못했습니다"
        )

    main_contact = shop_node.get("main_contact") or {}
    business_license = (
            shop_node.get("business_license") or {}
    )

    return {
        "company_name": str(
            business_license.get("company_name") or ""
        ).strip(),
        "representative_name": str(
            business_license.get("representative_name") or ""
        ).strip(),
        "email": str(
            main_contact.get("email") or ""
        ).strip(),
        "phone_number": combine_phone_numbers(
            main_contact.get("landline_number"),
            main_contact.get("mobile_number"),
        ),
    }


def fetch_seller_info(
        session: requests.Session,
        category_page_url: str,
        catalog_product_id: str,
) -> Dict[str, str]:
    detail_url = PRODUCT_DETAIL_URL.format(
        catalog_product_id=catalog_product_id
    )

    response = session.get(
        detail_url,
        headers=detail_headers(category_page_url),
        timeout=REQUEST_TIMEOUT,
    )
    response.raise_for_status()

    return parse_seller_info(response.text)


def collect_seller_rows(
        session: requests.Session,
        category_name: str,
        category_page_url: str,
        products: Sequence[Dict[str, str]],
) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    total_count = len(products)
    collection_date = datetime.now(
        ZoneInfo("Asia/Seoul")
    ).strftime("%Y-%m-%d")

    logger.info(
        "상세 전체 시작 | 대상=%s",
        total_count,
    )

    for index, product in enumerate(products, start=1):
        product_id = product["catalog_product_id"]
        shop_name = product["shop_name"]

        logger.info(
            "상세 시작 | %s/%s | 쇼핑몰=%s | product_id=%s",
            index,
            total_count,
            shop_name,
            product_id,
        )

        seller_info = {
            "company_name": "",
            "representative_name": "",
            "email": "",
            "phone_number": "",
        }
        status = "SUCCESS"

        try:
            seller_info = fetch_seller_info(
                session=session,
                category_page_url=category_page_url,
                catalog_product_id=product_id,
            )
        except Exception as exc:
            status = "FAIL"
            logger.exception(
                "상세 실패 | %s/%s | 쇼핑몰=%s | "
                "product_id=%s | error=%s",
                index,
                total_count,
                shop_name,
                product_id,
                exc,
            )

        rows.append(
            {
                "수집일": collection_date,
                "카테고리": category_name,
                "쇼핑몰명": shop_name,
                "업체명": seller_info["company_name"],
                "대표자": seller_info["representative_name"],
                "이메일": seller_info["email"],
                "전화번호": seller_info["phone_number"],
            }
        )

        logger.info(
            "상세 완료 | %s/%s | 상태=%s | 쇼핑몰=%s | "
            "업체명=%s | 대표자=%s | 이메일=%s | 전화번호=%s",
            index,
            total_count,
            status,
            shop_name,
            seller_info["company_name"] or "없음",
            seller_info["representative_name"] or "없음",
            seller_info["email"] or "없음",
            seller_info["phone_number"] or "없음",
            )

        if (
                index < total_count
                and DETAIL_DELAY_SECONDS > 0
        ):
            time.sleep(DETAIL_DELAY_SECONDS)

    logger.info(
        "상세 전체 완료 | 완료=%s",
        len(rows),
    )
    return rows


def save_to_excel(
        rows: Sequence[Dict[str, str]],
        output_dir: Path,
) -> Path:
    columns = [
        "수집일",
        "카테고리",
        "쇼핑몰명",
        "업체명",
        "대표자",
        "이메일",
        "전화번호",
    ]

    output_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now(
        ZoneInfo("Asia/Seoul")
    ).strftime("%Y%m%d_%H%M%S")

    output_path = (
            output_dir
            / f"zigzag_seller_{timestamp}.xlsx"
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "판매자정보"
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = "A1:G1"

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )
    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for column_index, column_name in enumerate(
            columns,
            start=1,
    ):
        cell = worksheet.cell(
            row=1,
            column=column_index,
            value=column_name,
        )
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    for row_index, row in enumerate(rows, start=2):
        for column_index, column_name in enumerate(
                columns,
                start=1,
        ):
            cell = worksheet.cell(
                row=row_index,
                column=column_index,
                value=row.get(column_name, ""),
            )
            cell.alignment = Alignment(
                vertical="center"
            )

    widths = {
        "A": 13,
        "B": 12,
        "C": 22,
        "D": 28,
        "E": 15,
        "F": 32,
        "G": 28,
    }

    for column_letter, width in widths.items():
        worksheet.column_dimensions[
            column_letter
        ].width = width

    worksheet.row_dimensions[1].height = 24
    workbook.save(output_path)

    logger.info(
        "엑셀 저장 완료 | rows=%s | path=%s",
        len(rows),
        output_path.resolve(),
    )
    return output_path


# TODO:
# Google Sheets API 또는 다른 스프레드시트 API로 rows를 전송할 예정입니다.
# def send_to_spreadsheet(rows: Sequence[Dict[str, str]]) -> None:
#     pass


def crawl_zigzag_sellers(
        category_name: str,
        category_page_url: str,
        max_pages: int = 2,
) -> Path:
    session = create_session()

    try:
        warm_up_session(
            session=session,
            category_page_url=category_page_url,
        )

        products = collect_product_list(
            session=session,
            category_name=category_name,
            category_page_url=category_page_url,
            max_pages=max_pages,
        )

        if DEDUPLICATE_SHOPS:
            unique_products = deduplicate_by_shop(products)

            logger.info(
                "쇼핑몰 중복 제거 | 상품=%s | 고유쇼핑몰=%s",
                len(products),
                len(unique_products),
            )
            products = unique_products

        if not products:
            raise RuntimeError(
                "목록에서 수집된 상품이 없습니다"
            )

        rows = collect_seller_rows(
            session=session,
            category_name=category_name,
            category_page_url=category_page_url,
            products=products,
        )

        output_path = save_to_excel(
            rows=rows,
            output_dir=OUTPUT_DIR,
        )

        # TODO: 스프레드시트 연동 시 아래 호출을 활성화합니다.
        # send_to_spreadsheet(rows)

        return output_path

    finally:
        session.close()


def main() -> None:
    configure_logging()

    try:
        output_path = crawl_zigzag_sellers(
            category_name=CATEGORY_NAME,
            category_page_url=CATEGORY_PAGE_URL,
            max_pages=MAX_PAGES,
        )

        logger.info(
            "프로그램 정상 종료 | 결과파일=%s",
            output_path.resolve(),
        )

    except Exception:
        logger.exception("프로그램 비정상 종료")
        raise


if __name__ == "__main__":
    main()
