# -*- coding: utf-8 -*-

import os
import re
import time
from datetime import datetime
from urllib.parse import urljoin, urldefrag, urlparse, parse_qs, urlencode, urlunparse

import pandas as pd
from bs4 import BeautifulSoup

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


BASE_URL = "https://www.swu.ac.kr"

UNIV_URL = "https://www.swu.ac.kr/www/swuniversity.html"
GRAD_TEAM_URL = "https://www.swu.ac.kr/www/swuprea_58.html"
GRAD_INDEX_URL = "https://www.swu.ac.kr/grdindex.do"
INSTITUTE_URL = "https://www.swu.ac.kr/www/swuprea_67.html"
RESEARCH_URL = "https://research.swu.ac.kr/skin/page/about04.html"
TEACHING_CENTER_STAFF_URL = "https://sweet.swu.ac.kr/skin/page/about03.html"

ART_DESIGN_URL = "https://sites.google.com/view/swuart/%EB%8C%80%ED%95%99%EC%86%8C%EA%B0%9C/%EC%9C%84%EC%B9%98-%EB%B0%8F-%EC%97%B0%EB%9D%BD%EC%B2%98"
GLOBAL_TRADE_URL = "https://www.swu.ac.kr/www/globa_1.html"

EMAIL_RE = r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}"
NOTICE_MAX_PAGES = 30


BAD_EMAIL_WORDS = [
    "wixpress.com",
    "sentry.io",
]

COLUMNS = [
    "대분류",
    "중분류",
    "소분류",
    "직위",
    "업무",
    "이름",
    "이메일",
    "URL",
    "홈페이지URL",
]

HTML_CACHE = {}
EMAIL_CACHE = {}


UNIV_EXTRA_CASES = [
    ("인문대학", "AI융합콘텐츠전공", "행정실", "https://meta.swu.ac.kr/notice/z8fd1svc2avb1owuthwh0zyx"),
    ("인문대학", "프랑스문화콘텐츠전공", "행정실", "https://france.swu.ac.kr/bbs/bbs/view.php?bbs_no=16&data_no=517&page_no=1&sub_id="),
    ("인문대학", "독일문화콘텐츠전공", "행정실", "https://german.swu.ac.kr/bbs/bbs/view.php?bbs_no=6&data_no=352&page_no=1&sub_id="),
    ("인문대학", "불어불문학과(*)", "행정실", "https://france.swu.ac.kr/bbs/bbs/view.php?bbs_no=16&data_no=517&page_no=1&sub_id="),
    ("인문대학", "독어불문학과(*)", "행정실", "https://german.swu.ac.kr/bbs/bbs/view.php?bbs_no=6&data_no=349&page_no=1&sub_id="),
    ("인문대학", "문헌정보학과", "행정실", "http://swulis.net/?page_id=51&mod=document&uid=489"),
    ("인문대학", "행정학과", "행정실", "https://swupuad.swu.ac.kr/bbs/bbs/view.php?bbs_no=6&data_no=126&page_no=1&sub_id="),
    ("사회과학대학", "AI뇌융합학습전공", "행정실", "https://edpsy.swu.ac.kr/bbs/bbs/view.php?bbs_no=9&data_no=290&page_no=1&sub_id="),
    ("사회과학대학", "응용심리전공", "행정실", "https://edpsy.swu.ac.kr/bbs/bbs/view.php?bbs_no=9&data_no=290&page_no=1&sub_id="),
    ("사회과학대학", "스포츠운동과학과", "행정실", "https://hms.swu.ac.kr/bbs/bbs/view.php?bbs_no=7&data_no=367&page_no=1&sub_id="),
    ("과학기술융합대학", "바이오헬스융합학과", "행정실", "https://biohealth.qshop.ai/%ED%95%99%EA%B3%BC%EC%9C%84%EC%B9%98"),
    ("과학기술융합대학", "원예생명조경학과", "행정실", "https://swu-hort.swu.ac.kr/bbs/bbs/view.php?bbs_no=6&data_no=582&page_no=1&sub_id="),
    ("과학기술융합대학", "식품생명공학과", "행정실", "https://swufood.swu.ac.kr/board/view?bd_id=bd01&wr_id=39"),
    ("과학기술융합대학", "식품공학전공", "행정실", "https://swufood.swu.ac.kr/board/view?bd_id=bd01&wr_id=39"),
]


def make_driver():
    options = Options()

    # === Selenium 속도 최적화 ===
    options.page_load_strategy = "eager"
    options.add_argument("--headless=new")
    options.add_argument("--disable-gpu")
    options.add_argument("--disable-extensions")
    options.add_argument("--disable-popup-blocking")
    options.add_argument("--disable-notifications")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--no-sandbox")
    options.add_argument("--blink-settings=imagesEnabled=false")
    options.add_argument("--log-level=3")

    prefs = {
        "profile.managed_default_content_settings.images": 2,
        "profile.default_content_setting_values.notifications": 2,
        "profile.default_content_setting_values.popups": 2,
    }
    options.add_experimental_option("prefs", prefs)

    driver = webdriver.Chrome(options=options)
    driver.set_page_load_timeout(12)
    driver.set_script_timeout(5)

    # === Selenium 속도 최적화: 무거운 리소스 차단 ===
    driver.execute_cdp_cmd("Network.enable", {})
    driver.execute_cdp_cmd("Network.setBlockedURLs", {
        "urls": [
            "*.png", "*.jpg", "*.jpeg", "*.gif", "*.webp", "*.svg", "*.ico",
            "*.css",
            "*.woff", "*.woff2", "*.ttf", "*.otf",
            "*google-analytics*", "*googletagmanager*", "*doubleclick*",
        ]
    })

    return driver


def clean_text(text):
    text = str(text or "")
    text = text.replace("\xa0", " ")
    text = text.replace("(*)", "")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def clean_name(text):
    name = clean_text(text)

    for word in ["명예교수", "석좌교수", "특임교수", "부교수", "조교수", "교수"]:
        name = name.replace(word, "")

    name = re.sub(r"\s+", "", name)
    return name.strip()


def abs_url(base, href):
    if not href:
        return ""

    return urldefrag(urljoin(base, href))[0]


def is_bad_email(email):
    email = str(email or "").lower()

    for word in BAD_EMAIL_WORDS:
        if word in email:
            return True

    return False


def uniq(items):
    result = []

    for item in items:
        item = clean_text(item)

        if item and item not in result:
            result.append(item)

    return result


def find_emails(text):
    emails = re.findall(EMAIL_RE, str(text or ""))
    emails = uniq(emails)

    result = []

    for email in emails:
        if not is_bad_email(email):
            result.append(email)

    return result


def first_mailto(tag):
    a = tag.select_one('a[href^="mailto:"]')

    if not a:
        return ""

    email = a.get("href", "").replace("mailto:", "")
    email = email.split("?")[0]
    email = clean_text(email)

    if is_bad_email(email):
        return ""

    return email


def get_html(driver, url, wait_css="", wait_sec=2):
    url = abs_url(BASE_URL, url)

    if not url:
        return ""

    if url in HTML_CACHE:
        return HTML_CACHE[url]

    try:
        driver.get(url)

        if wait_css:
            end_time = time.time() + wait_sec

            while time.time() < end_time:
                if driver.find_elements(By.CSS_SELECTOR, wait_css):
                    break

                time.sleep(0.15)

        html = driver.page_source

        try:
            driver.execute_script("window.stop();")
        except Exception:
            pass

        HTML_CACHE[url] = html
        return html

    except Exception as e:
        print("[WARN] 접속 실패:", url, str(e))

        try:
            html = driver.page_source
            HTML_CACHE[url] = html
            return html
        except Exception:
            return ""


def get_soup(driver, url, wait_css="", wait_sec=2):
    html = get_html(driver, url, wait_css, wait_sec)
    return BeautifulSoup(html, "html.parser")


def make_row(big, mid, small, position, work, name, email, url, home_url):
    return {
        "대분류": clean_text(big),
        "중분류": clean_text(mid),
        "소분류": clean_text(small),
        "직위": clean_text(position),
        "업무": clean_text(work),
        "이름": clean_text(name),
        "이메일": clean_text(email),
        "URL": clean_text(url),
        "홈페이지URL": clean_text(home_url),
    }


def collect_emails_from_url(driver, url):
    url = abs_url(BASE_URL, url)

    if not url:
        return []

    if url in EMAIL_CACHE:
        return EMAIL_CACHE[url]

    html = get_html(driver, url, "body", 3)
    soup = BeautifulSoup(html, "html.parser")
    text = soup.get_text(" ", strip=True) + " " + html

    emails = find_emails(text)
    EMAIL_CACHE[url] = emails

    return emails


def collect_homepage_urls(soup, page_url):
    urls = []

    for a in soup.select("a.btn.btn_xl.btn_blue_gray[href]"):
        text = clean_text(a.get_text(" ", strip=True))

        if "홈페이지" in text:
            urls.append(abs_url(page_url, a.get("href")))

    return uniq(urls)


def find_tab_link(soup, page_url, keyword):
    for a in soup.select(".tabui0 a[href], .tabui1 a[href]"):
        text = clean_text(a.get_text(" ", strip=True))

        if keyword in text:
            return abs_url(page_url, a.get("href"))

    return ""


def collect_professors_from_page(driver, info, prof_url, home_urls):
    rows = []

    if not prof_url:
        return rows

    soup = get_soup(driver, prof_url, "ul.pro_list", 3)

    for li in soup.select("ul.pro_list > li"):
        name_el = li.select_one(".CELL.CELL1 .name")

        if not name_el:
            continue

        name = clean_name(name_el.get_text(" ", strip=True))
        email = first_mailto(li)

        rows.append(make_row(
            info["대분류"],
            info["중분류"],
            info["소분류"],
            "교수",
            "교육·연구",
            name,
            email,
            prof_url,
            ", ".join(home_urls)
        ))

    return rows



def normalize_email(email):
    return clean_text(email).lower()


def is_notice_email_allowed(email, source_url):
    key = normalize_email(email)
    url = str(source_url or "").lower()

    # === 신규 === 기본 공지사항 기준은 @swu 포함 이메일만 사용한다.
    if "@swu" in key:
        return True

    # === 신규 === 바이오헬스융합학과 qshop 게시판 예외
    # 게시판 작성자 이메일이 swu.biohealth@gmail.com 형태로 노출된다.
    if "biohealth.qshop.ai" in url and key.startswith("swu.biohealth@"):
        return True

    return False


def find_notice_candidate_emails(text, professor_emails, source_url=""):
    professor_keys = []

    for email in professor_emails:
        key = normalize_email(email)

        if key and key not in professor_keys:
            professor_keys.append(key)

    result = []

    for email in find_emails(text):
        key = normalize_email(email)

        if not is_notice_email_allowed(email, source_url):
            continue

        # === 신규 === 교수 목록에서 이미 수집한 이메일은 제외한다.
        if key in professor_keys:
            continue

        if key and email not in result:
            result.append(email)

    return result


def find_notice_urls_from_homepage(driver, home_url):
    urls = []
    soup = get_soup(driver, home_url, "body", 3)

    for a in soup.select("a[href]"):
        text = clean_text(a.get_text(" ", strip=True))
        compact_text = text.replace(" ", "")

        # === 신규 === 일반 공지사항 + 언론영상학부 같은 "학부 공지" 메뉴도 포함한다.
        if "공지사항" in text or "학부공지" in compact_text:
            urls.append(abs_url(home_url, a.get("href")))

    return uniq(urls)


def make_notice_page_url(url, page_no):
    parsed = urlparse(url)
    qs = parse_qs(parsed.query)

    if "page_no" in qs:
        qs["page_no"] = [str(page_no)]
    elif "page" in qs:
        qs["page"] = [str(page_no)]
    elif "p" in qs:
        qs["p"] = [str(page_no)]
    elif "qshop.ai" in parsed.netloc:
        qs["page"] = [str(page_no)]
    else:
        qs["page_no"] = [str(page_no)]

    query = urlencode(qs, doseq=True)
    return urlunparse((parsed.scheme, parsed.netloc, parsed.path, parsed.params, query, parsed.fragment))


def make_notice_detail_url(list_url, data_no, page_no):
    parsed = urlparse(list_url)
    path = parsed.path

    if "view.php" not in path:
        path = path.replace("lists.php", "view.php")
        path = path.replace("list.php", "view.php")

    qs = parse_qs(parsed.query)
    qs["data_no"] = [str(data_no)]

    if "page_no" in qs:
        qs["page_no"] = [str(page_no)]
    else:
        qs["page_no"] = [str(page_no)]

    query = urlencode(qs, doseq=True)
    return urlunparse((parsed.scheme, parsed.netloc, path, parsed.params, query, parsed.fragment))


def collect_notice_detail_urls(list_url, soup, page_no):
    urls = []

    for tr in soup.select("table tbody tr"):
        if "secret" in tr.get("class", []):
            continue

        a = tr.select_one("a.btnRead[value]")

        if a:
            data_no = clean_text(a.get("value"))

            if data_no:
                urls.append(make_notice_detail_url(list_url, data_no, page_no))

            continue

        a = tr.select_one("td.subject a[href], .subject a[href], a[href]")

        if a:
            href = a.get("href", "")

            if href and href != "#":
                urls.append(abs_url(list_url, href))

    # === 신규 === 바이오헬스융합학과 qshop 게시판은 table row가 아니라 /board/... 카드 링크 구조다.
    for a in soup.select('a[href*="/board/"]'):
        href = a.get("href", "")

        if not href:
            continue

        if "/write" in href:
            continue

        urls.append(abs_url(list_url, href))

    return uniq(urls)


def collect_notice_emails_from_board(driver, notice_url, professor_emails):
    for page_no in range(1, NOTICE_MAX_PAGES + 1):
        page_url = make_notice_page_url(notice_url, page_no)
        soup = get_soup(driver, page_url, "body", 3)
        detail_urls = collect_notice_detail_urls(page_url, soup, page_no)

        if not detail_urls:
            if page_no == 1:
                area = soup.select_one(".bbs-wrap") or soup
                text = area.get_text(" ", strip=True) + " " + str(area)
                emails = find_notice_candidate_emails(text, professor_emails, page_url)

                if emails:
                    print("    [공지사항 이메일] 발견:", ", ".join(emails))
                    return emails

            break

        print("  [공지사항] page", page_no, "상세", len(detail_urls), "건")

        for detail_url in detail_urls:
            detail_soup = get_soup(driver, detail_url, ".bbs-wrap, body", 3)
            area = detail_soup.select_one(".bbs-wrap") or detail_soup
            text = area.get_text(" ", strip=True) + " " + str(area)
            emails = find_notice_candidate_emails(text, professor_emails, detail_url)

            if emails:
                print("    [공지사항 이메일] 발견:", ", ".join(emails))
                return emails

    return []


def collect_notice_emails_from_homepages(driver, home_urls, professor_emails):
    result = []

    for home_url in home_urls:
        notice_urls = find_notice_urls_from_homepage(driver, home_url)

        for notice_url in notice_urls:
            print("  [공지사항] 확인:", notice_url)
            emails = collect_notice_emails_from_board(driver, notice_url, professor_emails)
            result.extend(emails)

            if emails:
                break

    return uniq(result)

def collect_admin_from_homepages(driver, info, home_urls, professor_emails):
    rows = []

    for home_url in home_urls:
        home_emails = collect_emails_from_url(driver, home_url)
        notice_emails = collect_notice_emails_from_homepages(driver, [home_url], professor_emails)
        emails = uniq(home_emails + notice_emails)

        if emails:
            for email in emails:
                rows.append(make_row(
                    info["대분류"],
                    info["중분류"],
                    "행정실",
                    "",
                    "행정",
                    "",
                    email,
                    info["URL"],
                    home_url
                ))
        else:
            rows.append(make_row(
                info["대분류"],
                info["중분류"],
                "행정실",
                "",
                "행정",
                "",
                "",
                info["URL"],
                home_url
            ))

    return rows


def collect_university_list(driver):
    print("[대학] 목록 수집 시작")

    soup = get_soup(driver, UNIV_URL, "#main .section", 3)
    depts = []

    for section in soup.select("#main .section"):
        title = section.select_one(".titl1, .titl0")

        if not title:
            continue

        big = clean_text(title.get_text(" ", strip=True))
        links = section.select("ul.col_list0 li a[href]")

        if links:
            for a in links:
                name = clean_text(a.get_text(" ", strip=True))
                url = abs_url(UNIV_URL, a.get("href"))

                depts.append({
                    "대분류": big,
                    "중분류": name,
                    "소분류": name,
                    "URL": url,
                })
        else:
            a = title.select_one("a[href]")

            if a:
                name = clean_text(a.get_text(" ", strip=True))
                url = abs_url(UNIV_URL, a.get("href"))

                depts.append({
                    "대분류": name,
                    "중분류": name,
                    "소분류": name,
                    "URL": url,
                })

    print("[대학] 학과 수:", len(depts))
    return depts


def collect_major_tabs(soup, page_url, dept):
    majors = []

    for a in soup.select(".tabui4 a[href]"):
        name = clean_text(a.get_text(" ", strip=True))

        if not name:
            continue

        majors.append({
            "대분류": dept["대분류"],
            "중분류": name,
            "소분류": name,
            "URL": abs_url(page_url, a.get("href")),
        })

    result = []
    keys = set()

    for item in majors:
        key = item["중분류"] + "|" + item["URL"]

        if key not in keys:
            keys.add(key)
            result.append(item)

    return result


def collect_university_data(driver):
    rows = []
    depts = collect_university_list(driver)

    for idx, dept in enumerate(depts, 1):
        print("[대학]", idx, "/", len(depts), dept["대분류"], dept["중분류"])

        soup = get_soup(driver, dept["URL"], "body", 2)
        dept_home_urls = collect_homepage_urls(soup, dept["URL"])

        majors = collect_major_tabs(soup, dept["URL"], dept)

        if not majors:
            majors = [dept]

        for major in majors:
            major_soup = soup

            if major["URL"] != dept["URL"]:
                major_soup = get_soup(driver, major["URL"], "body", 2)

            home_urls = collect_homepage_urls(major_soup, major["URL"])

            if not home_urls:
                home_urls = dept_home_urls

            prof_url = find_tab_link(major_soup, major["URL"], "교수진")
            professor_rows = []
            professor_emails = []

            if prof_url:
                professor_rows = collect_professors_from_page(driver, major, prof_url, home_urls)
                rows.extend(professor_rows)

            for professor_row in professor_rows:
                professor_email = professor_row.get("이메일", "")

                if professor_email:
                    professor_emails.append(professor_email)

            rows.extend(collect_admin_from_homepages(driver, major, home_urls, professor_emails))

    rows.extend(collect_global_trade_extra(driver))
    rows.extend(collect_teaching_center(driver))
    rows.extend(collect_university_extra_cases(driver))
    rows.extend(collect_art_design_extra(driver))

    return rows


def collect_global_trade_extra(driver):
    print("[대학 예외] 글로벌통상학부 수집 시작")

    rows = []
    soup = get_soup(driver, GLOBAL_TRADE_URL, "#main", 3)
    text = soup.get_text(" ", strip=True) + " " + str(soup)
    emails = find_emails(text)

    if not emails:
        emails = [""]

    for email in emails:
        rows.append(make_row(
            "사회과학대학",
            "글로벌통상학부",
            "행정실",
            "",
            "행정",
            "",
            email,
            GLOBAL_TRADE_URL,
            ""
        ))

    return rows


def collect_teaching_center(driver):
    print("[대학 예외] 교직지원센터 수집 시작")

    rows = []
    soup = get_soup(driver, TEACHING_CENTER_STAFF_URL, ".se2.col-md-12", 4)
    area = soup.select_one(".se2.col-md-12")

    if not area:
        return rows

    for caption in area.select(".row.profile-wrap .caption"):
        name_el = caption.select_one("p.name")
        work_el = caption.select_one("p.area")
        email_el = caption.select_one("p.email")

        name_text = clean_text(name_el.get_text(" ", strip=True)) if name_el else ""
        work = clean_text(work_el.get_text("\n", strip=True)) if work_el else ""
        email_text = clean_text(email_el.get_text(" ", strip=True)) if email_el else ""

        parts = name_text.split()
        name = ""
        position = ""

        if len(parts) >= 2:
            position = parts[-1]
            name = " ".join(parts[:-1])
        else:
            name = name_text

        emails = find_emails(email_text)

        if not emails:
            emails = [""]

        for email in emails:
            rows.append(make_row(
                "교양대학",
                "교양대학",
                "교직지원센터",
                position,
                work,
                name,
                email,
                TEACHING_CENTER_STAFF_URL,
                TEACHING_CENTER_STAFF_URL
            ))

    return rows


def collect_university_extra_cases(driver):
    print("[대학 예외] 일반 예외 수집 시작")

    rows = []

    for idx, item in enumerate(UNIV_EXTRA_CASES, 1):
        big, mid, small, url = item
        print("[대학 예외]", idx, "/", len(UNIV_EXTRA_CASES), big, mid)

        emails = collect_emails_from_url(driver, url)

        if not emails:
            emails = [""]

        for email in emails:
            rows.append(make_row(
                big,
                mid,
                small,
                "",
                "행정",
                "",
                email,
                url,
                ""
            ))

    return rows


def map_art_design(text):
    if "조형" in text:
        return "아트앤디자인스쿨", "조형연구소", "행정실"

    if "현대미술" in text:
        return "아트앤디자인스쿨", "현대미술전공", "행정실"

    if "공예" in text:
        return "아트앤디자인스쿨", "공예_컬렉터블디자인전공", "행정실"

    if "시각디자인" in text:
        return "아트앤디자인스쿨", "시각디자인전공", "행정실"

    if "첨단미디어" in text:
        return "아트앤디자인스쿨", "첨단미디어디자인전공", "행정실"

    return "아트앤디자인스쿨", text, "행정실"


def collect_art_design_extra(driver):
    print("[대학 예외] 아트앤디자인스쿨 수집 시작")

    rows = []
    soup = get_soup(driver, ART_DESIGN_URL, "ul.n8H08c.UVNKR", 8)

    for ul in soup.select("ul.n8H08c.UVNKR"):
        for li in ul.select("li"):
            name_parts = []
            emails = []

            for span in li.select("span"):
                style = span.get("style", "")
                style_key = style.replace(" ", "").lower()
                text = clean_text(span.get_text(" ", strip=True))

                if "font-size:13.999999999999998pt" in style_key:
                    if text:
                        name_parts.append(text)

                if "color:#4d81d7" in style_key:
                    if text and not is_bad_email(text):
                        emails.append(text)

            office_text = clean_text(" ".join(name_parts))
            emails = uniq(emails)

            if not office_text:
                continue

            big, mid, small = map_art_design(office_text)

            if not emails:
                emails = [""]

            for email in emails:
                rows.append(make_row(
                    big,
                    mid,
                    small,
                    "",
                    "행정",
                    "",
                    email,
                    ART_DESIGN_URL,
                    ""
                ))

    return rows


def collect_grad_team(driver):
    print("[대학원 교학팀] 수집 시작")

    rows = []
    soup = get_soup(driver, GRAD_TEAM_URL, "div.table0.center", 3)

    for tr in soup.select("div.table0.center tbody tr"):
        tds = tr.select("td")

        if len(tds) < 4:
            continue

        position = clean_text(tds[0].get_text(" ", strip=True))
        name = clean_text(tds[1].get_text(" ", strip=True))
        email = first_mailto(tds[3])

        if not email:
            emails = find_emails(tds[3].get_text(" ", strip=True))
            email = emails[0] if emails else ""

        work = ""

        if len(tds) >= 5:
            work = clean_text(tds[4].get_text("\n", strip=True))

        rows.append(make_row(
            "대학원",
            "대학원",
            "대학원 교학팀",
            position,
            work,
            name,
            email,
            GRAD_TEAM_URL,
            ""
        ))

    return rows


def collect_grad_sites(driver):
    print("[대학원] 사이트 목록 수집 시작")

    soup = get_soup(driver, GRAD_INDEX_URL, ".site_box", 3)
    sites = []

    for a in soup.select(".site_box a[href]"):
        href = a.get("href", "")
        name = clean_text(a.get_text(" ", strip=True))

        if not name:
            continue

        if "index.do" in href:
            continue

        sites.append({
            "name": name,
            "url": abs_url(GRAD_INDEX_URL, href),
        })

    print("[대학원] 사이트 수:", len(sites))
    return sites


def direct_ul(tag):
    for child in tag.find_all("ul", recursive=False):
        return child

    return None


def first_direct_text(tag):
    for child in tag.find_all(["a", "span"], recursive=False):
        text = clean_text(child.get_text(" ", strip=True))

        if text:
            return text

    return clean_text(tag.get_text(" ", strip=True))


def first_direct_link(tag, base_url):
    for child in tag.find_all("a", recursive=False):
        href = child.get("href")

        if href:
            return abs_url(base_url, href)

    a = tag.select_one("a[href]")

    if a:
        return abs_url(base_url, a.get("href"))

    return ""


def collect_grad_depts(driver, site):
    soup = get_soup(driver, site["url"], "ul#gnb", 3)
    depts = []

    menu_li = None

    for li in soup.select("div.section.gnb_sec ul#gnb > li"):
        title = first_direct_text(li)

        if title in ["학과안내", "학과소개"]:
            menu_li = li
            break

    if not menu_li:
        return depts

    menu_ul = direct_ul(menu_li)

    if not menu_ul:
        return depts

    for li in menu_ul.find_all("li", recursive=False):
        child_ul = direct_ul(li)
        title = first_direct_text(li)

        if child_ul:
            series = title

            for sub_li in child_ul.find_all("li", recursive=False):
                dept_name = first_direct_text(sub_li)
                dept_url = first_direct_link(sub_li, site["url"])

                if dept_name and dept_url:
                    depts.append({
                        "대분류": site["name"] + " / " + series,
                        "중분류": dept_name,
                        "소분류": dept_name,
                        "URL": dept_url,
                    })
        else:
            dept_name = title
            dept_url = first_direct_link(li, site["url"])

            if dept_name and dept_url:
                depts.append({
                    "대분류": site["name"],
                    "중분류": dept_name,
                    "소분류": dept_name,
                    "URL": dept_url,
                })

    return depts


def collect_grad_data(driver):
    rows = []

    rows.extend(collect_grad_team(driver))

    sites = collect_grad_sites(driver)

    for site_idx, site in enumerate(sites, 1):
        print("[대학원]", site_idx, "/", len(sites), site["name"])

        depts = collect_grad_depts(driver, site)

        for dept_idx, dept in enumerate(depts, 1):
            print("  [대학원 학과]", dept_idx, "/", len(depts), dept["중분류"])

            soup = get_soup(driver, dept["URL"], "body", 2)
            prof_url = find_tab_link(soup, dept["URL"], "교수진 소개")

            if not prof_url:
                prof_url = find_tab_link(soup, dept["URL"], "교수진")

            if prof_url:
                rows.extend(collect_professors_from_page(driver, dept, prof_url, []))

    return rows


def collect_people_table(soup, big, mid, small, url, home_url):
    rows = []

    for tr in soup.select("div.table0.center tbody tr"):
        tds = tr.select("td")

        if len(tds) < 4:
            continue

        position = clean_text(tds[0].get_text(" ", strip=True))
        name = clean_text(tds[1].get_text(" ", strip=True))
        email = first_mailto(tds[3])

        if not email:
            emails = find_emails(tds[3].get_text(" ", strip=True))
            email = emails[0] if emails else ""

        work = ""

        if len(tds) >= 5:
            work = clean_text(tds[4].get_text("\n", strip=True))

        rows.append(make_row(
            big,
            mid,
            small,
            position,
            work,
            name,
            email,
            url,
            home_url
        ))

    return rows


def collect_institute_data(driver):
    print("[부설연구기관] 목록 수집 시작")

    rows = []
    soup = get_soup(driver, INSTITUTE_URL, "div.tabui0", 3)
    institutes = []

    for a in soup.select("div.tabui0 a[href]"):
        name = clean_text(a.get_text(" ", strip=True))
        url = abs_url(INSTITUTE_URL, a.get("href"))

        if name and url:
            institutes.append({
                "name": name,
                "url": url,
            })

    print("[부설연구기관] 기관 수:", len(institutes))

    for idx, item in enumerate(institutes, 1):
        print("[부설연구기관]", idx, "/", len(institutes), item["name"])

        soup = get_soup(driver, item["url"], "body", 2)
        member_url = find_tab_link(soup, item["url"], "구성원")

        if not member_url:
            continue

        member_soup = get_soup(driver, member_url, "body", 2)

        info = {
            "대분류": "부설연구기관",
            "중분류": item["name"],
            "소분류": item["name"],
            "URL": member_url,
        }

        rows.extend(collect_professors_from_page(driver, info, member_url, []))
        rows.extend(collect_people_table(
            member_soup,
            "부설연구기관",
            item["name"],
            item["name"],
            member_url,
            ""
        ))

    return rows


def collect_research_data(driver):
    print("[산학협력단] 수집 시작")

    rows = []
    soup = get_soup(driver, RESEARCH_URL, ".table-wrap", 3)

    for wrap in soup.select(".table-wrap"):
        title_el = wrap.select_one(".wrap-title h4")
        small = clean_text(title_el.get_text(" ", strip=True)) if title_el else "행정실"

        table = wrap.select_one(".wrap-content table.table.table-bordered.table-email")

        if not table:
            continue

        for tr in table.select("tbody tr"):
            tds = tr.select("td")

            if len(tds) < 4:
                continue

            position = clean_text(tds[0].get_text(" ", strip=True))
            name = clean_text(tds[1].get_text(" ", strip=True))
            email = first_mailto(tds[3])

            if not email:
                emails = find_emails(tds[3].get_text(" ", strip=True))
                email = emails[0] if emails else ""

            work = ""

            if len(tds) >= 5:
                work = clean_text(tds[4].get_text("\n", strip=True))

            rows.append(make_row(
                "산학협력단",
                "산학협력단",
                small,
                position,
                work,
                name,
                email,
                RESEARCH_URL,
                ""
            ))

    return rows


def format_excel(filename):
    wb = load_workbook(filename)
    ws = wb.active

    ws.freeze_panes = "A2"

    widths = {
        "A": 18,
        "B": 26,
        "C": 26,
        "D": 14,
        "E": 34,
        "F": 14,
        "G": 30,
        "H": 55,
        "I": 55,
    }

    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = widths.get(letter, 20)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

    wb.save(filename)


def save_excels(rows):
    ts = datetime.now().strftime("%Y%m%d%H%M%S")

    raw_file = "SWU_EMAIL_RESULT_원본_" + ts + ".xlsx"
    clean_file = "SWU_EMAIL_RESULT_고객용_" + ts + ".xlsx"

    df = pd.DataFrame(rows, columns=COLUMNS)
    df.to_excel(raw_file, index=False)

    clean_df = df[df["이메일"].astype(str).str.strip() != ""].copy()
    clean_df["이메일_KEY"] = clean_df["이메일"].astype(str).str.lower().str.strip()
    clean_df = clean_df.drop_duplicates("이메일_KEY", keep="first")
    clean_df = clean_df.drop(columns=["이메일_KEY"])
    clean_df.to_excel(clean_file, index=False)

    format_excel(raw_file)
    format_excel(clean_file)

    print("[완료] 원본 엑셀:", os.path.abspath(raw_file))
    print("[완료] 고객용 엑셀:", os.path.abspath(clean_file))


def main():
    rows = []
    driver = make_driver()

    try:
        rows.extend(collect_university_data(driver))
        rows.extend(collect_grad_data(driver))
        rows.extend(collect_institute_data(driver))
        rows.extend(collect_research_data(driver))

    finally:
        driver.quit()

    print("[전체] 수집 행 수:", len(rows))
    save_excels(rows)


if __name__ == "__main__":
    main()