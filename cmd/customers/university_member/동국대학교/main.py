# -*- coding: utf-8 -*-
"""
동국대학교 이메일 수집 최종 단일 실행 파일

실행 방법:
    python 동국대_이메일_수집_최종_단일.py

실행 결과:
    1) 동국대_전체_원본_YYYYMMDDHHMMSS.xlsx
       - 수집된 전체 데이터 그대로 저장
       - 이메일 빈값 유지
       - 이메일 형식이 아닌 값 유지
       - 이메일 중복 유지

    2) 동국대_전체_정리본_YYYYMMDDHHMMSS.xlsx
       - 이메일 빈값 제거
       - 이메일 형식이 아닌 값 제거
       - 이메일 소문자 기준 중복 제거
       - 같은 이메일은 가장 위에 나온 1개만 유지

중요:
    - 이 파일 하나만 실행하면 됩니다.
    - 다른 py 파일을 읽거나 실행하지 않습니다.
    - 중간 엑셀 3개를 만들지 않습니다.
    - 최종 엑셀 2개만 생성합니다.

필요 패키지:
    pip install selenium pandas beautifulsoup4 openpyxl requests
"""

import re
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment


EMAIL_RE_FINAL = r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}"

FINAL_COLUMNS = [
    "구분",
    "대분류",
    "중분류",
    "소분류",
    "직위",
    "업무",
    "이름",
    "이메일",
    "URL",
    "학과URL",
    "교수URL"
]


def collect_college_rows():
    # -*- coding: utf-8 -*-

    import re
    import time
    from urllib.parse import urljoin

    import pandas as pd
    from bs4 import BeautifulSoup

    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options


    START_URL = "https://www.dongguk.edu/page/853#none"
    OUT_FILE = "동국대_교수목록.xlsx"

    EMAIL_RE = r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}"
    NOTICE_MAX_PAGE = 20

    PROF_TEXTS = [
        "교수진",
        "현직교수",
        "현직 교수",
        "교수소개",
        "교수 소개",
        "교수님 소개",
        "전임교원"
    ]

    # === 신규 === 공지/게시판 메뉴명 확장
    NOTICE_TEXTS = [
        "학과사무실공지",
        "학사공지",
        "학부공지",
        "학사정보",
        "학과소식",
        "알림판",
        "공지사항",
        "게시판"
    ]

    NAME_ALIASES = {
        "불교학과": ["불교학과", "불교학부"],
        "영어영문학부": ["영어영문학부", "영어통번역학전공", "영어문학전공"]
    }

    # === 신규 === 학과별 강제 URL 예외
    SPECIAL_DEPT_URLS = {
        "의료인공지능공학과": "https://medai.dongguk.edu/main",
        "산업시스템공학과": "https://ise.dongguk.edu/main",
        "사회복지상담학과": "https://dswc.dongguk.edu/main",
        "글로벌무역학과": "https://gt.dongguk.edu"
    }

    SPECIAL_PROFESSOR_URLS = {
        "광고홍보학과": "http://dguadpr.kr/sh_page/page26.php",
        "의료인공지능공학과": "https://medai.dongguk.edu/professor/list?professor_type=PROF_006 | https://medai.dongguk.edu/professor/list?professor_type=PROF_015 | https://medai.dongguk.edu/professor/list?professor_type=PROF_032",
        "산업시스템공학과": "https://ise.dongguk.edu/professor/list?professor_haggwa_type=PROFH_039&professor_type=PROF_006 | https://ise.dongguk.edu/professor/list?professor_haggwa_type=PROFH_039&professor_type=PROF_010",
        "영화영상학과": "https://movie.dongguk.edu/movie23_1_1_3",
        "영화영상제작학과": "https://movie.dongguk.edu/movie23_1_1_3",
        "영상영화학과": "https://movie.dongguk.edu/movie23_1_1_3",
        "사회복지상담학과": "https://dswc.dongguk.edu/professor/list?professor_haggwa_type=PROFH_059&professor_type=PROF_007 | https://dswc.dongguk.edu/professor/list?professor_haggwa_type=PROFH_059&professor_type=PROF_034",
        "글로벌무역학과": "https://gt.dongguk.edu/ibuilder.do?menu_idx=49"
    }

    SPECIAL_NOTICE_URLS = {
        "광고홍보학과": "http://dguadpr.kr/bbs/board.php?bo_table=table31",
        "글로벌무역학과": "https://gt.dongguk.edu/bbs/data/list.do?menu_idx=58"
    }

    # === 신규 === 메인 목록에서 누락될 때 강제로 한 번 더 처리할 학과
    EXTRA_DEPTS = [
        {
            "대분류": "미래융합대학",
            "중분류": "글로벌무역학과",
            "URL": "https://gt.dongguk.edu"
        }
    ]


    def clean_text(text):
        return " ".join(str(text).split()).replace("홈페이지 바로가기", "").strip()


    def norm_text(text):
        return clean_text(text).replace(" ", "")


    def clean_name(text):
        text = clean_text(text)
        text = text.replace("교수님", "")
        text = text.replace("강의초빙교수", "")
        text = text.replace("초빙교수", "")
        text = text.replace("명예교수", "")
        text = text.replace("석좌교수", "")
        text = text.replace("겸임교수", "")
        text = text.replace("교수", "")
        return clean_text(text)


    def clean_prof_name_text(text):
        text = clean_name(text)
        text = text.replace(" ", "")
        return text


    def get_special_value(mapping, dept_name):
        dept_key = norm_text(dept_name)

        for key in mapping:
            if norm_text(key) == dept_key:
                return mapping[key]

        for key in mapping:
            key_text = norm_text(key)
            if key_text and (key_text in dept_key or dept_key in key_text):
                return mapping[key]

        return ""


    def get_dept_url(dept_name, dept_url):
        special_url = get_special_value(SPECIAL_DEPT_URLS, dept_name)

        if special_url:
            return special_url

        return urljoin(START_URL, dept_url)


    # === 신규 === 광고홍보학과 교수진 소개 전용 파서
    # 구조: #page26 .table_box 안에 p.name, p.mail 이 있음
    # 예: <p class="name">조 형 오</p>, <p class="mail">02-2260-3808<br>hocho@dongguk.edu</p>
    def parse_adpr_professors(soup):
        rows = []

        for box in soup.select("#page26 .table_box"):
            name = ""
            email = ""

            name_tag = box.select_one("p.name")
            if name_tag:
                name = clean_prof_name_text(name_tag.get_text(" ", strip=True))

            mail_tag = box.select_one("p.mail")
            if mail_tag:
                emails = find_dongguk_emails(str(mail_tag) + " " + mail_tag.get_text(" ", strip=True))
                if emails:
                    email = emails[0]

            if not email:
                emails = find_dongguk_emails(str(box) + " " + box.get_text(" ", strip=True))
                if emails:
                    email = emails[0]

            if name or email:
                rows.append({
                    "이름": name,
                    "이메일": email
                })

        return rows


    # === 신규 === 영화/영상 학과 교수 페이지 전용 파서
    # 구조: .tbl_basic9 테이블 안에 strong 이름, E-mail 행이 있음
    def parse_movie_professors(soup):
        rows = []

        for table in soup.select("table.tbl_basic9"):
            name = ""
            email = ""

            name_tag = table.select_one("strong")
            if name_tag:
                name = clean_prof_name_text(name_tag.get_text(" ", strip=True))

            emails = find_emails(str(table) + " " + table.get_text(" ", strip=True))
            if emails:
                email = emails[0]

            if name or email:
                rows.append({
                    "이름": name,
                    "이메일": email
                })

        return rows


    def find_emails(text):
        emails = []

        for email in re.findall(EMAIL_RE, str(text)):
            email = email.strip().strip(".,;:)'\"")
            emails.append(email)

        return list(dict.fromkeys(emails))


    def find_dongguk_emails(text):
        result = []

        for email in find_emails(text):
            if "@dongguk" in email.lower():
                result.append(email)

        return list(dict.fromkeys(result))


    def make_driver():
        options = Options()

        # 필요하면 주석 해제
        options.add_argument("--headless=new")

        options.add_argument("--window-size=1920,1080")
        options.add_argument("--ignore-certificate-errors")
        options.add_argument("--allow-running-insecure-content")
        options.add_argument("--blink-settings=imagesEnabled=false")
        options.set_capability("acceptInsecureCerts", True)

        driver = webdriver.Chrome(options=options)
        driver.set_page_load_timeout(20)
        return driver


    def get_soup(driver, url):
        driver.get(url)
        time.sleep(1)
        return BeautifulSoup(driver.page_source, "html.parser")


    def is_prof_menu(text):
        text = norm_text(text)

        for word in PROF_TEXTS:
            if norm_text(word) in text:
                return True

        return False


    def is_notice_menu(text):
        text = norm_text(text)

        for word in NOTICE_TEXTS:
            if norm_text(word) in text:
                return True

        return False


    def get_notice_text_score(text):
        text = norm_text(text)
        score = 0

        for index, word in enumerate(NOTICE_TEXTS):
            if norm_text(word) in text:
                score = max(score, 100 - index)

        return score


    def is_board_href(href):
        href = str(href).lower()
        checks = [
            "/article/",
            "/bbs/",
            "bo_table=",
            "board",
            "notice",
            "list.do",
            "list"
        ]

        for word in checks:
            if word in href:
                return True

        return False


    def is_same_dept(link_text, dept_name):
        link_text = norm_text(link_text)
        dept_name = norm_text(dept_name)

        if link_text == dept_name:
            return True

        names = NAME_ALIASES.get(dept_name, [])
        for name in names:
            if norm_text(name) == link_text:
                return True

        return False


    def find_professor_url(driver, page_url, dept_name):
        # === 신규 === 학과별 교수 URL 예외를 먼저 사용
        special_url = get_special_value(SPECIAL_PROFESSOR_URLS, dept_name)
        if special_url:
            return special_url

        if not page_url:
            return ""

        try:
            soup = get_soup(driver, page_url)
        except Exception:
            print("[교수URL 실패]", page_url)
            return ""

        matched_urls = []
        candidate_urls = []

        for a in soup.select("a"):
            text = clean_text(a.get_text(" ", strip=True))
            href = a.get("href", "").strip()

            if not is_prof_menu(text):
                continue

            if href and href != "#":
                return urljoin(page_url, href)

            parent = a.find_parent("li")
            if not parent:
                parent = a.find_parent()

            for sub_a in parent.select("a"):
                sub_text = clean_text(sub_a.get_text(" ", strip=True))
                sub_href = sub_a.get("href", "").strip()

                if not sub_href or sub_href == "#":
                    continue

                full_url = urljoin(page_url, sub_href)
                candidate_urls.append(full_url)

                if is_same_dept(sub_text, dept_name):
                    matched_urls.append(full_url)

        if matched_urls:
            return " | ".join(dict.fromkeys(matched_urls))

        if candidate_urls:
            return " | ".join(dict.fromkeys(candidate_urls))

        return ""


    def add_notice_candidate(candidates, page_url, text, href):
        if not href or href == "#" or href == "#none":
            return

        if not is_notice_menu(text):
            return

        full_url = urljoin(page_url, href.strip())
        score = get_notice_text_score(text)

        if is_board_href(full_url):
            score += 1000

        if "게시판" in norm_text(text):
            score += 50

        candidates.append({
            "score": score,
            "url": full_url,
            "text": clean_text(text)
        })


    def find_notice_url(driver, page_url, dept_name):
        # === 신규 === 학과별 공지사항 URL 예외를 먼저 사용
        special_url = get_special_value(SPECIAL_NOTICE_URLS, dept_name)
        if special_url:
            return special_url

        if not page_url:
            return ""

        try:
            soup = get_soup(driver, page_url)
        except Exception:
            print("[공지URL 실패]", page_url)
            return ""

        candidates = []

        for a in soup.select("a"):
            text = clean_text(a.get_text(" ", strip=True))
            href = a.get("href", "").strip()

            add_notice_candidate(candidates, page_url, text, href)

            if href and href not in ["#", "#none"]:
                continue

            if not is_notice_menu(text):
                continue

            parent = a.find_parent("li")
            if not parent:
                parent = a.find_parent()

            for sub_a in parent.select("a"):
                sub_text = clean_text(sub_a.get_text(" ", strip=True))
                sub_href = sub_a.get("href", "").strip()
                add_notice_candidate(candidates, page_url, sub_text, sub_href)

        candidates = sorted(candidates, key=lambda x: x.get("score", 0), reverse=True)

        if candidates:
            print("[공지URL]", dept_name, candidates[0].get("text", ""), candidates[0].get("url", ""))
            return candidates[0].get("url", "")

        return ""


    def build_notice_page_url(notice_url, page_index):
        # === 신규 === 그누보드 계열 공지사항 예외
        if "bo_table=" in notice_url:
            if page_index == 1:
                return notice_url
            sep = "&" if "?" in notice_url else "?"
            return notice_url + sep + "page=" + str(page_index)

        # === 신규 === iBuilder 게시판 계열 예외
        if "/bbs/data/list.do" in notice_url:
            if page_index == 1:
                return notice_url
            sep = "&" if "?" in notice_url else "?"
            return notice_url + sep + "pageIndex=" + str(page_index)

        base = notice_url.split("?")[0]
        return base + "?pageIndex=" + str(page_index) + "&searchCondition=&searchKeyword=&category_cd="


    def make_notice_detail_url(notice_url, seq, page_index):
        base = notice_url.split("?")[0]

        if "/list" in base:
            detail_base = base.replace("/list", "/detail")
        else:
            detail_base = base.rstrip("/") + "/detail"

        return detail_base.rstrip("/") + "/" + seq + "?pageIndex=" + str(page_index)


    def make_url_target(url):
        return {
            "type": "url",
            "key": url,
            "url": url
        }


    def make_fn_view_target(bm_id, bd_id, etc):
        return {
            "type": "fn_view",
            "key": "fn_view:" + bm_id + ":" + bd_id + ":" + etc,
            "bm_id": bm_id,
            "bd_id": bd_id,
            "etc": etc
        }


    def get_notice_detail_targets(soup, notice_url, page_index):
        targets = []

        for a in soup.select("table.board tbody a"):
            href = a.get("href", "").strip()
            onclick = a.get("onclick", "").strip()

            if href and href != "#none" and "/detail/" in href:
                targets.append(make_url_target(urljoin(notice_url, href)))
                continue

            match = re.search(r"goDetail\((\d+)\)", onclick)
            if match:
                targets.append(make_url_target(make_notice_detail_url(notice_url, match.group(1), page_index)))

        for a in soup.select("a"):
            href = a.get("href", "").strip()

            if not href or href == "#" or href == "#none":
                continue

            full_url = urljoin(notice_url, href)

            # === 신규 === 그누보드 계열 상세 URL 예외
            if "bo_table=" in full_url and "wr_id=" in full_url:
                targets.append(make_url_target(full_url))
                continue

            # === 신규 === iBuilder 게시판 javascript:fn_view 예외
            match = re.search(r"fn_view\(['\"]([^'\"]+)['\"]\s*,\s*['\"]([^'\"]+)['\"]\s*,\s*['\"]?([^'\"]*)['\"]?\)", href)
            if match:
                targets.append(make_fn_view_target(match.group(1), match.group(2), match.group(3)))

        result = []
        used = set()

        for target in targets:
            key = target.get("key", "")
            if key in used:
                continue

            used.add(key)
            result.append(target)

        return result


    def get_detail_soup_by_target(driver, list_url, target):
        if target.get("type") == "url":
            url = target.get("url", "")
            return get_soup(driver, url), url

        if target.get("type") == "fn_view":
            driver.get(list_url)
            time.sleep(0.5)
            driver.execute_script(
                "fn_view(arguments[0], arguments[1], arguments[2]);",
                target.get("bm_id", ""),
                target.get("bd_id", ""),
                target.get("etc", "")
            )
            time.sleep(1)
            return BeautifulSoup(driver.page_source, "html.parser"), driver.current_url

        return None, ""


    def find_notice_admin_email(driver, notice_url, professor_emails):
        if not notice_url:
            return None

        checked_keys = set()

        for page_index in range(1, NOTICE_MAX_PAGE + 1):
            list_url = build_notice_page_url(notice_url, page_index)

            try:
                soup = get_soup(driver, list_url)
            except Exception:
                print("[공지목록 실패]", list_url)
                continue

            detail_targets = get_notice_detail_targets(soup, notice_url, page_index)
            detail_targets = [target for target in detail_targets if target.get("key", "") not in checked_keys]

            if not detail_targets:
                print("[공지] 상세 URL 없음", list_url)
                break

            print("[공지] page=", page_index, "상세=", len(detail_targets))

            for target in detail_targets:
                checked_keys.add(target.get("key", ""))

                try:
                    detail_soup, detail_url = get_detail_soup_by_target(driver, list_url, target)
                except Exception:
                    print("[공지상세 실패]", target.get("key", ""))
                    continue

                if not detail_soup:
                    continue

                board_view = detail_soup.select_one(".board_view")
                if not board_view:
                    board_view = detail_soup.select_one(".board_view2")
                if not board_view:
                    board_view = detail_soup.select_one(".board_view_wrap")
                if not board_view:
                    board_view = detail_soup.select_one(".view")
                if not board_view:
                    board_view = detail_soup

                html_text = str(board_view)
                body_text = board_view.get_text(" ", strip=True)
                emails = find_dongguk_emails(html_text + " " + body_text)

                for email in emails:
                    if email.lower() not in professor_emails:
                        print("[공지메일 발견]", email, detail_url)
                        return {
                            "이메일": email,
                            "공지사항URL": detail_url
                        }

            time.sleep(0.2)

        return None


    def get_depth3_urls(soup, professor_url):
        urls = []

        for a in soup.select(".depth3 a"):
            href = a.get("href", "").strip()

            if not href or href == "#":
                continue

            urls.append(urljoin(professor_url, href))

        return list(dict.fromkeys(urls))


    def parse_prof_items(soup):
        rows = []

        for item in soup.select(".prof_item"):
            name = ""
            email = ""

            name_tag = item.select_one(".prof_info strong")
            if name_tag:
                name = clean_name(name_tag.get_text(" ", strip=True))

            mail_tag = item.select_one("li.mail .txt")
            if mail_tag:
                email = clean_text(mail_tag.get_text(" ", strip=True))

            if not email:
                emails = re.findall(EMAIL_RE, item.get_text(" ", strip=True))
                if emails:
                    email = emails[0]

            if name or email:
                rows.append({
                    "이름": name,
                    "이메일": email
                })

        return rows


    def parse_cons_teaching(soup):
        rows = []

        for item in soup.select(".cons-teaching"):
            name = ""
            email = ""

            photo_ps = item.select(".photo p")
            for p in photo_ps:
                text = clean_text(p.get_text(" ", strip=True))
                if "교수" in text:
                    name = clean_name(text)
                    break

            emails = re.findall(EMAIL_RE, item.get_text(" ", strip=True))
            if emails:
                email = clean_text(emails[0].replace("\xa0", ""))

            if name or email:
                rows.append({
                    "이름": name,
                    "이메일": email
                })

        return rows


    def guess_prof_name(text, email):
        text = clean_text(text).replace(email, " ")

        match = re.search(r"([가-힣]{2,5})\s*(명예교수|석좌교수|초빙교수|겸임교수|부교수|조교수|교수)", text)
        if match:
            return clean_name(match.group(1))

        match = re.search(r"(명예교수|석좌교수|초빙교수|겸임교수|부교수|조교수|교수)\s*([가-힣]{2,5})", text)
        if match:
            return clean_name(match.group(2))

        skip_words = [
            "광고홍보학과", "동국대학교", "교수소개", "교수진", "이메일", "메일", "연구실", "전화", "학력", "경력", "전공"
        ]

        for word in re.findall(r"[가-힣]{2,5}", text):
            if word in skip_words:
                continue
            if "학과" in word or "교수" in word or "소개" in word:
                continue
            return clean_name(word)

        return ""


    def parse_generic_professors(soup):
        rows = []
        used = set()

        for item in soup.select("tr, li, td, .prof, .professor, .member, .person, .item, .box"):
            item_text = item.get_text(" ", strip=True)
            emails = find_emails(str(item) + " " + item_text)

            for email in emails:
                key = email.lower()
                if key in used:
                    continue

                used.add(key)
                rows.append({
                    "이름": guess_prof_name(item_text, email),
                    "이메일": email
                })

        if not rows:
            body_text = soup.get_text(" ", strip=True)
            for email in find_emails(str(soup) + " " + body_text):
                key = email.lower()
                if key in used:
                    continue

                used.add(key)
                rows.append({
                    "이름": guess_prof_name(body_text, email),
                    "이메일": email
                })

        return rows


    def parse_professors(soup):
        rows = []
        result = []
        used = set()

        # === 신규 === 학과별 전용 구조를 먼저 파싱
        rows.extend(parse_adpr_professors(soup))
        rows.extend(parse_movie_professors(soup))
        rows.extend(parse_prof_items(soup))
        rows.extend(parse_cons_teaching(soup))

        # 전용/기존 구조에서 못 잡힌 경우만 보조 파서 사용
        if not rows:
            rows.extend(parse_generic_professors(soup))

        for row in rows:
            key = (row.get("이름", ""), row.get("이메일", ""))
            if key in used:
                continue

            used.add(key)
            result.append(row)

        return result


    def get_professors(driver, professor_url):
        if not professor_url:
            return []

        result = []

        for url in professor_url.split("|"):
            url = url.strip()
            if not url:
                continue

            try:
                soup = get_soup(driver, url)
            except Exception:
                print("[교수목록 실패]", url)
                continue

            depth3_urls = get_depth3_urls(soup, url)

            if depth3_urls:
                for depth_url in depth3_urls:
                    try:
                        depth_soup = get_soup(driver, depth_url)
                        profs = parse_professors(depth_soup)

                        for prof in profs:
                            prof["교수URL"] = depth_url
                            result.append(prof)

                        print("[교수수집]", depth_url, len(profs), "명")
                    except Exception:
                        print("[교수목록 실패]", depth_url)
            else:
                profs = parse_professors(soup)

                for prof in profs:
                    prof["교수URL"] = url
                    result.append(prof)

                print("[교수수집]", url, len(profs), "명")

        return result


    def add_professor_rows(rows, driver, big_name, dept_name, dept_url):
        professor_url = find_professor_url(driver, dept_url, dept_name)
        professors = get_professors(driver, professor_url)
        professor_emails = set()

        if not professors:
            rows.append({
                "구분": "대학",
                "대분류": big_name,
                "중분류": dept_name,
                "소분류": dept_name,
                "직위": "교수",
                "업무": "교육·연구",
                "이름": "",
                "이메일": "",
                "URL": dept_url,
                "교수URL": professor_url
            })

            print("[수집]", big_name, ">", dept_name, "| 교수 0명")
            return professor_emails

        for prof in professors:
            for email in find_emails(prof.get("이메일", "")):
                professor_emails.add(email.lower())

            rows.append({
                "구분": "대학",
                "대분류": big_name,
                "중분류": dept_name,
                "소분류": dept_name,
                "직위": "교수",
                "업무": "교육·연구",
                "이름": prof.get("이름", ""),
                "이메일": prof.get("이메일", ""),
                "URL": dept_url,
                "교수URL": prof.get("교수URL", professor_url)
            })

        print("[수집]", big_name, ">", dept_name, "| 교수", len(professors), "명")
        return professor_emails


    def add_notice_admin_row(rows, driver, big_name, dept_name, dept_url, professor_emails):
        notice_url = find_notice_url(driver, dept_url, dept_name)

        if not notice_url:
            print("[공지]", big_name, ">", dept_name, "| 공지사항 URL 없음")
            return

        found = find_notice_admin_email(driver, notice_url, professor_emails)

        if not found:
            print("[공지]", big_name, ">", dept_name, "| 행정 이메일 없음")
            return

        rows.append({
            "구분": "대학",
            "대분류": big_name,
            "중분류": dept_name,
            "소분류": "행정",
            "직위": "",
            "업무": "행정",
            "이름": "",
            "이메일": found.get("이메일", ""),
            "URL": found.get("공지사항URL", notice_url),
            "교수URL": ""
        })

        print("[수집]", big_name, ">", dept_name, "| 공지 행정", found.get("이메일", ""))


    def main():
        print("[시작] 동국대 교수 목록 수집")

        driver = make_driver()
        rows = []
        processed_depts = set()

        soup = get_soup(driver, START_URL)

        for item in soup.select(".univ .item"):
            h3 = item.select_one("h3.depart_tit")
            if not h3:
                continue

            big_name = clean_text(h3.get_text(" ", strip=True))

            # 대분류 = 중분류 = 소분류 row는 pass
            # 그래서 여기서는 대분류 자체 row를 만들지 않음

            for a in item.select("ul.univ_list a.link"):
                span = a.select_one("span")
                if not span:
                    continue

                dept_name = clean_text(span.get_text(" ", strip=True))
                dept_url = a.get("href", "").strip()
                dept_url = get_dept_url(dept_name, dept_url)

                processed_depts.add(norm_text(dept_name))

                professor_emails = add_professor_rows(rows, driver, big_name, dept_name, dept_url)
                add_notice_admin_row(rows, driver, big_name, dept_name, dept_url, professor_emails)

            for div in item.select("ul.univ_list div.link"):
                span = div.select_one("span")
                if not span:
                    continue

                dept_name = clean_text(span.get_text(" ", strip=True))

                rows.append({
                    "구분": "대학",
                    "대분류": big_name,
                    "중분류": dept_name,
                    "소분류": dept_name,
                    "직위": "교수",
                    "업무": "교육·연구",
                    "이름": "",
                    "이메일": "",
                    "URL": "",
                    "교수URL": ""
                })

                processed_depts.add(norm_text(dept_name))

                print("[수집]", big_name, ">", dept_name, "| URL 없음")

        # === 신규 === 메인 목록에서 빠진 예외 학과 보정 처리
        for extra in EXTRA_DEPTS:
            big_name = extra.get("대분류", "")
            dept_name = extra.get("중분류", "")
            dept_url = extra.get("URL", "")

            if norm_text(dept_name) in processed_depts:
                continue

            print("[예외학과]", big_name, ">", dept_name)
            professor_emails = add_professor_rows(rows, driver, big_name, dept_name, dept_url)
            add_notice_admin_row(rows, driver, big_name, dept_name, dept_url, professor_emails)

        driver.quit()

        print("[완료] 대학", len(rows), "건 수집")
        return rows



    return main()


def collect_graduate_rows():
    # -*- coding: utf-8 -*-

    import re
    import time
    from urllib.parse import urljoin

    import pandas as pd
    from bs4 import BeautifulSoup

    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.common.exceptions import UnexpectedAlertPresentException, TimeoutException, WebDriverException

    START_URL = "https://www.dongguk.edu/page/853#none"
    OUT_FILE = "동국대_대학원_교수목록.xlsx"

    EMAIL_RE = r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}"

    DEPT_MENU_NAMES = [
        "학과안내",
        "학과(전공)",
        "전공안내"
    ]

    PROF_TOP_MENU_NAMES = [
        "교수소개",
        "교수 소개",
        "교수진 소개",
        "교수진"
    ]

    PROF_LINK_TEXTS = [
        "교수진",
        "교수소개",
        "교수 소개",
        "교수진 소개"
    ]

    SKIP_TEXTS = [
        "교과과정 안내",
        "교과과정안내",
        "교과과정",
        "교과이수기준",
        "전공별 세부교육과정"
    ]

    def close_alert(driver):
        try:
            alert = driver.switch_to.alert
            text = alert.text
            alert.accept()
            print("[알림닫기]", text)
        except Exception:
            pass

    def clean_text(text):
        return " ".join(str(text).split()).strip()


    def norm_text(text):
        return clean_text(text).replace(" ", "")


    def is_menu_text(text, names):
        text = norm_text(text)

        for name in names:
            if text == norm_text(name):
                return True

        return False


    def is_skip_text(text):
        text = norm_text(text)

        for word in SKIP_TEXTS:
            if norm_text(word) in text:
                return True

        return False


    def add_dept_admin_email_rows(result_rows, base_row, soup):
        emails = find_emails(soup.get_text(" ", strip=True))
        emails = list(dict.fromkeys(emails))

        for email in emails:
            result_rows.append({
                "구분": "대학원",
                "대분류": base_row["대분류"],
                "중분류": base_row["중분류"],
                "소분류": "행정",
                "직위": "",
                "업무": "행정",
                "이름": "",
                "이메일": email,
                "URL": base_row["URL"],
                "학과URL": base_row.get("학과URL", ""),
                "교수URL": ""
            })

            print("[행정]", base_row["대분류"], ">", base_row["중분류"], email)




    def find_emails(text):
        return list(dict.fromkeys(re.findall(EMAIL_RE, text)))


    def make_driver():
        options = Options()
        options.add_argument("--headless=new")
        options.add_argument("--window-size=1920,1080")
        options.add_argument("--blink-settings=imagesEnabled=false")

        driver = webdriver.Chrome(options=options)
        driver.set_page_load_timeout(20)
        return driver


    def get_soup(driver, url):
        try:
            driver.get(url)
            time.sleep(1)
            return BeautifulSoup(driver.page_source, "html.parser")

        except UnexpectedAlertPresentException:
            close_alert(driver)
            print("[페이지 스킵 - 알림]", url)
            return BeautifulSoup("", "html.parser")

        except TimeoutException:
            print("[페이지 스킵 - 타임아웃]", url)
            return BeautifulSoup("", "html.parser")

        except WebDriverException as e:
            close_alert(driver)
            print("[페이지 스킵 - 실패]", url, str(e).splitlines()[0])
            return BeautifulSoup("", "html.parser")


    def get_graduate_list(driver):
        soup = get_soup(driver, START_URL)

        rows = []

        for li in soup.select(".ft-major-list li"):
            title_a = li.select_one("p.ft-major-tit a")
            if not title_a:
                continue

            title = clean_text(title_a.get_text(" ", strip=True))

            if title != "대학원":
                continue

            for a in li.select(".col a"):
                name = clean_text(a.get_text(" ", strip=True))
                url = a.get("href", "").strip()

                rows.append({
                    "구분": "대학원",
                    "대분류": name,
                    "중분류": name,
                    "소분류": name,
                    "URL": url,
                    "학과URL": ""
                })

                print("[대학원]", name, url)

        return rows


    def get_top_menu_links(soup, base_url, menu_names):
        links = []

        for li in soup.select("#gnb ul.depth01 > li"):
            menu_a = li.find("a", recursive=False)
            if not menu_a:
                continue

            menu_text = clean_text(menu_a.get_text(" ", strip=True))

            if not is_menu_text(menu_text, menu_names):
                continue

            for a in li.select(".depth02 a"):
                text = clean_text(a.get_text(" ", strip=True))
                href = a.get("href", "").strip()

                if not text or not href or href == "#":
                    continue

                links.append({
                    "text": text,
                    "url": urljoin(base_url, href)
                })

        for dl in soup.select(".gnbs .sitemap_nav3 dl"):
            dt = dl.select_one("dt")
            if not dt:
                continue

            menu_text = clean_text(dt.get_text(" ", strip=True))

            if not is_menu_text(menu_text, menu_names):
                continue

            for a in dl.select("dd a"):
                text = clean_text(a.get_text(" ", strip=True))
                href = a.get("href", "").strip()

                if not text or not href or href == "#":
                    continue

                links.append({
                    "text": text,
                    "url": urljoin(base_url, href)
                })

        return unique_links(links)


    def unique_links(links):
        result = []
        seen = set()

        for item in links:
            key = item["text"] + "|" + item["url"]
            if key in seen:
                continue

            seen.add(key)
            result.append(item)

        return result


    def get_dept_list(soup, grad_row):
        grad_name = grad_row["대분류"]
        grad_url = grad_row["URL"]

        links = get_top_menu_links(soup, grad_url, DEPT_MENU_NAMES)

        rows = []

        for item in links:
            dept_name = item["text"]
            dept_url = item["url"]

            if is_skip_text(dept_name):
                continue

            if is_menu_text(dept_name, PROF_LINK_TEXTS):
                continue

            rows.append({
                "구분": "대학원",
                "대분류": grad_name,
                "중분류": dept_name,
                "소분류": dept_name,
                "URL": grad_url,
                "학과URL": dept_url
            })

            print("[학과]", grad_name, ">", dept_name, dept_url)

        return rows


    def get_professor_links(soup, base_url):
        links = []

        # 교수소개 / 교수진 소개 메뉴 안의 링크
        links.extend(get_top_menu_links(soup, base_url, PROF_TOP_MENU_NAMES))

        # 영상대학원 case:
        # 학과 페이지 안의 .menu_tabs .depth3 에 교수진 소개 링크가 있음
        for a in soup.select(".menu_tabs .depth3 a, .menu_tabs .depth4 a"):
            text = clean_text(a.get_text(" ", strip=True))
            href = a.get("href", "").strip()

            if not text or not href or href == "#":
                continue

            if "/professor/list" in href or is_menu_text(text, PROF_LINK_TEXTS):
                links.append({
                    "text": text,
                    "url": urljoin(base_url, href)
                })

        # 전체 a 중에서 text가 교수진 / 교수소개 / 교수진 소개 인 것만 추가
        for a in soup.select("a"):
            text = clean_text(a.get_text(" ", strip=True))
            href = a.get("href", "").strip()

            if not text or not href or href == "#":
                continue

            if is_menu_text(text, PROF_LINK_TEXTS):
                links.append({
                    "text": text,
                    "url": urljoin(base_url, href)
                })

        # 왼쪽 메뉴 case
        for a in soup.select(".s_left_area_menu1 a"):
            text = clean_text(a.get_text(" ", strip=True))
            href = a.get("href", "").strip()

            if not text or not href or href == "#":
                continue

            if is_menu_text(text, PROF_LINK_TEXTS):
                links.append({
                    "text": text,
                    "url": urljoin(base_url, href)
                })

        return unique_links(links)


    def get_bo_cate_urls(soup, base_url):
        urls = []

        for a in soup.select("#bo_cate a"):
            text = clean_text(a.get_text(" ", strip=True))
            text = text.replace("열린 분류", "").strip()

            href = a.get("href", "").strip()

            if not text or not href or href == "#":
                continue

            urls.append({
                "text": text,
                "url": urljoin(base_url, href)
            })

        return unique_links(urls)


    def get_depth_professor_urls(soup, professor_url):
        urls = []

        # 교수진 카테고리가 있으면 그 링크들을 우선 사용
        bo_cate_urls = get_bo_cate_urls(soup, professor_url)

        if bo_cate_urls:
            return bo_cate_urls

        # 교수소개 내부 professor/list 카테고리
        for a in soup.select(".menu_tabs .depth4 a"):
            text = clean_text(a.get_text(" ", strip=True))
            href = a.get("href", "").strip()

            if not href or href == "#":
                continue

            if "/professor/list" in href:
                urls.append({
                    "text": text,
                    "url": urljoin(professor_url, href)
                })

        for a in soup.select(".menu_tabs .depth3 a"):
            text = clean_text(a.get_text(" ", strip=True))
            href = a.get("href", "").strip()

            if not href or href == "#":
                continue

            if "/professor/list" in href:
                urls.append({
                    "text": text,
                    "url": urljoin(professor_url, href)
                })

            elif is_menu_text(text, PROF_LINK_TEXTS):
                urls.append({
                    "text": text,
                    "url": urljoin(professor_url, href)
                })

        return unique_links(urls)


    def parse_prof_items(soup):
        rows = []

        for item in soup.select(".prof_item"):
            name = ""
            email = ""

            name_tag = item.select_one(".prof_info strong")
            if name_tag:
                name = clean_text(name_tag.get_text(" ", strip=True))

            mail_tag = item.select_one("li.mail .txt")
            if mail_tag:
                email = clean_text(mail_tag.get_text(" ", strip=True))

            if email == "-":
                email = ""

            if not email:
                emails = find_emails(item.get_text(" ", strip=True))
                if emails:
                    email = emails[0]

            if name or email:
                rows.append({
                    "이름": name,
                    "이메일": email
                })

        return rows


    def parse_tbl_basic2(soup):
        rows = []

        for table in soup.select("table.tbl_basic2"):
            name = ""
            email = ""

            name_tag = table.select_one("strong")
            if name_tag:
                name = clean_text(name_tag.get_text(" ", strip=True))

            emails = find_emails(table.get_text(" ", strip=True))
            if emails:
                email = emails[0]

            if name or email:
                rows.append({
                    "이름": name,
                    "이메일": email
                })

        return rows


    def parse_professors(soup):
        rows = []

        rows.extend(parse_prof_items(soup))
        rows.extend(parse_tbl_basic2(soup))

        return rows


    def get_professors(driver, professor_url):
        result = []

        soup = get_soup(driver, professor_url)

        if not soup.get_text(strip=True):
            print("[교수URL 스킵]", professor_url)
            return result

        depth_urls = get_depth_professor_urls(soup, professor_url)

        if not depth_urls:
            depth_urls = [{
                "text": "",
                "url": professor_url
            }]

        for item in depth_urls:
            url = item["url"]
            cate_name = item["text"]

            prof_soup = get_soup(driver, url)

            if not prof_soup.get_text(strip=True):
                print("[교수상세 스킵]", url)
                continue

            profs = parse_professors(prof_soup)

            for prof in profs:
                prof["교수URL"] = url
                prof["분류명"] = cate_name
                result.append(prof)

            print("[교수수집]", url, len(profs), "명")

        return result


    def get_professor_category_names(base_row, cate_name=""):
        # 학과가 있는 경우: 교수 탭명은 쓰지 않고 학과명을 중분류/소분류에 그대로 사용
        if base_row.get("학과URL"):
            return base_row["중분류"], base_row["소분류"]

        # 학과가 없는 경우: 중분류/소분류 모두 대분류와 동일하게 처리
        return base_row["대분류"], base_row["대분류"]


    def make_empty_professor_row(base_row, professor_url):
        mid_name, sub_name = get_professor_category_names(base_row)

        return {
            "구분": "대학원",
            "대분류": base_row["대분류"],
            "중분류": mid_name,
            "소분류": sub_name,
            "직위": "교수",
            "업무": "교육·연구",
            "이름": "",
            "이메일": "",
            "URL": base_row["URL"],
            "학과URL": base_row.get("학과URL", ""),
            "교수URL": professor_url
        }


    def collect_professor_rows(driver, result_rows, base_row, professor_links):
        if not professor_links:
            result_rows.append(make_empty_professor_row(base_row, ""))
            return

        total = 0

        for item in professor_links:
            professor_url = item["url"]
            professors = get_professors(driver, professor_url)
            total += len(professors)

            if not professors:
                result_rows.append(make_empty_professor_row(base_row, professor_url))

            for prof in professors:
                mid_name, sub_name = get_professor_category_names(base_row, prof.get("분류명", ""))

                result_rows.append({
                    "구분": "대학원",
                    "대분류": base_row["대분류"],
                    "중분류": mid_name,
                    "소분류": sub_name,
                    "직위": "교수",
                    "업무": "교육·연구",
                    "이름": prof.get("이름", ""),
                    "이메일": prof.get("이메일", ""),
                    "URL": base_row["URL"],
                    "학과URL": base_row.get("학과URL", ""),
                    "교수URL": prof.get("교수URL", professor_url)
                })

        print("[교수완료]", base_row["대분류"], ">", base_row["중분류"], total, "명")


    def add_education_admin_rows(result_rows, grad_row, soup):
        emails = find_emails(soup.get_text(" ", strip=True))[:2]

        for email in emails:
            result_rows.append({
                "구분": "대학원",
                "대분류": grad_row["대분류"],
                "중분류": grad_row["대분류"],
                "소분류": "행정",
                "직위": "",
                "업무": "행정",
                "이름": "",
                "이메일": email,
                "URL": grad_row["URL"],
                "학과URL": "",
                "교수URL": ""
            })

            print("[교육대학원 행정]", email)


    def is_direct_professor_grad(grad_name):
        names = [
            "미래융합대학원",
            "교육서비스과학대학원"
        ]

        return grad_name in names


    def main():
        print("[시작] 동국대 대학원 교수 목록 수집")

        driver = make_driver()

        grad_rows = get_graduate_list(driver)
        result_rows = []

        for i, grad_row in enumerate(grad_rows, start=1):
            grad_name = grad_row["대분류"]
            grad_url = grad_row["URL"]

            print("[진행]", i, "/", len(grad_rows), grad_name)

            soup = get_soup(driver, grad_url)

            if grad_name == "교육대학원":
                add_education_admin_rows(result_rows, grad_row, soup)
                continue

            # 미래융합대학원 / 교육서비스과학대학원은 학과URL보다 교수진 메뉴로 바로 진입
            if is_direct_professor_grad(grad_name):
                # 학과가 없는 대학원은 대학원 메인 페이지에서 행정 이메일 수집
                add_dept_admin_email_rows(result_rows, grad_row, soup)

                professor_links = get_professor_links(soup, grad_url)
                collect_professor_rows(driver, result_rows, grad_row, professor_links)
                continue

            dept_rows = get_dept_list(soup, grad_row)

            if dept_rows:

                for dept_row in dept_rows:
                    dept_soup = get_soup(driver, dept_row["학과URL"])

                    # 학과URL 메인 페이지 이메일 먼저 수집
                    add_dept_admin_email_rows(result_rows, dept_row, dept_soup)

                    # 그 다음 교수진 링크 수집
                    professor_links = get_professor_links(dept_soup, dept_row["학과URL"])

                    if not professor_links:
                        professor_links = get_professor_links(soup, grad_url)

                    collect_professor_rows(driver, result_rows, dept_row, professor_links)

            else:
                # 학과 목록이 없는 대학원은 대학원 메인 페이지에서 행정 이메일 수집
                add_dept_admin_email_rows(result_rows, grad_row, soup)

                professor_links = get_professor_links(soup, grad_url)
                collect_professor_rows(driver, result_rows, grad_row, professor_links)

        driver.quit()

        print("[완료] 대학원", len(result_rows), "건 수집")
        return result_rows



    return main()


def collect_sanhak_rows():
    import re
    import requests
    import pandas as pd
    from bs4 import BeautifulSoup
    from urllib.parse import urljoin

    BASE_URL = "https://rnd.dongguk.edu"
    START_URL = "https://rnd.dongguk.edu/ko/page/sub/sub_0105_00.do;jsessionid=D92158A27813FF90B739371244D6603F"

    EMAIL_RE = r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}"

    rows = []

    html = requests.get(START_URL).text
    soup = BeautifulSoup(html, "html.parser")

    last_nav = soup.select(".nav-02")[-1]

    menus = []

    for a in last_nav.select("li a[href]"):
        text = a.get_text(strip=True)

        if text == "조직도":
            continue

        menus.append({
            "text": text,
            "url": urljoin(BASE_URL, a.get("href"))
        })

    for idx, menu in enumerate(menus, start=1):
        print("[진행]", idx, "/", len(menus), menu["text"])

        html = requests.get(menu["url"]).text
        soup = BeautifulSoup(html, "html.parser")

        for tr in soup.select(".table-wrap table tbody tr"):
            tr_text = tr.get_text(" ", strip=True)
            emails = re.findall(EMAIL_RE, tr_text)

            if not emails:
                continue

            works = []

            for li in tr.select("td.left li"):
                works.append(li.get_text(" ", strip=True))

            work_text = "\n".join(works)

            for email in emails:
                rows.append({
                    "대분류": "산학협력단",
                    "중분류": menu["text"],
                    "소분류": "행정",
                    "직위": "",
                    "업무": work_text,
                    "이메일": email,
                    "URL": menu["url"]
                })

    print("[완료] 산학협력단", len(rows), "건 수집")
    return rows




def now_text():
    return datetime.now().strftime("%Y%m%d%H%M%S")


def clean_email_for_final(value):
    value = str(value).strip()

    if not value or value.lower() == "nan":
        return ""

    match = re.search(EMAIL_RE_FINAL, value)
    if not match:
        return ""

    return match.group(0).strip().strip(".,;:)'\"")


def normalize_rows(rows, default_gubun):
    fixed_rows = []

    for row in rows:
        new_row = {}

        for col in FINAL_COLUMNS:
            new_row[col] = row.get(col, "")

        if not new_row["구분"]:
            new_row["구분"] = default_gubun

        if new_row["구분"] == "산학협력단":
            if not new_row["대분류"]:
                new_row["대분류"] = "산학협력단"
            if not new_row["소분류"]:
                new_row["소분류"] = "행정"
            if not new_row["업무"]:
                new_row["업무"] = "행정"

        fixed_rows.append(new_row)

    return fixed_rows


def make_original_df(rows):
    return pd.DataFrame(rows, columns=FINAL_COLUMNS)


def make_clean_df(df):
    clean_df = df.copy()
    clean_df["이메일"] = clean_df["이메일"].apply(clean_email_for_final)
    clean_df = clean_df[clean_df["이메일"] != ""].copy()
    clean_df["_email_key"] = clean_df["이메일"].str.lower()
    clean_df = clean_df.drop_duplicates(subset=["_email_key"], keep="first")
    clean_df = clean_df.drop(columns=["_email_key"])
    return clean_df


def format_excel(filename):
    wb = load_workbook(filename)
    ws = wb.active
    ws.freeze_panes = "A2"

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter

        for cell in col:
            text = str(cell.value) if cell.value is not None else ""
            if len(text) > max_len:
                max_len = len(text)

        width = max_len + 2
        if width < 12:
            width = 12
        if width > 50:
            width = 50

        ws.column_dimensions[col_letter].width = width

    wb.save(filename)


def save_final_excels(df):
    ts = now_text()
    original_file = "동국대_전체_원본_" + ts + ".xlsx"
    clean_file = "동국대_전체_정리본_" + ts + ".xlsx"

    clean_df = make_clean_df(df)

    df.to_excel(original_file, index=False)
    clean_df.to_excel(clean_file, index=False)

    format_excel(original_file)
    format_excel(clean_file)

    print("\n[엑셀] 원본 저장 완료:", original_file, len(df), "건")
    print("[엑셀] 정리본 저장 완료:", clean_file, len(clean_df), "건")


def main():
    print("[시작] 동국대학교 이메일 수집 최종 단일 실행")
    print("[안내] 다른 py 파일을 읽지 않고, 최종 엑셀 2개만 생성합니다.")

    all_rows = []

    college_rows = collect_college_rows()
    all_rows.extend(normalize_rows(college_rows, "대학"))

    graduate_rows = collect_graduate_rows()
    all_rows.extend(normalize_rows(graduate_rows, "대학원"))

    sanhak_rows = collect_sanhak_rows()
    all_rows.extend(normalize_rows(sanhak_rows, "산학협력단"))

    df = make_original_df(all_rows)
    save_final_excels(df)

    print("\n[완료] 전체 수집 종료")


if __name__ == "__main__":
    main()
