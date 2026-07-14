import csv
import logging
import math
import os
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import urlparse, parse_qsl

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


# =============================================================================
# 설정
# =============================================================================

REQUEST_URL = "https://review6.cre.ma/api/blancnature.co.kr/reviews?secure_device_token=V293fddf293c56b5e7c028ec9237f6b4047339ad357af961adcd0032af21cbf0175938c4f3a8f7746dd66a17c8e9c806d4&fields=has_media,total_product_media_reviews_count,reviews.evaluation_properties,reviews.ai_summary,reviews.with_parent_reviews,reviews.customer_properties&product_code=13&iframe_id=crema-product-reviews-3&widget_id=2&widget_style=list_v3&locale=ko-KR&app=0&iframe=1&page=1"

REFERER = "https://review6.cre.ma/v2/blancnature.co.kr/product_reviews/list_v3?product_code=13&iframe_id=crema-product-reviews-3&widget_id=2&widget_style=list_v3&install_method=smart&locale=ko-KR&app=0&parent_url=https%3A%2F%2Fblancnature.co.kr%2Fproduct%2Fbest-%25EC%2595%2584%25ED%2581%25AC%25EB%2584%25A4-%25ED%2581%25B4%25EB%25A0%258C%25EC%25A7%2595%25ED%258F%25BC-150ml%2F13%2Fcategory%2F46%2Fdisplay%2F1%2F&secure_device_token=V293fddf293c56b5e7c028ec9237f6b4047339ad357af961adcd0032af21cbf0175938c4f3a8f7746dd66a17c8e9c806d4&iframe=1"

# 필요한 상품만 남기고, 안 쓸 상품은 주석 처리하면 됨
PRODUCTS = [
    # {"product_code": "13", "product_name": "아크네 클렌징폼", "expected_reviews": 12285},
    # {"product_code": "12", "product_name": "매직 티트리오일", "expected_reviews": 10138},
    # {"product_code": "11", "product_name": "스네일 모이스처라이징 크림", "expected_reviews": 5347},
    # {"product_code": "96", "product_name": "판테놀 시카세럼", "expected_reviews": 416},
    {"product_code": "175", "product_name": "어성초 토너", "expected_reviews": 1620},

    # 예시: 잠깐 제외하고 싶으면 이렇게 주석 처리
    # {"product_code": "13", "product_name": "아크네 클렌징폼", "expected_reviews": 12285},
]

OUTPUT_CSV = "crema_reviews.csv"
OUTPUT_XLSX = "crema_reviews.xlsx"
PAGE_STATUS_CSV = "crema_review_page_status.csv"
LOG_FILE = "crema_reviews.log"

# 요청 병렬 개수
MAX_WORKERS = 8

# CREMA 응답 샘플 기준 한 페이지 5개
PAGE_SIZE = 5

# 리뷰 수가 살짝 달라졌을 수 있으니 뒤에 여유 페이지 몇 개 더 요청
EXTRA_PAGE_COUNT = 5

REQUEST_TIMEOUT_SECONDS = 20
MAX_RETRY_COUNT = 3

CSV_HEADERS = [
    "상품코드",
    "상품명_입력",
    "페이지",
    "리뷰ID",
    "글쓴이",
    "제품명",
    "텍스트",
    "이미지",
    "영상 링크",
]

PAGE_STATUS_HEADERS = [
    "상품코드",
    "페이지",
    "상태",
    "수집건수",
    "오류",
]


# =============================================================================
# 로거
# =============================================================================

def setup_logger():
    logger = logging.getLogger("crema_review_crawler")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()

    formatter = logging.Formatter(
        fmt="%(asctime)s | %(levelname)s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S"
    )

    console_handler = logging.StreamHandler()
    console_handler.setFormatter(formatter)

    file_handler = logging.FileHandler(LOG_FILE, encoding="utf-8")
    file_handler.setFormatter(formatter)

    logger.addHandler(console_handler)
    logger.addHandler(file_handler)

    return logger


def format_elapsed(seconds):
    seconds = int(seconds)

    hour = seconds // 3600
    minute = (seconds % 3600) // 60
    second = seconds % 60

    if hour > 0:
        return f"{hour}시간 {minute}분 {second}초"

    if minute > 0:
        return f"{minute}분 {second}초"

    return f"{second}초"


# =============================================================================
# 기본 유틸
# =============================================================================

def build_base_params(request_url):
    parsed_url = urlparse(request_url)
    params = dict(parse_qsl(parsed_url.query))

    api_url = parsed_url.scheme + "://" + parsed_url.netloc + parsed_url.path

    return api_url, params


def get_headers():
    return {
        "accept": "application/json, text/plain, */*",
        "accept-language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
        "cache-control": "no-cache",
        "pragma": "no-cache",
        "referer": REFERER,
        "sec-ch-ua": '"Not;A=Brand";v="8", "Chromium";v="150", "Google Chrome";v="150"',
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": '"Windows"',
        "sec-fetch-dest": "empty",
        "sec-fetch-mode": "cors",
        "sec-fetch-site": "same-origin",
        "sec-fetch-storage-access": "none",
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/150.0.0.0 Safari/537.36",
    }


def safe_int(value, default_value=0):
    try:
        if value is None or value == "":
            return default_value

        return int(value)

    except (ValueError, TypeError):
        return default_value


def normalize_url(url):
    if not url:
        return ""

    if url.startswith("//"):
        return "https:" + url

    return url


def extract_media_urls(media_list):
    if not isinstance(media_list, list):
        return ""

    urls = []

    for media in media_list:
        if not isinstance(media, dict):
            continue

        url = (
                media.get("url")
                or media.get("video_url")
                or media.get("source_url")
                or media.get("file_url")
                or media.get("thumbnail_url")
        )

        url = normalize_url(url)

        if url:
            urls.append(url)

    return ", ".join(urls)


def extract_review_row(product, page_no, review):
    review_id = review.get("id", "")
    writer = review.get("user_display_name", "")
    api_product_name = review.get("product_name", "")
    text = review.get("filtered_message", "")

    image_urls = extract_media_urls(review.get("images", []))
    video_urls = extract_media_urls(review.get("videos", []))

    return [
        product.get("product_code", ""),
        product.get("product_name", ""),
        page_no,
        review_id,
        writer,
        api_product_name,
        text,
        image_urls,
        video_urls,
    ]


def ensure_csv_file(file_path, headers):
    if os.path.exists(file_path) and os.path.getsize(file_path) > 0:
        return

    with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers)


# =============================================================================
# 이어받기 처리
# =============================================================================

def load_saved_review_keys():
    """
    이미 저장된 리뷰 중복 방지용.
    상품코드 + 리뷰ID 기준으로 중복 판단.
    """
    ensure_csv_file(OUTPUT_CSV, CSV_HEADERS)

    saved_keys = set()
    saved_count = 0

    with open(OUTPUT_CSV, "r", newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)

        for row in reader:
            product_code = str(row.get("상품코드", "")).strip()
            review_id = str(row.get("리뷰ID", "")).strip()

            if product_code and review_id:
                saved_keys.add(f"{product_code}:{review_id}")

            saved_count += 1

    return saved_keys, saved_count


def load_completed_pages_from_status():
    """
    페이지 상태 파일 기준으로 이미 성공/빈 페이지 처리된 page는 다시 요청하지 않음.
    """
    ensure_csv_file(PAGE_STATUS_CSV, PAGE_STATUS_HEADERS)

    completed_pages = set()

    with open(PAGE_STATUS_CSV, "r", newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)

        for row in reader:
            product_code = str(row.get("상품코드", "")).strip()
            page_no = safe_int(row.get("페이지"), 0)
            status = str(row.get("상태", "")).strip()

            if not product_code or page_no <= 0:
                continue

            if status in ("SUCCESS", "EMPTY"):
                completed_pages.add(f"{product_code}:{page_no}")

    return completed_pages


def append_review_rows(rows):
    if not rows:
        return

    ensure_csv_file(OUTPUT_CSV, CSV_HEADERS)

    with open(OUTPUT_CSV, "a", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerows(rows)


def append_page_status(product_code, page_no, status, count, error_message=""):
    ensure_csv_file(PAGE_STATUS_CSV, PAGE_STATUS_HEADERS)

    with open(PAGE_STATUS_CSV, "a", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow([
            product_code,
            page_no,
            status,
            count,
            error_message,
        ])


# =============================================================================
# 요청 처리
# =============================================================================

thread_local = threading.local()


def get_thread_session():
    """
    스레드별 Session 사용.
    requests.Session은 스레드 간 공유하지 않는 게 안전함.
    """
    if not hasattr(thread_local, "session"):
        thread_local.session = requests.Session()

    return thread_local.session


def fetch_page(product, page_no, api_url, base_params, headers):
    product_code = product.get("product_code", "")

    params = dict(base_params)
    params["product_code"] = product_code
    params["page"] = str(page_no)

    session = get_thread_session()

    last_error = None

    for retry_index in range(1, MAX_RETRY_COUNT + 1):
        try:
            response = session.get(
                api_url,
                headers=headers,
                params=params,
                timeout=REQUEST_TIMEOUT_SECONDS
            )

            response.raise_for_status()

            data = response.json()
            reviews = data.get("reviews", [])

            if not isinstance(reviews, list):
                raise RuntimeError(f"reviews 타입 이상: {type(reviews)}")

            rows = []

            for review in reviews:
                rows.append(extract_review_row(product, page_no, review))

            if not reviews:
                status = "EMPTY"
            else:
                status = "SUCCESS"

            return {
                "product_code": product_code,
                "product_name": product.get("product_name", ""),
                "page_no": page_no,
                "status": status,
                "rows": rows,
                "review_count": len(reviews),
                "error": "",
            }

        except Exception as e:
            last_error = e
            time.sleep(1.5 * retry_index)

    return {
        "product_code": product_code,
        "product_name": product.get("product_name", ""),
        "page_no": page_no,
        "status": "FAILED",
        "rows": [],
        "review_count": 0,
        "error": str(last_error),
    }


def make_page_jobs(products, completed_pages):
    """
    expected_reviews 기준으로 요청할 page 목록 생성.
    한 페이지 5개 기준 + 여유 페이지 EXTRA_PAGE_COUNT.
    """
    jobs = []

    for product in products:
        product_code = product.get("product_code", "")
        expected_reviews = safe_int(product.get("expected_reviews"), 0)

        if expected_reviews <= 0:
            # 리뷰 수를 모르면 크게 잡아야 함
            # 보통은 위 PRODUCTS에 expected_reviews를 넣는 것을 권장
            max_page = 10000
        else:
            max_page = int(math.ceil(expected_reviews / float(PAGE_SIZE))) + EXTRA_PAGE_COUNT

        for page_no in range(1, max_page + 1):
            page_key = f"{product_code}:{page_no}"

            if page_key in completed_pages:
                continue

            jobs.append({
                "product": product,
                "page_no": page_no,
            })

    return jobs


# =============================================================================
# XLSX 변환
# =============================================================================

def convert_csv_to_xlsx(logger):
    if not os.path.exists(OUTPUT_CSV):
        logger.warning("CSV 파일이 없어 XLSX 변환을 건너뜁니다.")
        return

    logger.info(f"XLSX 변환 시작 | {OUTPUT_CSV} -> {OUTPUT_XLSX}")

    wb = Workbook()
    ws = wb.active
    ws.title = "reviews"

    with open(OUTPUT_CSV, "r", newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)

        for row in reader:
            ws.append(row)

    if ws.max_row <= 1:
        logger.warning("CSV에 데이터가 없어 XLSX 변환을 건너뜁니다.")
        return

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    column_widths = {
        "A": 12,    # 상품코드
        "B": 32,    # 상품명_입력
        "C": 10,    # 페이지
        "D": 14,    # 리뷰ID
        "E": 18,    # 글쓴이
        "F": 40,    # 제품명
        "G": 90,    # 텍스트
        "H": 100,   # 이미지
        "I": 100,   # 영상 링크
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"

    wb.save(OUTPUT_XLSX)

    logger.info(f"XLSX 변환 완료 | {OUTPUT_XLSX}")


# =============================================================================
# 메인
# =============================================================================

def main():
    logger = setup_logger()
    program_start_time = time.perf_counter()

    try:
        api_url, base_params = build_base_params(REQUEST_URL)
        headers = get_headers()

        saved_review_keys, existing_saved_count = load_saved_review_keys()
        completed_pages = load_completed_pages_from_status()

        jobs = make_page_jobs(PRODUCTS, completed_pages)

        total_jobs = len(jobs)
        finished_jobs = 0
        success_pages = 0
        empty_pages = 0
        failed_pages = 0
        newly_saved_count = 0
        duplicate_count = 0

        logger.info("수집 시작")
        logger.info(f"상품 수: {len(PRODUCTS)}개")
        logger.info(f"스레드 수: {MAX_WORKERS}개")
        logger.info(f"기존 저장 리뷰 수: {existing_saved_count}건")
        logger.info(f"이미 완료된 페이지 수: {len(completed_pages)}개")
        logger.info(f"이번 실행 요청 예정 페이지 수: {total_jobs}개")

        if total_jobs == 0:
            logger.info("요청할 페이지가 없습니다. XLSX 변환만 진행합니다.")
            convert_csv_to_xlsx(logger)
            return

        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            futures = []

            for job in jobs:
                future = executor.submit(
                    fetch_page,
                    job["product"],
                    job["page_no"],
                    api_url,
                    base_params,
                    headers
                )
                futures.append(future)

            for future in as_completed(futures):
                result = future.result()

                product_code = result.get("product_code", "")
                product_name = result.get("product_name", "")
                page_no = result.get("page_no", 0)
                status = result.get("status", "")
                rows = result.get("rows", [])
                error = result.get("error", "")

                filtered_rows = []

                for row in rows:
                    row_product_code = str(row[0]).strip()
                    review_id = str(row[3]).strip()
                    review_key = f"{row_product_code}:{review_id}"

                    if review_id and review_key in saved_review_keys:
                        duplicate_count += 1
                        continue

                    if review_id:
                        saved_review_keys.add(review_key)

                    filtered_rows.append(row)

                append_review_rows(filtered_rows)
                append_page_status(
                    product_code=product_code,
                    page_no=page_no,
                    status=status,
                    count=len(filtered_rows),
                    error_message=error
                )

                finished_jobs += 1

                if status == "SUCCESS":
                    success_pages += 1
                elif status == "EMPTY":
                    empty_pages += 1
                else:
                    failed_pages += 1

                newly_saved_count += len(filtered_rows)

                elapsed = time.perf_counter() - program_start_time
                progress_rate = (finished_jobs / total_jobs) * 100

                logger.info(
                    f"진행 | {finished_jobs}/{total_jobs} "
                    f"({progress_rate:.2f}%) | "
                    f"상품={product_code} {product_name} | "
                    f"page={page_no} | "
                    f"상태={status} | "
                    f"저장={len(filtered_rows)}건 | "
                    f"신규누적={newly_saved_count}건 | "
                    f"중복누적={duplicate_count}건 | "
                    f"실패페이지={failed_pages}개 | "
                    f"소요={format_elapsed(elapsed)}"
                )

                if status == "FAILED":
                    logger.warning(
                        f"페이지 실패 | 상품={product_code} | page={page_no} | 오류={error}"
                    )

        logger.info(
            f"수집 종료 | 성공페이지={success_pages} | "
            f"빈페이지={empty_pages} | 실패페이지={failed_pages} | "
            f"신규 저장={newly_saved_count}건 | 중복={duplicate_count}건"
        )

        convert_csv_to_xlsx(logger)

        total_elapsed = time.perf_counter() - program_start_time

        logger.info(
            f"프로그램 완료 | 전체 소요={format_elapsed(total_elapsed)} | "
            f"CSV={OUTPUT_CSV} | XLSX={OUTPUT_XLSX}"
        )

    except KeyboardInterrupt:
        logger.warning("사용자 중단 | 지금까지 저장된 CSV 기준으로 다음 실행 때 이어받을 수 있습니다.")

    except Exception as e:
        logger.exception(f"프로그램 오류 종료 | 다음 실행 때 이어받을 수 있습니다. 오류={e}")


if __name__ == "__main__":
    main()