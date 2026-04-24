from __future__ import annotations

import json
import os
import re
import sqlite3
import sys  # === 신규 ===
from datetime import datetime
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PySide6.QtCore import QRect, Qt, Signal
from PySide6.QtGui import QColor, QPainter
from PySide6.QtWidgets import (
    QAbstractItemView,
    QCheckBox,
    QDialog,
    QHeaderView,
    QHBoxLayout,
    QLabel,
    QMessageBox,
    QPushButton,
    QSplitter,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from src.core.global_state import GlobalState
from src.ui.style.style import create_common_button


class CheckBoxHeader(QHeaderView):
    toggled = Signal(bool)

    def __init__(self, orientation: Qt.Orientation, parent: Optional[QWidget] = None) -> None:
        super().__init__(orientation, parent)
        self.checked = False
        self.setSectionsClickable(True)

    def paintSection(self, painter: QPainter, rect: QRect, logical_index: int) -> None:
        super().paintSection(painter, rect, logical_index)

        if logical_index != 0:
            return

        size = 14
        x = rect.x() + (rect.width() - size) // 2
        y = rect.y() + (rect.height() - size) // 2

        painter.save()
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        painter.setPen(QColor("#111"))
        painter.setBrush(QColor("#111") if self.checked else QColor("white"))
        painter.drawRect(x, y, size, size)

        if self.checked:
            painter.setPen(QColor("white"))
            painter.drawLine(x + 3, y + 7, x + 6, y + 10)
            painter.drawLine(x + 6, y + 10, x + 11, y + 4)

        painter.restore()

    def mousePressEvent(self, event) -> None:
        index = self.logicalIndexAt(event.pos())
        if index == 0:
            self.checked = not self.checked
            self.toggled.emit(self.checked)
            self.viewport().update()
            return
        super().mousePressEvent(event)

    def set_checked(self, checked: bool) -> None:
        self.checked = checked
        self.viewport().update()


class DbTableWidget(QTableWidget):
    row_clicked_signal = Signal(int)

    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)

        self.row_meta: list[dict[str, Any]] = []

        header = CheckBoxHeader(Qt.Orientation.Horizontal, self)
        header.toggled.connect(self.toggle_all_checked)
        self.setHorizontalHeader(header)

        self.verticalHeader().setVisible(False)
        self.setAlternatingRowColors(True)
        self.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.setWordWrap(False)
        self.setShowGrid(True)
        self.setSortingEnabled(False)
        self.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.setStyleSheet(self._table_style())

    def _table_style(self) -> str:
        return """
        QTableWidget {
            background: white;
            color: #111;
            border: 1px solid #dcdcdc;
            border-radius: 10px;
            gridline-color: #ececec;
            selection-background-color: #eaf2ff;
            selection-color: #111;
            alternate-background-color: #fafafa;
            font-size: 12px;
        }

        QTableWidget::item:selected {
            background: #eaf2ff;
            color: #111;
        }

        QHeaderView::section {
            background: #f6f6f6;
            color: #111;
            font-weight: bold;
            border: none;
            border-right: 1px solid #ececec;
            border-bottom: 1px solid #ececec;
            padding: 8px 6px;
        }

        QScrollBar:vertical, QScrollBar:horizontal {
            background: #f3f3f3;
            border: none;
        }

        QScrollBar:vertical { width: 12px; }
        QScrollBar:horizontal { height: 12px; }

        QScrollBar::handle:vertical, QScrollBar::handle:horizontal {
            background: #c9c9c9;
            border-radius: 5px;
            min-height: 24px;
            min-width: 24px;
        }
        """

    def checkbox_style(self) -> str:
        return """
        QCheckBox::indicator {
            width: 14px;
            height: 14px;
            border: 1px solid #111;
            background: white;
        }
        QCheckBox::indicator:unchecked {
            background: white;
            border: 1px solid #111;
        }
        QCheckBox::indicator:checked {
            background: black;
            border: 1px solid #111;
            image: none;
        }
        """

    def make_checkbox_cell(self) -> QWidget:
        wrap = QWidget()
        layout = QHBoxLayout(wrap)
        layout.setContentsMargins(10, 0, 0, 0)
        layout.setSpacing(0)
        layout.setAlignment(Qt.AlignmentFlag.AlignCenter)

        cb = QCheckBox()
        cb.setCursor(Qt.CursorShape.PointingHandCursor)
        cb.setStyleSheet(self.checkbox_style())
        cb.stateChanged.connect(self.sync_header_check_state)

        layout.addWidget(cb)
        return wrap

    def load_rows(
            self,
            rows: list[dict[str, Any]],
            display_columns: list[str],
            header_labels: list[str],
            width_map: Optional[dict[str, int]] = None,
    ) -> None:
        self.blockSignals(True)
        try:
            self.clear()
            self.setRowCount(len(rows))
            self.setColumnCount(2 + len(display_columns))
            self.setHorizontalHeaderLabels(["", "번호", *header_labels])
            self.row_meta = rows

            self.setColumnWidth(0, 42)
            self.setColumnWidth(1, 56)

            for row_idx, row in enumerate(rows):
                self.setCellWidget(row_idx, 0, self.make_checkbox_cell())

                no_item = QTableWidgetItem(str(row_idx + 1))
                no_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.setItem(row_idx, 1, no_item)

                for col_idx, key in enumerate(display_columns, start=2):
                    value = row.get(key)
                    text = "" if value is None else str(value)
                    item = QTableWidgetItem(text)
                    item.setToolTip(text)
                    if len(text) <= 12:
                        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.setItem(row_idx, col_idx, item)

            self.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
            self.horizontalHeader().setMinimumSectionSize(40)
            self.horizontalHeader().setStretchLastSection(False)

            self.setColumnWidth(0, 42)
            self.setColumnWidth(1, 56)

            if width_map:
                for idx, key in enumerate(display_columns, start=2):
                    self.setColumnWidth(idx, width_map.get(key, 120))

            header = self.horizontalHeader()
            if isinstance(header, CheckBoxHeader):
                header.set_checked(False)

        finally:
            self.blockSignals(False)

    def mousePressEvent(self, event) -> None:
        item = self.itemAt(event.pos())
        if item is not None and item.column() != 0:
            self.selectRow(item.row())
            self.row_clicked_signal.emit(item.row())
        super().mousePressEvent(event)

    def get_checkbox(self, row: int) -> Optional[QCheckBox]:
        wrap = self.cellWidget(row, 0)
        if not wrap:
            return None
        return wrap.findChild(QCheckBox)

    def toggle_all_checked(self, checked: bool) -> None:
        for row in range(self.rowCount()):
            cb = self.get_checkbox(row)
            if cb:
                cb.blockSignals(True)
                cb.setChecked(checked)
                cb.blockSignals(False)

        header = self.horizontalHeader()
        if isinstance(header, CheckBoxHeader):
            header.set_checked(checked)

    def checked_rows(self) -> list[int]:
        result = []
        for row in range(self.rowCount()):
            cb = self.get_checkbox(row)
            if cb and cb.isChecked():
                result.append(row)
        return result

    def sync_header_check_state(self) -> None:
        checked = self.rowCount() > 0 and len(self.checked_rows()) == self.rowCount()
        header = self.horizontalHeader()
        if isinstance(header, CheckBoxHeader):
            header.set_checked(checked)


class DbSetPop(QDialog):
    log_signal = Signal(str)

    def __init__(
            self,
            parent: Optional[QWidget] = None,
            title: str = "DB목록",
            config_path: Optional[str] = None,
            config_data: Optional[dict[str, Any]] = None,
    ) -> None:
        super().__init__(parent)

        self._parent = parent
        self.config_path = config_path or self._resolve_site_config_path()
        self.config_data = config_data or self._load_config_data()

        self.db_name = self._safe_table_name(self.config_data.get("db_name") or "")
        self.db_common_name = self._safe_table_name(self.config_data.get("db_common_name") or "worker_job_hist")
        self.db_path = self._resolve_db_path()

        self.hist_rows: list[dict[str, Any]] = []
        self.detail_rows: list[dict[str, Any]] = []
        self.current_hist_row: Optional[dict[str, Any]] = None

        self.setWindowTitle(title)
        self.resize(1650, 860)
        self.setMinimumSize(1500, 760)
        self.setStyleSheet("background-color: white; color: #111;")

        self.left_header_map = {
            "hist_id": "이력ID",
            "user_id": "사용자ID",
            "start_at": "시작시간",
            "end_at": "종료시간",
            "status": "상태",
            "success_count": "성공건수",
            "fail_count": "실패건수",
            "error_message": "오류메시지",
            "memo": "메모",
        }
        self.left_columns = list(self.left_header_map.keys())
        self.left_width_map = {
            "hist_id": 80,
            "user_id": 110,
            "start_at": 150,
            "end_at": 150,
            "status": 80,
            "success_count": 90,
            "fail_count": 90,
            "error_message": 220,
            "memo": 200,
        }

        self.right_columns, self.right_header_map = self._build_detail_columns()
        self.right_width_map = {
            "keyword": 110,
            "crawled_at": 150,
            "product_name": 260,
            "category": 170,
            "product_no": 120,
            "list_price": 100,
            "low_price": 100,
            "sale_price": 100,
            "delivery_fee": 100,
            "discount_ratio": 90,
            "brand": 100,
            "review_count": 90,
            "purchase_count": 90,
            "wish_count": 90,
            "store_name": 130,
            "mall_prod_mbl_url": 230,
            "mall_product_url": 230,
            "pc_url": 230,
            "total_visit_count": 110,
            "page": 70,
            "no": 70,
        }

        self.left_count_label: Optional[QLabel] = None
        self.right_count_label: Optional[QLabel] = None
        self.left_table: Optional[DbTableWidget] = None
        self.right_table: Optional[DbTableWidget] = None

        self.init_ui()
        self.load_hist_rows()

    def init_ui(self) -> None:
        root_layout = QVBoxLayout(self)
        root_layout.setContentsMargins(12, 8, 12, 12)
        root_layout.setSpacing(6)

        title_label = QLabel(self.windowTitle(), self)
        title_label.setFixedHeight(30)
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title_label.setStyleSheet("font-size: 18px; font-weight: bold; color: #111; padding: 0;")
        root_layout.addWidget(title_label, 0)

        splitter = QSplitter(Qt.Orientation.Horizontal, self)
        splitter.setHandleWidth(8)
        splitter.setChildrenCollapsible(False)
        splitter.addWidget(self.build_left_panel())
        splitter.addWidget(self.build_right_panel())
        splitter.setStretchFactor(0, 4)
        splitter.setStretchFactor(1, 7)
        splitter.setSizes([520, 1000])
        root_layout.addWidget(splitter, 1)

        btn_layout = QHBoxLayout()
        btn_layout.setContentsMargins(0, 4, 0, 0)

        confirm_btn = create_common_button("확인", self.accept, "black", 140)

        btn_layout.addStretch()
        btn_layout.addWidget(confirm_btn)
        root_layout.addLayout(btn_layout, 0)

    def build_left_panel(self) -> QWidget:
        wrap = QWidget()
        layout = QVBoxLayout(wrap)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(8)

        title = QLabel("작업목록")
        title.setFixedHeight(30)
        title.setStyleSheet("font-size: 17px; font-weight: bold; color: #111; padding: 0;")
        layout.addWidget(title)

        top_row = QHBoxLayout()
        top_row.setContentsMargins(0, 0, 0, 0)

        delete_btn = QPushButton("삭제")
        delete_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        delete_btn.setFixedSize(90, 34)
        delete_btn.setStyleSheet(self.gray_button_style())
        delete_btn.clicked.connect(self.delete_left_checked)

        self.left_count_label = QLabel("전체 row수 0")
        self.left_count_label.setStyleSheet("font-size: 13px; font-weight: bold; color: #333;")

        top_row.addWidget(delete_btn)
        top_row.addStretch()
        top_row.addWidget(self.left_count_label)
        layout.addLayout(top_row)

        self.left_table = DbTableWidget(self)
        self.left_table.row_clicked_signal.connect(self.on_left_row_clicked)
        layout.addWidget(self.left_table, 1)

        return wrap

    def build_right_panel(self) -> QWidget:
        wrap = QWidget()
        layout = QVBoxLayout(wrap)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(8)

        title = QLabel("상세목록")
        title.setFixedHeight(30)
        title.setStyleSheet("font-size: 17px; font-weight: bold; color: #111; padding: 0;")
        layout.addWidget(title)

        top_row = QHBoxLayout()
        top_row.setContentsMargins(0, 0, 0, 0)
        top_row.setSpacing(8)

        delete_btn = QPushButton("삭제")
        delete_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        delete_btn.setFixedSize(90, 34)
        delete_btn.setStyleSheet(self.gray_button_style())
        delete_btn.clicked.connect(self.delete_right_checked)

        excel_btn = QPushButton("엑셀 저장")
        excel_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        excel_btn.setFixedSize(110, 34)
        excel_btn.setStyleSheet(self.black_button_style())
        excel_btn.clicked.connect(self.save_detail_to_excel)

        self.right_count_label = QLabel("전체 row수 0")
        self.right_count_label.setStyleSheet("font-size: 13px; font-weight: bold; color: #333;")

        top_row.addWidget(delete_btn)
        top_row.addWidget(excel_btn)
        top_row.addStretch()
        top_row.addWidget(self.right_count_label)
        layout.addLayout(top_row)

        self.right_table = DbTableWidget(self)
        layout.addWidget(self.right_table, 1)

        return wrap

    def gray_button_style(self) -> str:
        return """
        QPushButton {
            background: #efefef;
            color: #111;
            border: none;
            border-radius: 8px;
            font-size: 13px;
            font-weight: bold;
        }
        QPushButton:hover { background: #e4e4e4; }
        """

    def black_button_style(self) -> str:
        return """
        QPushButton {
            background: black;
            color: white;
            border: none;
            border-radius: 8px;
            font-size: 13px;
            font-weight: bold;
        }
        QPushButton:hover { background: #222; }
        """

    def _safe_table_name(self, name: str) -> str:
        name = str(name or "").strip()
        if not name:
            raise ValueError("테이블명이 없습니다.")
        if not re.match(r"^[A-Za-z_][A-Za-z0-9_]*$", name):
            raise ValueError(f"잘못된 테이블명입니다: {name}")
        return name

    def _resolve_site_config_path(self) -> Optional[str]:
        try:
            state = GlobalState()
            app_config = state.get(GlobalState.APP_CONFIG) or {}
            site = str(state.get(GlobalState.SITE) or "").strip()
            runtime_dir = str(app_config.get("runtime_dir") or "").strip()

            if not runtime_dir or not site:
                return None

            app_json_path = os.path.join(runtime_dir, "app.json")
            if not os.path.exists(app_json_path):
                return None

            with open(app_json_path, "r", encoding="utf-8") as f:
                app_json = json.load(f)

            for site_item in app_json.get("site_list") or []:
                if isinstance(site_item, dict) and str(site_item.get("key") or "").strip() == site:
                    config_rel_path = str(site_item.get("config_path") or "").strip()
                    return os.path.normpath(os.path.join(runtime_dir, *config_rel_path.split("/")))

            return None
        except Exception:
            return None

    def _load_config_data(self) -> dict[str, Any]:
        if self._parent is not None:
            parent_config = getattr(self._parent, "config_data", None)
            if isinstance(parent_config, dict):
                return parent_config

        if self.config_path and os.path.exists(self.config_path):
            with open(self.config_path, "r", encoding="utf-8") as f:
                return json.load(f)

        return {}

    def _resolve_db_path(self) -> str:
        if self.config_path:
            runtime_dir = os.path.dirname(os.path.dirname(os.path.dirname(self.config_path)))
            return os.path.join(runtime_dir, "customers", "common", "db", "worker_hist.db")

        try:
            state = GlobalState()
            app_config = state.get(GlobalState.APP_CONFIG) or {}
            runtime_dir = str(app_config.get("runtime_dir") or "").strip()
            if runtime_dir:
                return os.path.join(runtime_dir, "customers", "common", "db", "worker_hist.db")
        except Exception:
            pass

        # === 신규 === 빌드 후에는 현재 작업경로(os.getcwd)가 아니라 exe 위치 기준
        if getattr(sys, "frozen", False):
            base_dir = os.path.dirname(sys.executable)
        else:
            base_dir = os.getcwd()

        return os.path.join(base_dir, "runtime", "customers", "common", "db", "worker_hist.db")

    def _get_folder_path(self) -> str:
        for row in self.config_data.get("setting") or []:
            if isinstance(row, dict) and str(row.get("code") or "").strip() == "folder_path":
                value = str(row.get("value") or "").strip()
                if value:
                    return value
        return os.path.dirname(self.db_path)

    def _build_detail_columns(self) -> tuple[list[str], dict[str, str]]:
        code_list = []
        header_map: dict[str, str] = {}

        for row in self.config_data.get("columns") or []:
            if not isinstance(row, dict):
                continue

            code = str(row.get("code") or "").strip()
            name = str(row.get("value") or code).strip()

            if code:
                code_list.append(code)
                header_map[code] = name

        return code_list, header_map

    def _connect(self) -> sqlite3.Connection:
        # === 신규 === DB 없을 때 빈 DB가 자동 생성되는 것 방지
        if not os.path.exists(self.db_path):
            raise FileNotFoundError(f"DB 파일이 없습니다: {self.db_path}")

        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        return conn

    def load_hist_rows(self) -> None:
        if not self.left_table or not self.left_count_label:
            return

        self.current_hist_row = None
        self.detail_rows = []
        rows: list[dict[str, Any]] = []

        try:
            with self._connect() as conn:
                cur = conn.execute(
                    f"""
                    SELECT
                        hist_id,
                        job_id,
                        user_id,
                        start_at,
                        end_at,
                        status,
                        success_count,
                        fail_count,
                        error_message,
                        memo
                    FROM {self.db_common_name}
                    WHERE UPPER(table_name) = UPPER(?)
                    ORDER BY hist_id DESC
                    """,
                    (self.db_name,),
                )
                rows = [dict(row) for row in cur.fetchall()]
        except Exception as e:
            QMessageBox.warning(self, "오류", f"작업목록 조회 실패\n{e}")

        self.hist_rows = rows
        self.left_table.load_rows(
            rows,
            self.left_columns,
            [self.left_header_map[x] for x in self.left_columns],
            self.left_width_map,
        )
        self.left_count_label.setText(f"전체 row수 {len(rows)}")
        self.clear_detail_rows()

    def clear_detail_rows(self) -> None:
        self.current_hist_row = None
        self.detail_rows = []

        if self.right_table:
            self.right_table.load_rows(
                [],
                self.right_columns,
                [self.right_header_map.get(x, x) for x in self.right_columns],
                self.right_width_map,
            )

        if self.right_count_label:
            self.right_count_label.setText("전체 row수 0")

    def on_left_row_clicked(self, row_index: int) -> None:
        if row_index < 0 or row_index >= len(self.hist_rows):
            return

        self.current_hist_row = self.hist_rows[row_index]
        self.load_detail_rows_by_job_id(str(self.current_hist_row.get("job_id") or ""))

    def load_detail_rows_by_job_id(self, job_id: str) -> None:
        if not self.right_table or not self.right_count_label:
            return

        rows: list[dict[str, Any]] = []

        try:
            with self._connect() as conn:
                cur = conn.execute(
                    f"SELECT rowid AS __rowid__, * FROM {self.db_name} WHERE job_id = ? ORDER BY rowid DESC",
                    (job_id,),
                )
                rows = [dict(row) for row in cur.fetchall()]
        except Exception as e:
            QMessageBox.warning(self, "오류", f"상세목록 조회 실패\n{e}")

        self.detail_rows = rows
        self.right_table.load_rows(
            rows,
            self.right_columns,
            [self.right_header_map.get(x, x) for x in self.right_columns],
            self.right_width_map,
        )
        self.right_count_label.setText(f"전체 row수 {len(rows)}")

    def delete_left_checked(self) -> None:
        if not self.left_table:
            return

        checked = self.left_table.checked_rows()
        if not checked:
            QMessageBox.information(self, "알림", "삭제할 항목을 선택해주세요.")
            return

        ok = QMessageBox.question(
            self,
            "삭제 확인",
            "삭제 하겠습니까?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
            )
        if ok != QMessageBox.StandardButton.Yes:
            return

        targets = [self.hist_rows[i] for i in checked if i < len(self.hist_rows)]

        try:
            with self._connect() as conn:
                for row in targets:
                    hist_id = row.get("hist_id")
                    job_id = row.get("job_id")

                    conn.execute(f"DELETE FROM {self.db_name} WHERE hist_id = ? OR job_id = ?", (hist_id, job_id))
                    conn.execute(f"DELETE FROM {self.db_common_name} WHERE hist_id = ?", (hist_id,))

            QMessageBox.information(self, "알림", "삭제되었습니다.")
            self.load_hist_rows()

        except Exception as e:
            QMessageBox.warning(self, "오류", f"삭제 실패\n{e}")

    def delete_right_checked(self) -> None:
        if not self.right_table:
            return

        checked = self.right_table.checked_rows()
        if not checked:
            QMessageBox.information(self, "알림", "삭제할 항목을 선택해주세요.")
            return

        ok = QMessageBox.question(
            self,
            "삭제 확인",
            "삭제 하겠습니까?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
            )
        if ok != QMessageBox.StandardButton.Yes:
            return

        targets = [self.detail_rows[i] for i in checked if i < len(self.detail_rows)]

        try:
            with self._connect() as conn:
                for row in targets:
                    conn.execute(f"DELETE FROM {self.db_name} WHERE rowid = ?", (row.get("__rowid__"),))

            QMessageBox.information(self, "알림", "삭제되었습니다.")

            if self.current_hist_row:
                self._refresh_left_summary_counts()
            else:
                self.clear_detail_rows()

        except Exception as e:
            QMessageBox.warning(self, "오류", f"삭제 실패\n{e}")

    def _refresh_left_summary_counts(self) -> None:
        if not self.current_hist_row:
            self.load_hist_rows()
            return

        hist_id = self.current_hist_row.get("hist_id")
        job_id = str(self.current_hist_row.get("job_id") or "")

        try:
            with self._connect() as conn:
                success_count = conn.execute(
                    f"""
                    SELECT COUNT(*) AS cnt
                    FROM {self.db_name}
                    WHERE hist_id = ?
                      AND UPPER(COALESCE(row_status, '')) = 'SUCCESS'
                    """,
                    (hist_id,),
                ).fetchone()["cnt"]

                fail_count = conn.execute(
                    f"""
                    SELECT COUNT(*) AS cnt
                    FROM {self.db_name}
                    WHERE hist_id = ?
                      AND UPPER(COALESCE(row_status, '')) = 'FAIL'
                    """,
                    (hist_id,),
                ).fetchone()["cnt"]

                conn.execute(
                    f"""
                    UPDATE {self.db_common_name}
                    SET total_count = ?, success_count = ?, fail_count = ?, updated_at = ?
                    WHERE hist_id = ?
                    """,
                    (
                        int(success_count) + int(fail_count),
                        int(success_count),
                        int(fail_count),
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        hist_id,
                    ),
                )
        except Exception:
            pass

        selected_hist_id = hist_id
        self.load_hist_rows()

        if not self.left_table:
            return

        for row_index, row in enumerate(self.hist_rows):
            if row.get("hist_id") == selected_hist_id:
                self.left_table.selectRow(row_index)
                self.current_hist_row = row
                self.load_detail_rows_by_job_id(job_id)
                break

    def save_detail_to_excel(self) -> None:
        if not self.detail_rows:
            QMessageBox.information(self, "알림", "저장할게 없습니다.")
            return

        folder_root = self._get_folder_path()
        save_dir = os.path.join(folder_root, "output_db")
        os.makedirs(save_dir, exist_ok=True)

        job_id = ""
        if self.current_hist_row:
            job_id = str(self.current_hist_row.get("job_id") or "").strip()

        filename = f"{self.db_name.lower()}_{job_id or 'detail'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path = os.path.join(save_dir, filename)

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "상세목록"

            headers = [self.right_header_map.get(x, x) for x in self.right_columns]
            ws.append(headers)

            for row in self.detail_rows:
                ws.append([row.get(col, "") for col in self.right_columns])

            header_fill = PatternFill(fill_type="solid", fgColor="F3F3F3")
            header_font = Font(bold=True)
            thin = Side(style="thin", color="DDDDDD")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(vertical="center")
                    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for idx, col in enumerate(self.right_columns, start=1):
                width = self.right_width_map.get(col, 120)
                ws.column_dimensions[get_column_letter(idx)].width = max(12, int(width / 8))

            wb.save(file_path)
            QMessageBox.information(self, "알림", "엑셀이 저장되었습니다.")

        except Exception as e:
            QMessageBox.warning(self, "오류", f"엑셀 저장 실패\n{e}")