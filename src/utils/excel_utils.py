import pandas as pd
import os
from openpyxl import load_workbook, Workbook
import csv
import re
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

class ExcelUtils:
    def __init__(self, log_func=None):
        self.log_func = log_func

    # =========================
    # 경로 유틸
    # =========================
    def get_default_output_dir(self):
        return os.path.join(os.path.expanduser("~"), "Documents")

    def resolve_output_dir(self, folder_path=None):
        folder_path = str(folder_path or "").strip()

        if folder_path:
            output_dir = folder_path
        else:
            output_dir = self.get_default_output_dir()

        os.makedirs(output_dir, exist_ok=True)

        if self.log_func:
            self.log_func(f"[EXCEL] 저장 폴더: {output_dir}")

        return output_dir

    def build_file_path(self, filename, folder_path=None, sub_dir=None):
        output_dir = self.resolve_output_dir(folder_path)

        if sub_dir:
            output_dir = os.path.join(output_dir, sub_dir)

        os.makedirs(output_dir, exist_ok=True)

        filename = str(filename or "").strip()
        if not filename:
            raise ValueError("filename 이 비어 있습니다.")

        filename = os.path.basename(filename)

        full_path = os.path.join(output_dir, filename)

        if self.log_func:
            self.log_func(f"[EXCEL] 저장 파일 경로: {full_path}")

        return full_path

    def init_csv(self, filename, columns, folder_path=None, sub_dir=None):
        filename = self.build_file_path(filename, folder_path, sub_dir)

        df = pd.DataFrame(columns=columns)
        df.to_csv(filename, index=False, encoding="utf-8-sig")
        if self.log_func:
            self.log_func(f"CSV 초기화 완료: {filename}")

    def append_to_csv(self, filename, data_list, columns, folder_path=None, sub_dir=None):
        if not data_list:
            return

        filename = self.build_file_path(filename, folder_path, sub_dir)

        df = pd.DataFrame(data_list, columns=columns)
        df.to_csv(filename, mode="a", header=False, index=False, encoding="utf-8-sig")
        data_list.clear()
        if self.log_func:
            self.log_func("csv 저장완료")

    def append_to_excel(self, filename, data_list, columns, sheet_name="Sheet1", folder_path=None, sub_dir=None):
        if not data_list:
            return

        filename = self.build_file_path(filename, folder_path, sub_dir)

        df = pd.DataFrame(data_list, columns=columns)

        if os.path.exists(filename):
            with pd.ExcelWriter(filename, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                start_row = writer.sheets[sheet_name].max_row if sheet_name in writer.sheets else 0
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=start_row)
        else:
            with pd.ExcelWriter(filename, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=True)

        data_list.clear()
        if self.log_func:
            self.log_func("excel 저장완료")

    def _clean_excel_cell_value(self, value):
        if pd.isna(value):
            return ""

        text = str(value).strip()

        # 엑셀 저장 불가 제어문자 제거
        # 허용: \t, \n, \r
        text = re.sub(r"[\x00-\x08\x0B-\x0C\x0E-\x1F\x7F]", "", text)

        return text


    def _apply_header_style_and_filter(self, ws):
        max_col = ws.max_column
        max_row = ws.max_row

        if max_col <= 0 or max_row <= 0:
            return

        header_fill = PatternFill(fill_type="solid", fgColor="BFBFBF")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"




    def _get_header_index_map(self, ws):
        header_map = {}

        for col_idx in range(1, ws.max_column + 1):
            header_text = str(ws.cell(row=1, column=col_idx).value or "").strip()
            if header_text:
                header_map[header_text] = col_idx

        return header_map


    def _build_hyperlink_url(self, url_prefix, value_text):
        value_text = str(value_text or "").strip()
        url_prefix = str(url_prefix or "").strip()

        if not value_text:
            return ""

        if value_text.startswith("http://") or value_text.startswith("https://"):
            return value_text

        return f"{url_prefix}{value_text}"


    def _apply_hyperlink_cells(self, ws, hyperlink_columns=None):
        header_map = self._get_header_index_map(ws)

        for spec in hyperlink_columns or []:
            target_col = str(spec.get("컬럼", "")).strip()
            value_col = str(spec.get("값컬럼", "")).strip()
            display_col = str(spec.get("표시컬럼", "")).strip() or target_col
            url_prefix = str(spec.get("url", "")).strip()

            if not target_col or not value_col:
                continue

            target_idx = header_map.get(target_col)
            value_idx = header_map.get(value_col)
            display_idx = header_map.get(display_col) or target_idx

            if not target_idx or not value_idx:
                continue

            for row_idx in range(2, ws.max_row + 1):
                value_cell = ws.cell(row=row_idx, column=value_idx)
                display_cell = ws.cell(row=row_idx, column=display_idx)
                target_cell = ws.cell(row=row_idx, column=target_idx)

                value_text = self._clean_excel_cell_value(value_cell.value)
                display_text = self._clean_excel_cell_value(display_cell.value)

                if not value_text or not display_text:
                    continue

                link_url = self._build_hyperlink_url(url_prefix, value_text)
                if not link_url:
                    continue

                target_cell.value = display_text
                target_cell.hyperlink = link_url
                target_cell.style = "Hyperlink"

        # 기존 호환: 셀 값 자체가 URL이면 자동 링크
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                if cell.hyperlink:
                    continue

                text = self._clean_excel_cell_value(cell.value)
                if not text:
                    continue

                if text.startswith("http://") or text.startswith("https://"):
                    cell.value = text
                    cell.hyperlink = text
                    cell.style = "Hyperlink"


    def _apply_column_widths(self, ws, column_widths=None, default_width=16):
        width_map = {}

        for item in column_widths or []:
            col_name = str(item.get("컬럼", "")).strip()
            width = item.get("너비", "")

            if not col_name:
                continue

            try:
                width_map[col_name] = float(width)
            except Exception:
                continue

        for col_idx in range(1, ws.max_column + 1):
            header_value = ws.cell(row=1, column=col_idx).value
            header_text = str(header_value or "").strip()

            width = width_map.get(header_text, default_width)
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width


    def convert_csv_to_excel_and_delete(self, csv_filename, sheet_name="Sheet1", folder_path=None, sub_dir=None, keep_csv=False, column_widths=None, default_width=16, hyperlink_columns=None):
        csv_filename = self.build_file_path(csv_filename, folder_path, sub_dir)

        if not os.path.exists(csv_filename):
            if self.log_func:
                self.log_func(f"❌ CSV 파일이 존재하지 않습니다: {csv_filename}")
            return False

        try:
            df = pd.read_csv(csv_filename, encoding="utf-8-sig", dtype=str, keep_default_na=False)

            if df.empty:
                if self.log_func:
                    self.log_func(f"⚠️ CSV에 데이터가 없습니다: {csv_filename}")
                return False

            for col in df.columns:
                df[col] = df[col].apply(self._clean_excel_cell_value)

            excel_filename = os.path.splitext(csv_filename)[0] + ".xlsx"

            with pd.ExcelWriter(excel_filename, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name=sheet_name)

                ws = writer.sheets[sheet_name]
                for r in ws.iter_rows(min_row=2, max_row=len(df) + 1):
                    for cell in r:
                        if cell.value is not None:
                            cell.value = self._clean_excel_cell_value(cell.value)
                self._apply_header_style_and_filter(ws)
                self._apply_hyperlink_cells(ws, hyperlink_columns=hyperlink_columns)
                self._apply_column_widths(ws, column_widths=column_widths, default_width=default_width)

            if not keep_csv:
                os.remove(csv_filename)
                if self.log_func:
                    self.log_func(f"🗑️ CSV 파일 삭제 완료: {csv_filename}")
            else:
                if self.log_func:
                    self.log_func(f"📄 CSV 파일 유지: {csv_filename}")

            if self.log_func:
                self.log_func(f"✅ 엑셀 파일 저장 완료: {excel_filename}")

            return True
        except Exception as e:
            if self.log_func:
                self.log_func(f"❌ 변환 중 오류 발생: {e}")
            return False

    def obj_to_row(self, o, cols):
        if isinstance(o, dict):
            return {c: o.get(c) for c in cols}
        return {c: getattr(o, c, None) for c in cols}

    def obj_list_to_dataframe(self, obj_list, columns=None):
        if not obj_list:
            return None

        if isinstance(obj_list[0], dict):
            if columns:
                rows = [{col: obj.get(col) for col in columns} for obj in obj_list]
                return pd.DataFrame(rows, columns=columns)
            return pd.DataFrame(obj_list)

        if columns:
            rows = [self.obj_to_row(o, columns) for o in obj_list]
            return pd.DataFrame(rows, columns=columns)

        first = obj_list[0]
        if hasattr(first, "__dict__") and first.__dict__:
            cols = list(first.__dict__.keys())
            rows = [self.obj_to_row(o, cols) for o in obj_list]
            return pd.DataFrame(rows, columns=cols)

        cols = [k for k in dir(first) if not k.startswith("_") and not callable(getattr(first, k, None))]
        rows = [self.obj_to_row(o, cols) for o in obj_list]
        return pd.DataFrame(rows, columns=cols)

    def save_obj_list_to_excel(
            self,
            filename,
            obj_list,
            columns=None,
            sheet_name="Sheet1",
            folder_path=None,
            sub_dir=None
    ):

        if not obj_list:
            return None

        filename = self.build_file_path(filename, folder_path, sub_dir)

        df = self.obj_list_to_dataframe(obj_list, columns=columns)
        if df is None or df.empty:
            if self.log_func:
                self.log_func("⚠️ 저장할 데이터가 없습니다.")
            return None

        if os.path.exists(filename):
            with pd.ExcelWriter(filename, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                ws = writer.sheets.get(sheet_name)
                if ws is not None:
                    start_row = ws.max_row if ws.max_row is not None else 0
                    df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=start_row)
                else:
                    df.to_excel(writer, sheet_name=sheet_name, index=False, header=True)
        else:
            with pd.ExcelWriter(filename, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=True)

        wb = load_workbook(filename)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    val = str(cell.value) if cell.value else ""
                    if val.startswith("http://") or val.startswith("https://"):
                        cell.hyperlink = val
                        cell.style = "Hyperlink"

        wb.save(filename)

        obj_list.clear()
        if self.log_func:
            self.log_func("excel(객체 리스트) 저장완료 (URL 하이퍼링크 처리)")

        return filename

    def append_rows_text_excel(
            self,
            filename,
            rows,
            columns,
            sheet_name="Sheet1",
            folder_path=None,
            sub_dir=None
    ):
        if not rows:
            if self.log_func:
                self.log_func("[EXCEL] 저장할 데이터 없음")
            return

        filename = self.build_file_path(filename, folder_path, sub_dir)

        if os.path.exists(filename):
            wb = load_workbook(filename)

            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(title=sheet_name)
                ws.append(columns)

            if self.log_func:
                self.log_func(f"[EXCEL] 기존 파일 로드: {filename}")
                self.log_func(f"[EXCEL] 대상 시트: {sheet_name}")

        else:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            ws.append(columns)
            if self.log_func:
                self.log_func(f"[EXCEL] 신규 파일 생성: {filename}")
                self.log_func(f"[EXCEL] 대상 시트: {sheet_name}")

        saved = 0
        for r in rows:
            out = {}
            for c in columns:
                out[c] = r.get(c, "")

            ws.append([str(out.get(c, "")) for c in columns])

            for col in range(1, len(columns) + 1):
                ws.cell(ws.max_row, col).number_format = "@"

            saved += 1

        wb.save(filename)
        if self.log_func:
            self.log_func(f"[EXCEL] 저장 완료 ({sheet_name}) (추가 {saved}건)")

    def append_row_to_csv(self, csv_filename, item, columns, folder_path=None, sub_dir=None):
        if not columns:
            return

        csv_filename = self.build_file_path(csv_filename, folder_path, sub_dir)

        row = {}
        for c in columns:
            v = item.get(c)
            row[c] = "" if v is None else str(v)

        file_exists = os.path.exists(csv_filename)

        with open(csv_filename, "a", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
            if not file_exists:
                writer.writeheader()
            writer.writerow(row)

        if self.log_func:
            self.log_func("csv 1행 저장완료")

    def close(self):
        if self.log_func:
            self.log_func("✅ [엑셀] 해재")
